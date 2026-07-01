"""
Material Inventory Automation System - v5
Complete rebuild: CO fix, invoice unmatched handling, leftover tracking,
summary sheets, per-project + all-projects excel, UI editing, graphs page,
total cost with tax in top bar.
PostgreSQL-backed storage (db.py) — replaces all JSON file I/O.
"""
import os, json, re, shutil
from pathlib import Path
from datetime import datetime

# ── API Key ───────────────────────────────────────────────────────────────────
import os

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
os.environ["ANTHROPIC_API_KEY"] = ANTHROPIC_API_KEY

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import anthropic

# Schedule / BT estimate dependencies
try:
    import pandas as pd
    import xlrd
    PANDAS_OK = True
except ImportError:
    PANDAS_OK = False
    pd = None

BASE_DIR   = Path(__file__).parent
# PROJECTS dir kept for Excel/PDF temp files only (generated on-demand, not stored)
PROJECTS   = BASE_DIR / "projects"
import tempfile as _tmp_s
UPLOAD_DIR = Path(_tmp_s.gettempdir()) / "matinv_uploads"
PROJECTS.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(exist_ok=True)

# ── Postgres storage layer ────────────────────────────────────────────────────
import db as _db
_db.init_db()   # creates tables if they don't exist yet (idempotent)

app = FastAPI(title="Material Inventory Automation v5")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

_claude = None
def get_claude():
    global _claude
    if _claude is None:
        # Pass api_key explicitly so it works regardless of env var state
        _claude = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    return _claude

# ── Styles ────────────────────────────────────────────────────────────────────
THIN        = Side(border_style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", start_color="1B3A5C")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
TYPE_FILLS  = {
    "Lumber":    PatternFill("solid", start_color="DDEEFF"),
    "Panels":    PatternFill("solid", start_color="FFF3CD"),
    "LVL":       PatternFill("solid", start_color="D4EDDA"),
    "Each":      PatternFill("solid", start_color="F8D7DA"),
    "Siding":    PatternFill("solid", start_color="E8F5E9"),
    "HouseWrap": PatternFill("solid", start_color="FCE4EC"),
    "Unknown":   PatternFill("solid", start_color="E2E2E2"),
}
SUBTOTAL_FILLS = {
    "Lumber":    PatternFill("solid", start_color="AACCEE"),
    "Panels":    PatternFill("solid", start_color="FFD966"),
    "LVL":       PatternFill("solid", start_color="93C47D"),
    "Each":      PatternFill("solid", start_color="E06666"),
    "Siding":    PatternFill("solid", start_color="A5D6A7"),
    "HouseWrap": PatternFill("solid", start_color="F48FB1"),
    "Unknown":   PatternFill("solid", start_color="BBBBBB"),
}
GRAND_FILL  = PatternFill("solid", start_color="1B3A5C")
GRAND_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
SUBTOT_FONT = Font(bold=True, name="Arial", size=10)
NORMAL_FONT = Font(name="Arial", size=9)
MONEY_FMT   = '"$"#,##0.00'
INT_FMT     = '#,##0'
DEC_FMT     = '#,##0.00'
TAX_RATE    = 0.06

# ── Math helpers ──────────────────────────────────────────────────────────────
def n(v):
    try: return float(v or 0)
    except: return 0.0

def compute_lf(typ, qty, t, w, length):
    """Compute linear feet per formula table:
    Lumber: LF = Qty × Length  (qty=pieces, length=ft)
    LVL:    LF = Qty × Length  (qty=pieces, length=ft; if length=0, qty is already LF)
    Others: LF = 0
    """
    q, ln = n(qty), n(length)
    if typ == "LVL":
        # If length given: qty = pieces, LF = pieces × length
        # If length = 0: qty was stored as LF directly (e.g. from CO with uom=LF)
        return round(q * ln, 4) if ln else round(q, 4)
    if typ == "Lumber":
        return round(q * ln, 4)
    return 0.0

def compute_bf_sf(typ, qty, t, w, length):
    q, tv, wv, ln = n(qty), n(t), n(w), n(length)
    if typ == "Lumber":
        return round((q * tv * wv * ln) / 12, 4)
    if typ == "Panels":
        # Each panel sheet is 4ft x 8ft = 32 SF.
        # t_num is the thickness (23/32 etc.) used only for identification.
        # w_num=8 represents the 8-ft dimension; the 4-ft is always implied.
        # SF per sheet = 4 x w_num (default w=8 => 32 SF/sheet)
        sheet_sf = 4.0 * (wv if wv > 0 else 8.0)
        return round(q * sheet_sf, 4)
    return 0.0

def compute_cost(typ, qty, t, w, length, unit_cost, cost_formula=None):
    """Compute cost with optional formula override.
    cost_formula: None=auto, 'each'/'pc'=qty*uc, 'mbf'/'msf'=BF*uc/1000, 'lf'=LF*uc
    """
    lf   = compute_lf(typ, qty, t, w, length)
    bfsf = compute_bf_sf(typ, qty, t, w, length)
    uc   = n(unit_cost)
    q    = n(qty)
    # Apply explicit formula override if set
    if cost_formula:
        cf = str(cost_formula).lower().strip()
        if cf in ("each", "pc", "per_pc", "per_piece"):
            return round(q * uc, 2)
        if cf in ("mbf", "msf", "per_mbf", "per_msf", "bf", "sf"):
            return round((bfsf * uc) / 1000, 2)
        if cf in ("lf", "per_lf", "linear"):
            return round(lf * uc, 2)
    # Auto formula based on type
    if typ == "Lumber":                          return round((bfsf * uc) / 1000, 2)
    if typ == "Panels":                          return round((bfsf * uc) / 1000, 2)
    if typ == "LVL":
        # LVL is ALWAYS priced per LF. lf = compute_lf result.
        return round(lf * uc, 2)
    if typ in ("Each", "Siding", "HouseWrap"):   return round(q * uc, 2)
    return 0.0

def item_fields(item, qty=None):
    typ = item.get("type", "")
    t   = n(item.get("t_num", 0))
    w   = n(item.get("w_num", 0))
    ln  = n(item.get("length_num", 0))
    uc  = n(item.get("unit_cost", 0))
    cf  = item.get("cost_formula", None)   # formula override
    if qty is None:
        qty = n(item.get("po_qty", 0)) + n(item.get("co_qty", 0))
    lf   = compute_lf(typ, qty, t, w, ln)
    bfsf = compute_bf_sf(typ, qty, t, w, ln)
    cost = compute_cost(typ, qty, t, w, ln, uc, cf)
    return lf, bfsf, cost


# ── PDF helpers ───────────────────────────────────────────────────────────────
def pdf_to_text(pdf_path):
    """Extract text from PDF. For structured tables, also try table extraction."""
    parts = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for pg_num, page in enumerate(pdf.pages, 1):
            parts.append(f"--- PAGE {pg_num} ---")
            # Try table extraction first (better for Matheus CO format)
            tables = page.extract_tables()
            if tables:
                for tbl in tables:
                    for row in tbl:
                        if row:
                            clean = [str(c or "").strip() for c in row]
                            if any(c for c in clean):
                                parts.append("  |  ".join(clean))
            # Always also get raw text (catches headers, totals)
            t = page.extract_text()
            if t:
                parts.append(t)
    return "\n".join(parts)

def safe_json_parse(raw):
    raw = re.sub(r'^```(?:json)?', '', raw.strip()).rstrip('`').strip()
    try: return json.loads(raw)
    except:
        lb = raw.rfind('}')
        if lb != -1:
            for suffix in [']', '']:
                try: return json.loads(raw[:lb+1] + suffix)
                except: pass
    return []


# ── Claude parsers ────────────────────────────────────────────────────────────
def parse_po_with_claude(text, po_category="lumber"):
    """Parse a PO PDF. po_category controls how items are typed:
    'lumber'   → normal typing (Lumber/LVL/Panels/Each)
    'housewrap'→ ALL items → HouseWrap (cost = qty × unit_cost)
    'siding'   → ALL items → Siding   (cost = qty × unit_cost)
    """
    # Category-specific instructions
    if po_category == "housewrap":
        type_rule = (
            'TYPE RULE: This is a HOUSE WRAP PO. Set "type": "HouseWrap" for EVERY item. '
            'No exceptions. These are wrap/tape/flashing products sold per piece or roll.'
        )
        forced_type = "HouseWrap"
    elif po_category == "siding":
        type_rule = (
            'TYPE RULE: This is a SIDING PO. Set "type": "Siding" for EVERY item. '
            'No exceptions. These are siding/trim products.'
        )
        forced_type = "Siding"
    else:
        type_rule = (
            "TYPE RULE: Classify each row by material: "
            "Lumber=dimensional lumber sold per MBF. "
            "LVL=laminated veneer/PSL/GLB sold per LF. "
            "Panels=OSB/Plywood/ZIP sold per MSF. "
            "Each=adhesive/sill seal/tape/hardware sold per piece. "
            "Siding=siding products. HouseWrap=Tyvek/wrap."
        )
        forced_type = None

    prompt = (
        "You are parsing a Matheus Lumber JOB PROPOSAL (Purchase Order).\n"
        "Table columns vary by PO format but always contain: Qty | T | W | Length | Description | LF | BF/SF | Unit Cost | Total.\n"
        "Extract EVERY single data row — do NOT skip any. Return ONLY a valid JSON array.\n\n"
        + type_rule + "\n\n"
        "━━━ QUANTITY (qty field) ━━━\n"
        "- Read the first number column (Amount/Qty/Ordered). Remove commas. Exact integer or decimal.\n"
        "- This is ALWAYS the piece/roll/unit count for all types.\n\n"
        "━━━ DIMENSIONS ━━━\n"
        "Lumber: t_num=T column (e.g. 2), w_num=W column (e.g. 6), length_num=Length column (12→12, R/L or 1→1).\n"
        "  CRITICAL: Each unique T×W×Length combo = separate row. 2x6x12 and 2x6x10 are DIFFERENT rows.\n"
        "LVL/PSL/GLB/LSL: t_num=first dim (1-3/4→1.75, 3-1/2→3.5), w_num=second dim (9-1/4→9.25, 11-7/8→11.875, 14→14).\n"
        "  length_num=length in feet (12'→12, 10'→10, 26'→26, R/L or 1→1).\n"
        "  CRITICAL: Each unique T×W×Length combo = separate row.\n"
        "Panels: t_num=thickness fraction (7/16→0.4375, 15/32→0.46875, 1/2→0.5, 19/32→0.59375, 5/8→0.625, 23/32→0.71875).\n"
        "  w_num=8 (standard 4x8 sheet). length_num=0.\n"
        "Each/Siding/HouseWrap: t_num=0, w_num=0, length_num=0.\n\n"
        "━━━ UNIT COST (unit_cost field) ━━━\n"
        "Read the Cost/Price column. Strip $, /MBF, /MSF, /LF, /EA, commas.\n"
        "  '$635.00/MBF'→635.0,  '6.85/LF'→6.85,  '$7.250/EA'→7.25,  '695.00/MSF'→695.0\n"
        "This is the raw unit rate — NOT the total amount.\n\n"
        "━━━ IMPORTANT ━━━\n"
        "- Include ALL rows from ALL pages. A large PO may have 50-100+ rows.\n"
        "- Each row with a different qty, dimensions, or unit_cost = separate JSON object.\n"
        "- Do NOT merge rows. Do NOT skip rows.\n"
        "- Ignore: headers, Sub Total, Tax, Grand Total, signature blocks, payment terms.\n"
        "- Numbers only: no $, no commas in output.\n\n"
        "Return ONLY the JSON array, no explanation. Each object:\n"
        '{"type":string,"description":string,"t_num":number,"w_num":number,"length_num":number,"qty":number,"unit_cost":number}\n\n'
        "PO TEXT:\n"
    ) + text + "\n\nReturn ONLY the JSON array. Include ALL rows."

    msg = get_claude().messages.create(
        model="claude-sonnet-4-5", max_tokens=16000,
        messages=[{"role": "user", "content": prompt}])
    items = safe_json_parse(next((getattr(b,"text","") for b in msg.content if hasattr(b,"text")),"").strip())

    # If category forces a type, override whatever Claude chose
    if forced_type:
        for item in items:
            item["type"] = forced_type
            item["t_num"] = 0; item["w_num"] = 0; item["length_num"] = 0

    return items


def normalize_panel_desc(desc):
    """Normalize PT/MCA and strip CO# references from panel descriptions."""
    import re
    d = desc.upper().strip()
    # Remove CO reference lines like "CO #001" or "CO#001"
    d = re.sub(r'\bCO\s*#?\s*\d+\b', '', d).strip()
    # Treat MCA TREATED same as PT
    d = d.replace("MCA TREATED", "PT").replace("MCA", "PT")
    # Collapse multiple spaces
    d = re.sub(r'\s+', ' ', d).strip()
    return d


def parse_co_with_claude(text, po_items):
    """
    Parse CO PDF and return clean line items.
    Rules applied here (before matching):
      - Strip 'CO #001' / 'CO #008' reference text from descriptions
      - Treat MCA TREATED == PT
      - Panels: t_num = thickness fraction (e.g. 23/32=0.71875), w_num=8
      - PRICE column = unit_cost exactly as printed (per MSF or per MBF)
    """
    prompt = (
        "You are parsing a Matheus Lumber ORDER ADD / CHANGE document (Change Order / CO).\n"
        "This document has 5 columns: ORDERED | DESCRIPTION | FOOTAGE | PRICE | AMOUNT\n\n"
        "ORDERED COLUMN: qty with unit embedded in the number. Examples:\n"
        "  129LF → qty_change=129 uom=LF   5111LF → 5111   7686PC → 7686\n"
        "  -14PC → -14   -14041LF → -14041   940EA → 940   -329EA → -329\n"
        "CRITICAL: Remove commas from numbers. 5,111LF=5111. -27,387LF=-27387.\n\n"
        "TYPES: Lumber=SYP/MCA/GLB sold/MBF. LVL=anything with LVL. "
        "Panels=OSB/PLYWOOD/ZIP/GYP-SHEATHING sold/MSF. "
        "Each=SILL SEALER/ADHESIVE/TYVEK/TREX/TAPE/FLASHING sold/EA or PC.\n\n"
        "DESCRIPTION: Remove MCA TREATED→PT, MCA→PT, /BTR→remove. "
        "Skip CO #001 / CO #003 reference lines.\n\n"
        "DIMENSIONS lumber: 2X6X12→t=2,w=6,len=12. R/L or XR/L→len=1.\n"
        "DIMENSIONS panels: t=fractional (7/16=0.4375,15/32=0.46875,1/2=0.5,"
        "19/32=0.59375,23/32=0.71875,5/8=0.625), w=8, len=0. NEVER t=4.\n"
        "DIMENSIONS LVL: t=1.75 for 1-3/4, w=second dim, len=span or 1 for R/L.\n\n"
        "PRICE→unit_cost: strip /MBF /MSF /LF /EA. 755.00/MBF→755.0.\n\n"
        "Include ALL rows from ALL pages (~80 items total). Never skip a row.\n\n"
        "Return ONLY a valid JSON array. Each object:\n"
        "{\"type\":\"Lumber\"|\"LVL\"|\"Panels\"|\"Each\"|\"Unknown\","
        "\"description\":string,\"t_num\":number,\"w_num\":number,\"length_num\":number,"
        "\"qty_change\":number,\"uom\":\"PC\"|\"LF\"|\"EA\","
        "\"unit_cost\":number,\"footage\":number,\"amount\":number}\n\n"
        "CO TEXT:\n"
    ) + text + "\n\nReturn ONLY the JSON array."
    msg = get_claude().messages.create(
        model="claude-sonnet-4-5", max_tokens=16000,
        messages=[{"role": "user", "content": prompt}])
    raw_items = safe_json_parse(next((getattr(b,"text","") for b in msg.content if hasattr(b,"text")),"").strip())

    # Post-process: strip CO# refs and normalise MCA→PT in Python too
    import re
    for item in raw_items:
        d = item.get("description", "")
        # Remove CO #001 / CO#008 patterns
        d = re.sub(r'\bCO\s*#?\s*\d+\b', '', d, flags=re.IGNORECASE).strip()
        # Normalise MCA → PT
        d = re.sub(r'\bMCA TREATED\b', 'PT', d, flags=re.IGNORECASE)
        d = re.sub(r'\bMCA\b', 'PT', d, flags=re.IGNORECASE)
        # Collapse spaces
        d = re.sub(r'\s+', ' ', d).strip()
        item["description"] = d

        # Fix ALL panel dimensions — Claude often misreads "4X8" as T=4, W=4, L=8
        # For panels: t=thickness fraction, w=8 (standard sheet width), length=0
        if item.get("type") == "Panels":
            # Always force w=8 and length=0 for standard 4x8 panels
            item["w_num"]      = 8
            item["length_num"] = 0
            # Fix t_num: must be fractional thickness, not 4
            t_val = n(item.get("t_num", 0))
            if t_val >= 1.0 or t_val == 0.0:
                # t_num is wrong (e.g. 4 from "4X8") — parse from description
                THICKNESS_MAP = {
                    "7/16": 0.4375, "15/32": 0.46875, "1/2": 0.5,
                    "19/32": 0.59375, "5/8": 0.625, "23/32": 0.71875,
                    "3/4": 0.75, "1": 1.0
                }
                found = False
                for frac, val in THICKNESS_MAP.items():
                    if frac in d:
                        item["t_num"] = val
                        found = True
                        break
                if not found:
                    # Try generic fraction pattern
                    m2 = re.search(r'(\d+)/(\d+)', d)
                    if m2:
                        item["t_num"] = round(int(m2.group(1)) / int(m2.group(2)), 6)
                    else:
                        item["t_num"] = 0.46875  # default 15/32

    return raw_items


def parse_invoice_with_claude(text, po_items):
    """
    Parse invoice and match each item to PO/CO items.
    KEY RULE: deliveries for R/L lumber store LINEAR FEET, not piece count.
      - If invoice has 256 PC of 2X6X12 matched to 2x6x1 R/L → store 256*12 = 3072 LF
      - If invoice has 192 PC of 2X8X16 matched to 2x8x1 R/L → store 192*16 = 3072 LF
      - Panels and Each items: store piece count as-is.
    """
    po_summary = json.dumps([
        {"description": i["description"], "type": i["type"],
         "t_num": i.get("t_num", 0), "w_num": i.get("w_num", 0),
         "length_num": i.get("length_num", 0), "unit_cost": i.get("unit_cost", 0)}
        for i in po_items], indent=2)

    prompt = """You are a PRECISE data extraction assistant reading a Matheus Lumber INVOICE.
TASK: Extract EVERY line item with 100% accuracy. This is used for financial accounting — every item must be captured.

INVOICE FORMAT (multi-line per item):
  Line 1:  QTY | UOM | ITEM# | DESCRIPTION/GRADE | FOOTAGE | UNIT PRICE | AMOUNT
  Line 2:  (code) | DIMENSIONS e.g. "2X6X12" or "23/32 4X8"
  Line 3:  "ABOVE SUBSTITUTED FOR" (ignore this line)
  Line 4:  Original spec (use for PO matching)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
EXTRACTION RULES — MUST FOLLOW EXACTLY:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. QTY = the number in column 1. Read it EXACTLY. Do not confuse with footage.
2. UOM "EA" or "PC" = pieces. "MSF" = thousand square feet (panels).
3. UNIT PRICE: extract the number only. "7.250/EA" → 7.25. "1285.000/MSF" → 1285.
4. AMOUNT = last column dollar value. Read it EXACTLY.
5. NEVER skip any line item. All items must appear in output.
6. "ABOVE SUBSTITUTED FOR" lines = ignore, already captured from Line 1.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SILL SEAL / SILL SEALER SIZE NORMALIZATION:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Sill seals use ACTUAL size on invoice but NOMINAL size in PO. Map accordingly:
  - "3-1/2"" or "3.5"" on invoice → matches "4in" or "4"" in PO
  - "5-1/2"" or "5.5"" on invoice → matches "6in" or "6"" in PO
  - "7-1/2"" or "7.5"" on invoice → matches "8in" or "8"" in PO
  - "1-1/2"" on invoice → matches "2in" in PO
Apply this rule ONLY for SILL SEAL / SILL SEALER items.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
LUMBER/PANEL NORMALIZATION:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- "MCA TREATED" or "MCA" → PT (pressure treated)
- "/BTR" → ignore  
- "SYP DRY #2/BTR S4S" → SYP #2
- "R/L" or "XR/L" → random length, invoice_length=1
- "23/32" → T dimension for panels

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
MATCHING RULES:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Match invoice item to PO item by: type + description similarity + unit_cost (within $2)
- For sill seals: use nominal size after normalization above for matching
- For lumber: T + W + length + unit_cost must match
- For panels: unit_cost match is primary (each spec has unique price)
- MCA items → match PT items
- If no match in PO (different dimensions AND different cost) → matched_description = "UNMATCHED"
- DO NOT skip unmatched items. Every item must appear — either matched or UNMATCHED.

RETURN: ONLY a valid JSON array, no markdown. Each element:
{
  "matched_description": "<EXACT description from PO list, or UNMATCHED>",
  "qty_delivered_pcs":   <integer — exact QTY from invoice>,
  "invoice_length":      <number — from dimensions, 1 for R/L, 0 if not lumber>,
  "invoice_t":           <number>,
  "invoice_w":           <number>,
  "invoice_unit_cost":   <number — unit price from invoice>,
  "invoice_amount":      <number — AMOUNT column total>,
  "invoice_description": "<description as printed on invoice>",
  "invoice_uom":         "<UOM: EA, PC, MSF, etc>",
  "rl_match":            <true if matched R/L PO item with specific invoice length>
}

PO/CO ITEMS:
""" + po_summary + "\n\nINVOICE TEXT:\n" + text + "\n\nReturn ONLY the JSON array."

    msg = get_claude().messages.create(
        model="claude-sonnet-4-5", max_tokens=8000,
        messages=[{"role": "user", "content": prompt}])
    return safe_json_parse(next((getattr(b,"text","") for b in msg.content if hasattr(b,"text")),"").strip())


def extract_invoice_header(text):
    """Extract invoice number and date from invoice PDF text."""
    prompt = """Extract the invoice number and invoice date from this Matheus Lumber invoice text.

Return ONLY a JSON object with exactly these keys:
{
  "invoice_no": string (the invoice number, e.g. "60126010" or "60126001"),
  "invoice_date": string (date in MM/DD/YYYY format, e.g. "11/20/2025")
}

If you cannot find either value, use "" for that field.

INVOICE TEXT:
""" + text[:3000] + """

Return ONLY the JSON object."""
    try:
        msg = get_claude().messages.create(
            model="claude-sonnet-4-5", max_tokens=200,
            messages=[{"role": "user", "content": prompt}])
        raw = next((getattr(b,"text","") for b in msg.content if hasattr(b,"text")),"").strip()
        raw = re.sub(r'^```(?:json)?', '', raw).rstrip('`').strip()
        result = json.loads(raw)
        return result.get("invoice_no",""), result.get("invoice_date","")
    except:
        return "", ""


# ── Data helpers ──────────────────────────────────────────────────────────────
def project_xlsx(project):    return PROJECTS / project / "inventory.xlsx"
def project_po_xlsx(project): return PROJECTS / project / "po_report.xlsx"
def project_co_xlsx(project): return PROJECTS / project / "co_report.xlsx"
def project_meta(project):    return PROJECTS / project / "meta.json"

# ── Storage: all JSON data goes through db.py (Postgres) ─────────────────────
def load_meta(project):
    return _db.load_meta(project) or \
           {"invoices": [], "co_count": 0, "change_orders": [], "unmatched_items": []}

def save_meta(project, meta):
    _db.save_meta(project, meta)

def load_items(project):
    return _db.load_items(project) or []

def save_items(project, items):
    _db.save_items(project, items)

def normalize_item(it):
    it.setdefault("t_num", 0.0)
    it.setdefault("w_num", 0.0)
    it.setdefault("length_num", 0.0)
    it.setdefault("po_qty", 0.0)
    it.setdefault("co_qty", 0.0)
    it.setdefault("deliveries", {})
    it.setdefault("unit_cost", 0.0)
    it.setdefault("leftover_lf", 0.0)
    it.setdefault("source", "po")

    typ  = it.get("type", "")
    desc = (it.get("description") or "").strip()
    tv   = n(it.get("t_num", 0))
    wv   = n(it.get("w_num", 0))
    lv   = n(it.get("length_num", 0))

    # ── Auto-build full description from dimensions + material name ───────
    # If description is very short (just "PT", "SYP#2", "SYP #2" etc.),
    # prepend dimensions so the tracker shows "2x4x1 PT" instead of "PT".
    if typ in ("Lumber", "LVL") and tv and wv:
        # Dimension prefix e.g. "2x4x12" or "2x4x1" for R/L
        t_str = str(int(tv)) if tv == int(tv) else str(tv)
        w_str = str(int(wv)) if wv == int(wv) else str(wv)
        l_str = str(int(lv)) if lv == int(lv) else str(lv) if lv else "RL"
        dim_prefix = f"{t_str}x{w_str}x{l_str}"
        # Only prepend if description doesn't already contain dimensions
        import re as _re
        has_dims = bool(_re.search(r'\d+\s*[xX]\s*\d+', desc))
        if not has_dims and desc:
            it["description"] = f"{dim_prefix} {desc}"
        elif not desc:
            it["description"] = dim_prefix

    # For Panels: keep fractional t_num for cost calculation, set display t=4
    if typ == "Panels":
        raw_t = n(it.get("t_num", 0))
        if raw_t > 0 and raw_t < 1.0:
            it["thickness_frac"] = raw_t   # save fractional e.g. 0.46875
        it["t_num"] = 4.0                  # display as 4 (4ft sheet dim)

    return it

def score_match(item, co):
    """Score how well a CO item matches a PO item. Higher = better."""
    s = 0
    # Type match
    if item.get("type") == co.get("type"): s += 4
    elif co.get("type") == "Unknown": s += 1  # unknown type — partial credit
    # Dimension match
    if abs(n(item.get("t_num", 0))      - n(co.get("t_num", 0)))      < 0.001: s += 5
    if abs(n(item.get("w_num", 0))      - n(co.get("w_num", 0)))      < 0.001: s += 5
    if abs(n(item.get("length_num", 0)) - n(co.get("length_num", 0))) < 0.001: s += 4
    # Description word overlap (normalised)
    def words(s): return set(s.lower().replace("x"," ").split())
    cw = words(co.get("description", ""))
    iw = words(item.get("description", ""))
    s += len(cw & iw)
    return s


# ── Excel builder: Inventory tracker ─────────────────────────────────────────
def rebuild_excel(project):
    items    = load_items(project)
    meta     = load_meta(project)
    invoices = meta.get("invoices", [])
    cos      = meta.get("change_orders", [])
    n_inv    = len(invoices)
    C        = get_column_letter
    # Per-project tax rates
    PO_TAX   = n(meta.get("po_tax_rate",       meta.get("tax_rate", TAX_RATE)))
    DEL_TAX  = n(meta.get("delivery_tax_rate",  meta.get("tax_rate", TAX_RATE)))
    wb = openpyxl.Workbook()

    # ── Sheet 1: Inventory ────────────────────────────────────────────────────
    ws = wb.active
    assert ws is not None  # type: ignore
    ws.title = "Inventory"

    # Column layout
    # 1=Type 2=Desc 3=T 4=W 5=Len
    # 6=POQty 7=COQty 8=PO+COQty
    # 9=UnitCost 10=LF 11=BF/SF 12=TotalCost 13=TotalCost+Tax
    # 14..13+n_inv = invoice date cols
    # 14+n_inv = Invoice# text
    # 14+n_inv+1 = TotalDeliveredQty
    # 14+n_inv+2 = DeliveredLF
    # 14+n_inv+3 = DeliveredBF/SF
    # 14+n_inv+4 = DeliveredCost
    # 14+n_inv+5 = DeliveredCost+Tax
    # 14+n_inv+6 = RemainingQty
    # 14+n_inv+7 = RemainingLF
    # 14+n_inv+8 = RemainingBF/SF
    # 14+n_inv+9 = RemainingCost
    # 14+n_inv+10 = RemainingCost+Tax
    # 14+n_inv+11 = LeftoverLF
    # 14+n_inv+12 = MaterialUsedLF
    # 14+n_inv+13 = MaterialUsedCost

    IDS  = 14
    inc  = IDS + n_inv
    dqc  = inc + 1
    dlfc = inc + 2
    dbfc = inc + 3
    dcc  = inc + 4
    dcx  = inc + 5
    rqc  = inc + 6
    rlfc = inc + 7
    rbfc = inc + 8
    rcc  = inc + 9
    rcx  = inc + 10
    lofc = inc + 11   # Leftover LF
    ulfc = inc + 12   # Material Used LF
    ucc  = inc + 13   # Material Used Cost
    LAST_COL = ucc

    headers = ["Type","Description","T","W","Length",
               "PO Qty","CO Qty","PO+CO Qty",
               "Unit Cost","LF (Linear Ft)","BF/SF","Total Cost","Total Cost+Tax"]
    for inv in invoices: headers.append(inv['date'])
    headers += ["Invoice #",
                "Total Delivered","Delivered LF","Delivered BF/SF",
                "Delivered Cost","Delivered Cost+Tax",
                "Remaining Qty","Remaining LF","Remaining BF/SF",
                "Remaining Cost","Remaining Cost+Tax",
                "Leftover LF (Site)","Material Used LF","Material Used Cost"]

    ws.row_dimensions[1].height = 44
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    TYPE_ORDER = ["Lumber", "LVL", "Each", "Panels", "Siding", "HouseWrap", "Unknown"]
    grouped    = {t: [i for i in items if i.get("type") == t] for t in TYPE_ORDER}
    row = 2
    subtotal_rows = {}

    for typ in TYPE_ORDER:
        type_items = grouped[typ]
        if not type_items: continue
        ds = row

        for item in type_items:
            fill = TYPE_FILLS.get(typ, PatternFill())
            invd = item.get("deliveries", {})
            r    = row
            uc   = n(item.get("unit_cost", 0))
            tv   = n(item.get("t_num", 0))
            wv   = n(item.get("w_num", 0))
            lv   = n(item.get("length_num", 0))
            lo_lf= n(item.get("leftover_lf", 0))

            H = f"H{r}"; I = f"I{r}"

            # LF formula
            if typ in ("Lumber", "LVL"):
                lf_formula = f"={H}*{lv}" if lv else 0
            else:
                lf_formula = 0

            # BF/SF formula
            if typ == "Lumber":
                bf_formula = f"=({H}*{tv}*{wv}*{lv})/12" if (tv and wv and lv) else 0
            elif typ == "Panels":
                bf_formula = f"={H}*{tv}*{wv}" if (tv and wv) else 0
            else:
                bf_formula = 0

            LF_col = f"J{r}"; BF_col = f"K{r}"

            # Total Cost
            if typ == "Lumber":   tc = f"=({BF_col}*{I})/1000"
            elif typ == "Panels": tc = f"=({BF_col}*{I})/1000"
            elif typ == "LVL":    tc = f"={LF_col}*{I}"
            elif typ == "Each":   tc = f"={H}*{I}"
            else:                 tc = f"={H}*{I}"
            tcx = f"=L{r}*{round(1 + PO_TAX, 6)}"

            # Invoice delivery cols
            inv_vals = [invd.get(inv["invoice_no"], 0) for inv in invoices]
            inv_nos  = "\n".join(inv["invoice_no"] for inv in invoices
                                 if invd.get(inv["invoice_no"], 0) > 0)

            dq_col = C(dqc)
            dqf = f"=SUM({C(IDS)}{r}:{C(IDS+n_inv-1)}{r})" if invoices else 0

            # Is this an R/L item? (length=1 Lumber/LVL)
            # For R/L items, deliveries stores LF directly, so:
            #   TotalDelivered col = LF already (no multiply by length)
            #   DeliveredLF = TotalDelivered (same value)
            #   Remaining = TotalLF (col J) - DeliveredLF
            is_rl_item = (typ in ("Lumber","LVL") and lv == 1.0)

            # Delivered LF/BF/Cost
            if is_rl_item:
                # deliveries stored as LF → TotalDelivered IS the LF
                dlf_f = f"={dq_col}{r}"                              # LF = delivered directly
                dbf_f = f"=({dq_col}{r}*{tv}*{wv})/12" if (tv and wv) else 0
                dcf   = f"=({C(dbfc)}{r}*{I})/1000" if typ=="Lumber" else f"={C(dlfc)}{r}*{I}"
            elif typ in ("Lumber", "LVL"):
                dlf_f = f"={dq_col}{r}*{lv}" if lv else 0
                dbf_f = f"=({dq_col}{r}*{tv}*{wv}*{lv})/12" if (tv and wv and lv) else 0
                dcf   = f"=({C(dbfc)}{r}*{I})/1000" if typ=="Lumber" else f"={C(dlfc)}{r}*{I}"
            elif typ == "Panels":
                dlf_f = 0
                dbf_f = f"={dq_col}{r}*{tv}*{wv}" if (tv and wv) else 0
                dcf   = f"=({C(dbfc)}{r}*{I})/1000"
            else:
                dlf_f = 0; dbf_f = 0
                dcf   = f"={dq_col}{r}*{I}"
            dxf = f"={C(dcc)}{r}*{round(1 + DEL_TAX, 6)}"

            # Remaining Qty / LF / BF / Cost
            if is_rl_item:
                # Remaining LF = Total LF (col J) - Delivered LF (col dlfc)
                rqf   = f"={LF_col}-{C(dlfc)}{r}"                    # remaining LF
                RQ    = C(rqc)
                rlf_f = f"={RQ}{r}"                                   # remaining LF = rqf
                rbf_f = f"=({RQ}{r}*{tv}*{wv})/12" if (tv and wv) else 0
                rcf   = f"=({C(rbfc)}{r}*{I})/1000" if typ=="Lumber" else f"={RQ}{r}*{I}"
            else:
                rqf = f"={H}-{dq_col}{r}"
                RQ  = C(rqc)
                if typ in ("Lumber", "LVL"):
                    rlf_f = f"={RQ}{r}*{lv}" if lv else 0
                elif typ == "Panels":
                    rlf_f = 0
                else:
                    rlf_f = 0
                if typ == "Lumber":
                    rbf_f = f"=({RQ}{r}*{tv}*{wv}*{lv})/12" if (tv and wv and lv) else 0
                elif typ == "Panels":
                    rbf_f = f"={RQ}{r}*{tv}*{wv}" if (tv and wv) else 0
                else:
                    rbf_f = 0
                if typ == "Lumber":   rcf = f"=({C(rbfc)}{r}*{I})/1000"
                elif typ == "Panels": rcf = f"=({C(rbfc)}{r}*{I})/1000"
                elif typ == "LVL":    rcf = f"={C(rlfc)}{r}*{I}"
                else:                 rcf = f"={RQ}{r}*{I}"
            rxf = f"={C(rcc)}{r}*{round(1 + PO_TAX, 6)}"

            # Leftover and Material Used LF / Cost
            # MaterialUsedLF = DeliveredLF - LeftoverLF
            # MaterialUsedCost: formula based on type
            lo_val = lo_lf  # actual leftover LF entered by user
            DLF_col = C(dlfc)
            LO_col  = C(lofc)
            ul_col  = C(ulfc)

            if typ in ("Lumber", "LVL") and lv:
                used_lf_f = f"={DLF_col}{r}-{LO_col}{r}"
                if typ == "Lumber":
                    used_cost_f = f"=(({ul_col}{r}*{tv}*{wv})/12*{I})/1000" if (tv and wv) else 0
                else:  # LVL
                    used_cost_f = f"={ul_col}{r}*{I}"
            else:
                used_lf_f   = 0
                used_cost_f = 0

            row_data = [
                item.get("type",""), item.get("description",""),
                tv or "", wv or "", lv or "",
                n(item.get("po_qty",0)), n(item.get("co_qty",0)), f"=F{r}+G{r}",
                uc, lf_formula, bf_formula, tc, tcx,
            ] + inv_vals + [
                inv_nos, dqf, dlf_f, dbf_f, dcf, dxf,
                rqf, rlf_f, rbf_f, rcf, rxf,
                lo_val, used_lf_f, used_cost_f,
            ]

            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.fill = fill; cell.font = NORMAL_FONT; cell.border = BORDER
                cell.alignment = Alignment(
                    horizontal="left" if ci == 2 else "center",
                    vertical="center", wrap_text=True)
                if ci == 9:                              cell.number_format = DEC_FMT
                elif ci in (12,13,dcc,dcx,rcc,rcx,ucc): cell.number_format = MONEY_FMT
                elif ci in (6,7,8,dqc,rqc):             cell.number_format = INT_FMT
                elif ci in (10,11,dlfc,dbfc,rlfc,rbfc,lofc,ulfc): cell.number_format = DEC_FMT
                elif IDS <= ci < IDS+n_inv:              cell.number_format = INT_FMT
            row += 1

        # Subtotal
        de = row - 1
        subtotal_rows[typ] = row
        sfill = SUBTOTAL_FILLS.get(typ, SUBTOTAL_FILLS["Unknown"])
        sd = {1: f"{typ.upper()} SUBTOTAL"}
        for ci in [6,7,8,10,11,12,13,dqc,dlfc,dbfc,dcc,dcx,rqc,rlfc,rbfc,rcc,rcx,lofc,ulfc,ucc]:
            sd[ci] = f"=SUM({C(ci)}{ds}:{C(ci)}{de})"
        for idx in range(n_inv):
            ci = IDS+idx; sd[ci] = f"=SUM({C(ci)}{ds}:{C(ci)}{de})"

        ws.row_dimensions[row].height = 18
        for ci in range(1, LAST_COL+1):
            val  = sd.get(ci, "")
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = sfill; cell.font = SUBTOT_FONT; cell.border = BORDER
            cell.alignment = Alignment(horizontal="left" if ci==1 else "center", vertical="center")
            if ci in (12,13,dcc,dcx,rcc,rcx,ucc): cell.number_format = MONEY_FMT
            elif ci in (6,7,8,dqc,rqc):           cell.number_format = INT_FMT
            elif ci in (10,11,dlfc,dbfc,rlfc,rbfc,lofc,ulfc): cell.number_format = DEC_FMT
        row += 1
        ws.row_dimensions[row].height = 5
        row += 1

    # Grand Total
    srows = list(subtotal_rows.values())
    def gsum(ci): return "=" + "+".join(f"{C(ci)}{sr}" for sr in srows)
    ws.row_dimensions[row].height = 22
    for ci in range(1, LAST_COL+1):
        val = ""
        if ci == 1: val = "GRAND TOTAL"
        elif ci in [6,7,8,10,11,12,13,dqc,dlfc,dbfc,dcc,dcx,rqc,rlfc,rbfc,rcc,rcx,lofc,ulfc,ucc]:
            val = gsum(ci)
        elif IDS <= ci < IDS+n_inv:
            val = gsum(ci)
        cell = ws.cell(row=row, column=ci, value=val)
        cell.fill = GRAND_FILL; cell.font = GRAND_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal="left" if ci==1 else "center", vertical="center")
        if ci in (12,13,dcc,dcx,rcc,rcx,ucc): cell.number_format = MONEY_FMT
        elif ci in (6,7,8,dqc,rqc):           cell.number_format = INT_FMT
        elif ci in (10,11,dlfc,dbfc,rlfc,rbfc,lofc,ulfc): cell.number_format = DEC_FMT

    # Column widths
    widths = {1:10,2:34,3:7,4:6,5:7,6:9,7:9,8:9,9:11,10:12,11:12,12:14,13:14}
    for idx in range(n_inv): widths[IDS+idx] = 12
    for ci, w in zip([inc,dqc,dlfc,dbfc,dcc,dcx,rqc,rlfc,rbfc,rcc,rcx,lofc,ulfc,ucc],
                     [18,13,12,12,14,15,12,12,12,14,15,14,14,14]):
        widths[ci] = w
    for ci, w in widths.items():
        ws.column_dimensions[C(ci)].width = w
    ws.freeze_panes = "B2"

    # ── Sheet 2: Change Orders ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Change Orders")
    co_hdrs = ["CO #","CO Date","Type","Description","T","W","Len",
               "Qty Change","Unit Cost","Footage/SF","LF","BF/SF","Amount","Amount+Tax","Action","Matched PO Item"]
    ws2.row_dimensions[1].height = 30
    for ci, h in enumerate(co_hdrs, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font=HEADER_FONT; cell.fill=HEADER_FILL
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=BORDER

    cr = 2
    for co in cos:
        for ci_item in co.get("items", []):
            fill = TYPE_FILLS.get(ci_item.get("type","Unknown"), TYPE_FILLS["Unknown"])
            amt  = n(ci_item.get("amount", 0))
            qty  = n(ci_item.get("qty_change", 0))
            tv   = n(ci_item.get("t_num", 0))
            wv   = n(ci_item.get("w_num", 0))
            lv   = n(ci_item.get("length_num", 0))
            typ  = ci_item.get("type","")
            uc   = n(ci_item.get("unit_cost", 0))
            ftg  = n(ci_item.get("footage", 0))
            abs_qty = abs(qty) if qty else 1
            lf   = compute_lf(typ, abs_qty, tv, wv, lv)
            bfsf = compute_bf_sf(typ, abs_qty, tv, wv, lv)
            if bfsf == 0 and ftg: bfsf = ftg
            row_vals = [
                co.get("co_no",""), co.get("date",""), typ,
                ci_item.get("description",""), tv or "", wv or "", lv or "",
                qty, uc, ftg, lf, bfsf, amt, round(amt*(1+PO_TAX),2),
                ci_item.get("action",""), ci_item.get("matched_po",""),
            ]
            row_vals[13] = round(amt*(1+PO_TAX),2)  # use per-project PO tax
            for ci, val in enumerate(row_vals, 1):
                cell = ws2.cell(row=cr, column=ci, value=val)
                cell.fill=fill; cell.font=NORMAL_FONT; cell.border=BORDER
                cell.alignment=Alignment(
                    horizontal="left" if ci in (4,15,16) else "center",
                    vertical="center", wrap_text=True)
                if ci in (9,13,14): cell.number_format = MONEY_FMT
                elif ci == 8:       cell.number_format = "#,##0"
            cr += 1
    for ci, w in enumerate([8,12,10,34,6,6,6,10,11,12,10,10,14,14,16,28],1):
        ws2.column_dimensions[C(ci)].width = w
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Invoices ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Invoices")
    inv_hdrs = ["Invoice #","Date","Description","Type","T","W","Length",
                "Qty Delivered","Delivered LF","Delivered BF/SF","Delivered Cost","Delivered Cost+Tax"]
    ws3.row_dimensions[1].height = 30
    for ci, h in enumerate(inv_hdrs, 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.font=HEADER_FONT; cell.fill=HEADER_FILL
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=BORDER

    ir = 2
    for inv in invoices:
        inv_no   = inv["invoice_no"]
        inv_date = inv.get("date","")
        for item in items:
            qty_del = n(item.get("deliveries",{}).get(inv_no, 0))
            if qty_del == 0: continue
            typ = item.get("type","")
            tv  = n(item.get("t_num",0))
            wv  = n(item.get("w_num",0))
            lv  = n(item.get("length_num",0))
            uc  = n(item.get("unit_cost",0))
            dlf  = compute_lf(typ, qty_del, tv, wv, lv)
            dbf  = compute_bf_sf(typ, qty_del, tv, wv, lv)
            dcost= compute_cost(typ, qty_del, tv, wv, lv, uc)
            fill = TYPE_FILLS.get(typ, TYPE_FILLS["Unknown"])
            row_vals = [inv_no, inv_date, item.get("description",""), typ,
                        tv or "", wv or "", lv or "", qty_del,
                        dlf, dbf, dcost, round(dcost*(1+DEL_TAX),2)]
            for ci, val in enumerate(row_vals, 1):
                cell = ws3.cell(row=ir, column=ci, value=val)
                cell.fill=fill; cell.font=NORMAL_FONT; cell.border=BORDER
                cell.alignment=Alignment(horizontal="left" if ci==3 else "center", vertical="center")
                if ci in (11,12): cell.number_format = MONEY_FMT
                elif ci == 8:     cell.number_format = INT_FMT
                elif ci in (9,10): cell.number_format = DEC_FMT
            ir += 1
    for ci, w in enumerate([14,12,34,10,6,6,8,12,12,12,14,14],1):
        ws3.column_dimensions[C(ci)].width = w
    ws3.freeze_panes = "A2"

    # ── Sheet 4: Unmatched Items ──────────────────────────────────────────────
    unmatched = load_meta(project).get("unmatched_items", [])
    if unmatched:
        ws4 = wb.create_sheet("Unmatched Items")
        um_hdrs = ["Invoice #","Date","Invoice Description","T","W","Length",
                   "Qty","Unit Cost","Notes","Status"]
        ws4.row_dimensions[1].height = 30
        for ci, h in enumerate(um_hdrs, 1):
            cell = ws4.cell(row=1, column=ci, value=h)
            cell.font=HEADER_FONT; cell.fill=HEADER_FILL
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            cell.border=BORDER
        ur = 2
        warn_fill = PatternFill("solid", start_color="FFF3CD")
        for um in unmatched:
            row_vals = [um.get("invoice_no",""), um.get("invoice_date",""),
                        um.get("description","UNKNOWN"),
                        um.get("t",""), um.get("w",""), um.get("length",""),
                        um.get("qty",0), um.get("unit_cost",0),
                        um.get("notes","Not identified — please assign to a line item"),
                        um.get("status","Pending")]
            for ci, val in enumerate(row_vals, 1):
                cell = ws4.cell(row=ur, column=ci, value=val)
                cell.fill=warn_fill; cell.font=NORMAL_FONT; cell.border=BORDER
                cell.alignment=Alignment(horizontal="left" if ci in (3,9,10) else "center", vertical="center")
                if ci == 8: cell.number_format = DEC_FMT
            ur += 1
        for ci, w in enumerate([14,12,34,6,6,8,10,12,40,14],1):
            ws4.column_dimensions[C(ci)].width = w
        ws4.freeze_panes = "A2"

    # Save inventory.xlsx FIRST — this is the most important file
    try:
        wb.save(str(project_xlsx(project)))
    except Exception as e:
        raise RuntimeError(f"Could not save inventory.xlsx: {e}. "
                           "If the file is open in Excel, please close it and try again.")

    # PO and CO reports are built separately — failures here don't affect inventory.xlsx
    try:
        _build_po_excel(project, items, meta)
    except Exception as e:
        print(f"[WARN] PO Excel build failed (file may be open in Excel): {e}")

    try:
        _build_co_excel(project, meta)
    except Exception as e:
        print(f"[WARN] CO Excel build failed (file may be open in Excel): {e}")


def _build_inventory_excel_only(project, items_list, meta):
    """Build inventory Excel matching the reference format exactly:
    Sheet 1: Inventory (Type,Desc,T,W,Len,POQty,COQty,PO+CO,UnitCost,LF,BF/SF,TotalCost,TotalCost+Tax,
                        InvDate×N, Invoice#, TotalDel, DelLF, DelBF/SF, DelCost, DelCost+Tax,
                        RemQty, RemLF, RemBF/SF, RemCost, RemCost+Tax,
                        LeftoverLF, UsedLF, UsedCost)
    Sheet 2: Change Orders
    Sheet 3: Invoices
    """
    import openpyxl as _xl
    from openpyxl.styles import (Font as _F, PatternFill as _PF, Alignment as _A,
                                  Border as _B, Side as _S)
    from openpyxl.utils import get_column_letter as _gcl

    po_tax  = n(meta.get("po_tax_rate",       meta.get("tax_rate", TAX_RATE)))
    del_tax = n(meta.get("delivery_tax_rate",  meta.get("tax_rate", TAX_RATE)))
    invoices = meta.get("invoices", [])
    cos      = meta.get("change_orders", [])

    # ── Styles ────────────────────────────────────────────────────────────────
    def _side(): return _S(border_style="thin", color="CCCCCC")
    def _bdr():  return _B(left=_side(),right=_side(),top=_side(),bottom=_side())

    HFILL  = _PF("solid", start_color="1B3A5C")
    HFONT  = _F(bold=True, color="FFFFFF", name="Arial", size=10)
    NFONT  = _F(name="Arial", size=9)
    SFILL  = _PF("solid", start_color="2C3E50")
    SFONT  = _F(bold=True, color="FFFFFF", name="Arial", size=9)
    GFILL  = _PF("solid", start_color="1A252F")
    GFONT  = _F(bold=True, color="FFD700",  name="Arial", size=10)
    MFMT   = '"$"#,##0.00'
    NFMT   = '#,##0.0##'
    IFMT   = '#,##0'

    TYPE_FILLS = {
        "Lumber":    _PF("solid", start_color="DDEEFF"),
        "LVL":       _PF("solid", start_color="E8D5FF"),
        "Each":      _PF("solid", start_color="FFF3CD"),
        "Panels":    _PF("solid", start_color="D4EDDA"),
        "Siding":    _PF("solid", start_color="FCE4EC"),
        "HouseWrap": _PF("solid", start_color="E0F7FA"),
        "Unknown":   _PF("solid", start_color="F8F8F8"),
    }
    TYPE_SFILLS = {
        "Lumber":    _PF("solid", start_color="A8C8E8"),
        "LVL":       _PF("solid", start_color="C8A8E8"),
        "Each":      _PF("solid", start_color="E8D888"),
        "Panels":    _PF("solid", start_color="A8C8A8"),
        "Siding":    _PF("solid", start_color="E8A8C8"),
        "HouseWrap": _PF("solid", start_color="A8D8E8"),
        "Unknown":   _PF("solid", start_color="C8C8C8"),
    }

    def hdr(ws, r, c, v, fill=None, font=None, halign="center", wrap=False):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = font or HFONT
        cell.fill      = fill or HFILL
        cell.border    = _bdr()
        cell.alignment = _A(horizontal=halign, vertical="center", wrap_text=wrap)
        return cell

    def val(ws, r, c, v, fill=None, font=None, fmt=None, halign="center"):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = font or NFONT
        cell.fill      = fill or _PF("solid", start_color="FFFFFF")
        cell.border    = _bdr()
        cell.alignment = _A(horizontal=halign, vertical="center")
        if fmt: cell.number_format = fmt
        return cell

    wb = _xl.Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1: INVENTORY
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    assert ws is not None  # type: ignore
    ws.title = "Inventory"
    ws.row_dimensions[1].height = 44
    ws.freeze_panes = "B2"

    # Column layout (matching reference exactly)
    n_inv = len(invoices)
    INV_START = 14          # first invoice date column
    INC       = INV_START + n_inv
    DQ  = INC + 1           # Total Delivered
    DLF = INC + 2           # Delivered LF
    DBF = INC + 3           # Delivered BF/SF
    DC  = INC + 4           # Delivered Cost
    DCX = INC + 5           # Delivered Cost+Tax
    RQ  = INC + 6           # Remaining Qty
    RLF = INC + 7           # Remaining LF
    RBF = INC + 8           # Remaining BF/SF
    RC  = INC + 9           # Remaining Cost
    RCX = INC + 10          # Remaining Cost+Tax
    LO  = INC + 11          # Leftover LF
    ULF = INC + 12          # Used LF
    UC_ = INC + 13          # Used Cost
    LAST = UC_

    # Fixed headers
    fixed_hdrs = ["Type","Description","T","W","Length",
                  "PO Qty","CO Qty","PO+CO Qty","Unit Cost",
                  "LF (Linear Ft)","BF/SF","Total Cost","Total Cost+Tax"]
    for ci,h in enumerate(fixed_hdrs,1):
        hdr(ws,1,ci,h, wrap=True)

    # Invoice date columns (one per invoice)
    for i,inv in enumerate(invoices):
        hdr(ws,1, INV_START+i, inv.get("date",""), wrap=True)

    # Remaining headers
    post_hdrs = {
        INC: "Invoice #",
        DQ:  "Total Delivered",  DLF: "Delivered LF",
        DBF: "Delivered BF/SF",  DC:  "Delivered Cost",
        DCX: "Delivered Cost+Tax",
        RQ:  "Remaining Qty",    RLF: "Remaining LF",
        RBF: "Remaining BF/SF",  RC:  "Remaining Cost",
        RCX: "Remaining Cost+Tax",
        LO:  "Leftover LF (Site)", ULF: "Material Used LF",
        UC_: "Material Used Cost",
    }
    for ci,h in post_hdrs.items():
        hdr(ws,1,ci,h, wrap=True)

    TYPE_ORDER = ["Lumber","LVL","Each","Panels","Siding","HouseWrap","Unknown"]
    row = 2
    subtotal_rows = {}
    grand = {k:0.0 for k in
             ["lf","bf","cost","cost_tax","del_lf","del_bf","del_cost","del_cost_tax",
              "rem_lf","rem_bf","rem_cost","rem_cost_tax","used_lf","used_cost"]}

    for typ in TYPE_ORDER:
        type_items = [it for it in items_list if it.get("type")==typ]
        if not type_items: continue
        ds = row
        fill  = TYPE_FILLS.get(typ, _PF("solid",start_color="F8F8F8"))
        sfill = TYPE_SFILLS.get(typ, _PF("solid",start_color="C8C8C8"))

        g = {k:0.0 for k in grand}

        for item in type_items:
            tv  = n(item.get("t_num",0)); wv=n(item.get("w_num",0)); lv=n(item.get("length_num",0))
            uc  = n(item.get("unit_cost",0))
            poq = n(item.get("po_qty",0)); coq=n(item.get("co_qty",0)); tq=poq+coq
            lf  = compute_lf(typ,tq,tv,wv,lv)
            bf  = compute_bf_sf(typ,tq,tv,wv,lv)
            cf  = item.get("cost_formula", None)
            cost= compute_cost(typ,tq,tv,wv,lv,uc,cf)
            ctax= round(cost*(1+po_tax),2)

            invd     = item.get("deliveries",{})
            inv_vals = [n(invd.get(inv["invoice_no"],0)) for inv in invoices]
            del_total= sum(inv_vals)
            inv_nos  = "\n".join(inv["invoice_no"] for inv in invoices
                                 if invd.get(inv["invoice_no"],0))

            dlf  = compute_lf(typ,del_total,tv,wv,lv)
            dbf  = compute_bf_sf(typ,del_total,tv,wv,lv)
            dcost= compute_cost(typ,del_total,tv,wv,lv,uc,cf)
            dctax= round(dcost*(1+del_tax),2)

            rem  = tq - del_total
            rlf  = compute_lf(typ,max(0,rem),tv,wv,lv)
            rbf  = compute_bf_sf(typ,max(0,rem),tv,wv,lv)
            rcost= compute_cost(typ,max(0,rem),tv,wv,lv,uc,cf)
            rctax= round(rcost*(1+po_tax),2)

            lo_lf   = n(item.get("leftover_lf",0))
            used_lf = max(0.0, dlf-lo_lf)
            if typ=="Lumber" and tv and wv:
                used_cost = round((used_lf*tv*wv/12*uc)/1000,2)
            elif typ=="LVL":
                used_cost = round(used_lf*uc,2)
            elif del_total>0:
                used_cost = round(dcost*max(0,(del_total-lo_lf))/del_total,2)
            else:
                used_cost=0.0

            # Accumulate
            for k,v in [("lf",lf),("bf",bf),("cost",cost),("cost_tax",ctax),
                        ("del_lf",dlf),("del_bf",dbf),("del_cost",dcost),("del_cost_tax",dctax),
                        ("rem_lf",rlf),("rem_bf",rbf),("rem_cost",rcost),("rem_cost_tax",rctax),
                        ("used_lf",used_lf),("used_cost",used_cost)]:
                g[k]+=v; grand[k]+=v

            # Write row
            def w(c,v,fmt=None,ha="center"):
                val(ws,row,c,v,fill=fill,fmt=fmt,halign=ha)

            val(ws,row,1,typ,fill=fill,halign="center")
            val(ws,row,2,item.get("description",""),fill=fill,halign="left")
            for ci,v in enumerate([tv or "",wv or "",lv or "",poq,coq,tq,uc],3):
                fmt = MFMT if ci==9 else (NFMT if ci in (10,11) else IFMT if ci in (6,7,8) else None)
                val(ws,row,ci,v,fill=fill,fmt=fmt)
            val(ws,row,10, lf, fill=fill, fmt=NFMT)
            val(ws,row,11, bf, fill=fill, fmt=NFMT)
            val(ws,row,12, cost, fill=fill, fmt=MFMT)
            val(ws,row,13, ctax, fill=fill, fmt=MFMT)
            for i,v in enumerate(inv_vals): val(ws,row,INV_START+i,v,fill=fill,fmt=IFMT)
            val(ws,row,INC, inv_nos, fill=fill, halign="left")
            val(ws,row,DQ,  del_total, fill=fill, fmt=IFMT)
            val(ws,row,DLF, dlf,       fill=fill, fmt=NFMT)
            val(ws,row,DBF, dbf,       fill=fill, fmt=NFMT)
            val(ws,row,DC,  dcost,     fill=fill, fmt=MFMT)
            val(ws,row,DCX, dctax,     fill=fill, fmt=MFMT)
            val(ws,row,RQ,  rem,       fill=fill, fmt=IFMT)
            val(ws,row,RLF, rlf,       fill=fill, fmt=NFMT)
            val(ws,row,RBF, rbf,       fill=fill, fmt=NFMT)
            val(ws,row,RC,  rcost,     fill=fill, fmt=MFMT)
            val(ws,row,RCX, rctax,     fill=fill, fmt=MFMT)
            val(ws,row,LO,  lo_lf,     fill=fill, fmt=NFMT)
            val(ws,row,ULF, used_lf,   fill=fill, fmt=NFMT)
            val(ws,row,UC_, used_cost,  fill=fill, fmt=MFMT)
            row += 1

        # Subtotal row
        subtotal_rows[typ] = row
        for ci in range(1, LAST+1):
            cell = ws.cell(row=row, column=ci)
            cell.font=SFONT; cell.fill=sfill; cell.border=_bdr()
            cell.alignment=_A(horizontal="left" if ci==1 else "center",vertical="center")
        ws.cell(row=row,column=1).value = f"{typ.upper()} SUBTOTAL"  # type: ignore[union-attr]
        sub_map = {
            6:sum(n(it.get("po_qty",0)) for it in type_items),
            7:sum(n(it.get("co_qty",0)) for it in type_items),
            8:sum(n(it.get("po_qty",0))+n(it.get("co_qty",0)) for it in type_items),
            10:g["lf"],11:g["bf"],12:g["cost"],13:g["cost_tax"],
            DQ:sum(sum(n(it.get("deliveries",{}).get(inv["invoice_no"],0)) for inv in invoices) for it in type_items),
            DLF:g["del_lf"],DBF:g["del_bf"],DC:g["del_cost"],DCX:g["del_cost_tax"],
            RLF:g["rem_lf"],RBF:g["rem_bf"],RC:g["rem_cost"],RCX:g["rem_cost_tax"],
            LO:sum(n(it.get("leftover_lf",0)) for it in type_items),
            ULF:g["used_lf"],UC_:g["used_cost"],
        }
        for ci,v in sub_map.items():
            c=ws.cell(row=row,column=ci); c.value=v; c.font=SFONT; c.fill=sfill; c.border=_bdr()
            c.alignment=_A(horizontal="center",vertical="center")
            if ci in (12,13,DC,DCX,RC,RCX,UC_): c.number_format=MFMT
            elif ci in (10,11,DLF,DBF,RLF,RBF,LO,ULF): c.number_format=NFMT
            elif ci in (6,7,8,DQ,RQ): c.number_format=IFMT
        row += 1
        # Spacer
        ws.row_dimensions[row].height=5; row+=1

    # Grand Total
    ws.row_dimensions[row].height=22
    for ci in range(1,LAST+1):
        c=ws.cell(row=row,column=ci); c.font=GFONT; c.fill=GFILL; c.border=_bdr()
        c.alignment=_A(horizontal="left" if ci==1 else "center",vertical="center")
    ws.cell(row=row,column=1).value="GRAND TOTAL"  # type: ignore[union-attr]
    grand_map = {
        10:grand["lf"],11:grand["bf"],12:grand["cost"],13:grand["cost_tax"],
        DLF:grand["del_lf"],DBF:grand["del_bf"],DC:grand["del_cost"],DCX:grand["del_cost_tax"],
        RLF:grand["rem_lf"],RBF:grand["rem_bf"],RC:grand["rem_cost"],RCX:grand["rem_cost_tax"],
        ULF:grand["used_lf"],UC_:grand["used_cost"],
    }
    for ci,v in grand_map.items():
        c=ws.cell(row=row,column=ci); c.value=v; c.font=GFONT; c.fill=GFILL; c.border=_bdr()
        c.alignment=_A(horizontal="center",vertical="center")
        if ci in (12,13,DC,DCX,RC,RCX,UC_): c.number_format=MFMT
        elif ci in (10,11,DLF,DBF,RLF,RBF,ULF): c.number_format=NFMT

    # Column widths
    ws.column_dimensions["A"].width=10
    ws.column_dimensions["B"].width=36
    for c in "CDEFGHI": ws.column_dimensions[c].width=9
    ws.column_dimensions["J"].width=12; ws.column_dimensions["K"].width=12
    ws.column_dimensions["L"].width=14; ws.column_dimensions["M"].width=14
    for i in range(n_inv): ws.column_dimensions[_gcl(INV_START+i)].width=12
    for ci,w in zip(range(INC,LAST+1),[18,13,13,13,14,15,12,12,12,14,15,14,14,14]):  # type: ignore[assignment]
        ws.column_dimensions[_gcl(ci)].width=w

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: CHANGE ORDERS  (matching reference exactly)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Change Orders")
    ws2.row_dimensions[1].height=36
    co_hdrs=["CO #","CO Date","Type","Description","T","W","Len",
             "Qty Change","Unit Cost","Footage/SF","LF","BF/SF",
             "Amount","Amount+Tax","Action","Matched PO Item"]
    for ci,h in enumerate(co_hdrs,1):
        hdr(ws2,1,ci,h,wrap=True)

    cr=2
    for co in cos:
        for ci_item in co.get("items",[]):
            typ  = ci_item.get("type","")
            fill = TYPE_FILLS.get(typ,_PF("solid",start_color="F8F8F8"))
            qty  = n(ci_item.get("qty_change",0))
            tv   = n(ci_item.get("t_num",0)); wv=n(ci_item.get("w_num",0)); lv=n(ci_item.get("length_num",0))
            uc   = n(ci_item.get("unit_cost",0))
            sign = -1 if qty<0 else 1
            abs_qty=abs(qty)
            ftg  = compute_bf_sf(typ,abs_qty,tv,wv,lv)*sign
            lf_  = compute_lf(typ,abs_qty,tv,wv,lv)*sign
            bf_  = compute_bf_sf(typ,abs_qty,tv,wv,lv)*sign
            amt  = compute_cost(typ,abs_qty,tv,wv,lv,uc)*sign
            amtt = round(amt*(1+po_tax),2)
            row_vals=[co.get("co_no",""),co.get("date",""),typ,
                      ci_item.get("description",""),tv or "",wv or "",lv or "",
                      qty,uc,ftg,lf_,bf_,amt,amtt,
                      ci_item.get("action",""),ci_item.get("matched_po","")]
            fmts=[None,None,None,None,None,None,None,IFMT,MFMT,NFMT,NFMT,NFMT,MFMT,MFMT,None,None]
            for ci,(v,f) in enumerate(zip(row_vals,fmts),1):
                c=ws2.cell(row=cr,column=ci,value=v)
                c.font=NFONT; c.fill=fill; c.border=_bdr()
                c.alignment=_A(horizontal="left" if ci in (4,15,16) else "center",vertical="center")
                if f: c.number_format=f
            cr+=1
    for ci,w in enumerate([8,12,10,34,6,6,6,10,11,12,10,10,14,14,16,28],1):  # type: ignore[assignment]
        ws2.column_dimensions[_gcl(ci)].width=w
    ws2.freeze_panes="A2"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3: INVOICES  (matching reference exactly)
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Invoices")
    ws3.row_dimensions[1].height=36
    inv_hdrs=["Invoice #","Date","Description","Type","T","W","Length",
              "Qty Delivered","Delivered LF","Delivered BF/SF",
              "Delivered Cost","Delivered Cost+Tax"]
    for ci,h in enumerate(inv_hdrs,1):
        hdr(ws3,1,ci,h,wrap=True)

    ir=2
    for inv in invoices:
        inv_no=inv["invoice_no"]; inv_date=inv.get("date","")
        for item in items_list:
            qty_del=n(item.get("deliveries",{}).get(inv_no,0))
            if qty_del==0: continue
            typ=item.get("type","")
            fill=TYPE_FILLS.get(typ,_PF("solid",start_color="F8F8F8"))
            tv=n(item.get("t_num",0)); wv=n(item.get("w_num",0)); lv=n(item.get("length_num",0))
            uc=n(item.get("unit_cost",0))
            dlf  = compute_lf(typ,qty_del,tv,wv,lv)
            dbf  = compute_bf_sf(typ,qty_del,tv,wv,lv)
            dcost= compute_cost(typ,qty_del,tv,wv,lv,uc)
            dctax= round(dcost*(1+del_tax),2)
            row_vals=[inv_no,inv_date,item.get("description",""),typ,
                      tv or "",wv or "",lv or "",qty_del,dlf,dbf,dcost,dctax]
            fmts=[None,None,None,None,None,None,None,IFMT,NFMT,NFMT,MFMT,MFMT]
            for ci,(v,f) in enumerate(zip(row_vals,fmts),1):
                c=ws3.cell(row=ir,column=ci,value=v)
                c.font=NFONT; c.fill=fill; c.border=_bdr()
                c.alignment=_A(horizontal="left" if ci==3 else "center",vertical="center")
                if f: c.number_format=f
            ir+=1
    for ci,w in enumerate([14,12,36,10,6,6,8,12,12,12,14,14],1):  # type: ignore[assignment]
        ws3.column_dimensions[_gcl(ci)].width=w
    ws3.freeze_panes="A2"

    wb.save(str(project_xlsx(project)))



def _build_po_excel(project, items, meta):
    """Standalone PO report Excel with per-project PO tax rate."""
    C = get_column_letter
    po_tax = n(meta.get("po_tax_rate", meta.get("tax_rate", TAX_RATE)))
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None  # type: ignore
    ws.title = "PO Report"
    headers = ["Type","Description","T","W","Length","PO Qty","Unit Cost",
               "LF","BF/SF","Total Cost","Total Cost+Tax"]
    ws.row_dimensions[1].height = 30
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font=HEADER_FONT; cell.fill=HEADER_FILL
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=BORDER

    row = 2
    TYPE_ORDER = ["Lumber","LVL","Each","Panels","Unknown"]
    for typ in TYPE_ORDER:
        type_items = [i for i in items if i.get("type")==typ and n(i.get("po_qty",0))>0]
        if not type_items: continue
        ds = row
        for item in type_items:
            fill = TYPE_FILLS.get(typ, PatternFill())
            tv=n(item.get("t_num",0)); wv=n(item.get("w_num",0)); lv=n(item.get("length_num",0))
            uc=n(item.get("unit_cost",0)); qty=n(item.get("po_qty",0))
            lf=compute_lf(typ,qty,tv,wv,lv); bf=compute_bf_sf(typ,qty,tv,wv,lv)
            cost=compute_cost(typ,qty,tv,wv,lv,uc)
            row_vals=[typ,item.get("description",""),tv or "",wv or "",lv or "",qty,uc,lf,bf,cost,round(cost*(1+po_tax),2)]
            for ci,val in enumerate(row_vals,1):
                cell=ws.cell(row=row,column=ci,value=val)
                cell.fill=fill; cell.font=NORMAL_FONT; cell.border=BORDER
                cell.alignment=Alignment(horizontal="left" if ci==2 else "center",vertical="center")
                if ci in (10,11): cell.number_format=MONEY_FMT
                elif ci==7: cell.number_format=DEC_FMT
                elif ci in (6,8,9): cell.number_format=DEC_FMT
            row += 1
        # subtotal
        de=row-1; sfill=SUBTOTAL_FILLS.get(typ,SUBTOTAL_FILLS["Unknown"])
        sd={1:f"{typ.upper()} SUBTOTAL"}
        for ci in [6,8,9,10,11]: sd[ci]=f"=SUM({C(ci)}{ds}:{C(ci)}{de})"
        for ci in range(1,12):
            cell=ws.cell(row=row,column=ci,value=sd.get(ci,""))
            cell.fill=sfill; cell.font=SUBTOT_FONT; cell.border=BORDER
            cell.alignment=Alignment(horizontal="left" if ci==1 else "center",vertical="center")
            if ci in (10,11): cell.number_format=MONEY_FMT
        row += 1

    for ci,w in enumerate([10,34,7,6,7,9,11,12,12,14,14],1):  # type: ignore[assignment]
        ws.column_dimensions[C(ci)].width=w
    ws.freeze_panes="B2"
    wb.save(str(project_po_xlsx(project)))


def _build_co_excel(project, meta):
    """Standalone CO report Excel with subtotals per CO."""
    C = get_column_letter
    tax_rate = n(meta.get("tax_rate", TAX_RATE))
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None  # type: ignore
    ws.title = "CO Report"
    cos = meta.get("change_orders",[])
    headers = ["CO #","CO Date","Type","Description","T","W","Len",
               "Qty Change","Unit Cost","LF","BF/SF","Cost","Cost+Tax","Action","Matched PO Item"]
    ws.row_dimensions[1].height = 30
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font=HEADER_FONT; cell.fill=HEADER_FILL
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=BORDER
    cr = 2
    grand_start = 2
    for co in cos:
        co_start = cr
        for ci_item in co.get("items",[]):
            fill=TYPE_FILLS.get(ci_item.get("type","Unknown"),TYPE_FILLS["Unknown"])
            qty=n(ci_item.get("qty_change",0))
            tv=n(ci_item.get("t_num",0)); wv=n(ci_item.get("w_num",0)); lv=n(ci_item.get("length_num",0))
            typ=ci_item.get("type",""); uc=n(ci_item.get("unit_cost",0))
            abs_qty=abs(qty)
            sign=-1 if qty<0 else 1
            lf=compute_lf(typ,abs_qty,tv,wv,lv)*sign
            bfsf=compute_bf_sf(typ,abs_qty,tv,wv,lv)*sign
            cost=compute_cost(typ,abs_qty,tv,wv,lv,uc)*sign
            row_vals=[co.get("co_no",""),co.get("date",""),typ,ci_item.get("description",""),
                      tv or "",wv or "",lv or "",qty,uc,lf,bfsf,cost,round(cost*(1+tax_rate),2),
                      ci_item.get("action",""),ci_item.get("matched_po","")]
            for ci,val in enumerate(row_vals,1):
                cell=ws.cell(row=cr,column=ci,value=val)
                cell.fill=fill; cell.font=NORMAL_FONT; cell.border=BORDER
                cell.alignment=Alignment(horizontal="left" if ci in (4,14,15) else "center",vertical="center")
                if ci in (9,12,13): cell.number_format=MONEY_FMT
                elif ci==8: cell.number_format="#,##0"
                elif ci in (10,11): cell.number_format=DEC_FMT
            cr += 1
        # CO subtotal row
        if cr > co_start:
            co_end = cr-1
            sfill = PatternFill("solid", start_color="1B3A5C")
            sfont = Font(bold=True, color="FFFFFF", name="Arial", size=9)
            for ci in range(1,16):
                val=""
                if ci==1: val=f"CO #{co.get('co_no','')} SUBTOTAL"
                elif ci in (8,10,11,12,13):
                    val=f"=SUM({C(ci)}{co_start}:{C(ci)}{co_end})"
                cell=ws.cell(row=cr,column=ci,value=val)
                cell.fill=sfill; cell.font=sfont; cell.border=BORDER
                cell.alignment=Alignment(horizontal="left" if ci==1 else "center",vertical="center")
                if ci in (12,13): cell.number_format=MONEY_FMT
                elif ci in (10,11): cell.number_format=DEC_FMT
            cr += 1
            # spacer
            cr += 1
    # Grand total
    if cr > 2:
        gtfill=GRAND_FILL; gtfont=GRAND_FONT
        for ci in range(1,16):
            val=""
            if ci==1: val="GRAND TOTAL"
            elif ci in (12,13): val=f"=SUM({C(ci)}2:{C(ci)}{cr-2})"
            cell=ws.cell(row=cr,column=ci,value=val)
            cell.fill=gtfill; cell.font=gtfont; cell.border=BORDER
            cell.alignment=Alignment(horizontal="left" if ci==1 else "center",vertical="center")
            if ci in (12,13): cell.number_format=MONEY_FMT

    for ci,w in enumerate([8,12,10,34,6,6,6,10,11,12,12,14,14,16,28],1):  # type: ignore[assignment]
        ws.column_dimensions[C(ci)].width=w
    ws.freeze_panes="A2"
    wb.save(str(project_co_xlsx(project)))


def build_summary_excel():
    """Build cross-project summary Excel."""
    C = get_column_letter
    projects = _db.list_projects()
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None  # type: ignore
    ws.title = "All Projects Summary"

    headers = ["Project","Total Items","PO+CO Qty","Total Cost+Tax",
               "Total Delivered","Delivered Cost+Tax",
               "Remaining Qty","Remaining Cost+Tax",
               "# Invoices","# Change Orders","% Delivered"]
    ws.row_dimensions[1].height = 30
    for ci,h in enumerate(headers,1):
        cell=ws.cell(row=1,column=ci,value=h)
        cell.font=HEADER_FONT; cell.fill=HEADER_FILL
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=BORDER

    row=2
    alt_fill = PatternFill("solid", start_color="F0F4FF")
    for proj in sorted(projects):
        items=load_items(proj); meta=load_meta(proj)
        total_qty=sum(n(i.get("po_qty",0))+n(i.get("co_qty",0)) for i in items)
        total_del=sum(sum(i.get("deliveries",{}).values()) for i in items)
        total_cost=sum(item_fields(i)[2] for i in items)
        total_del_cost=sum(item_fields(i, sum(i.get("deliveries",{}).values()))[2] for i in items)
        rem_qty=total_qty-total_del
        rem_cost=total_cost-total_del_cost
        pct=round(total_del/total_qty*100,1) if total_qty else 0
        fill = alt_fill if row%2==0 else PatternFill()
        row_vals=[proj,len(items),total_qty,round(total_cost*(1+TAX_RATE),2),
                  total_del,round(total_del_cost*(1+TAX_RATE),2),
                  rem_qty,round(rem_cost*(1+TAX_RATE),2),
                  len(meta.get("invoices",[])),meta.get("co_count",0),pct]
        for ci,val in enumerate(row_vals,1):
            cell=ws.cell(row=row,column=ci,value=val)
            cell.font=NORMAL_FONT; cell.border=BORDER
            cell.alignment=Alignment(horizontal="left" if ci==1 else "center",vertical="center")
            if ci in (4,6,8): cell.number_format=MONEY_FMT
            elif ci in (3,5,7): cell.number_format=INT_FMT
            elif ci==11: cell.number_format='0.0"%"'
        row+=1

    # Grand Total row
    if row > 2:
        ws.row_dimensions[row].height=20
        for ci in range(1,12):
            val=""
            if ci==1: val="GRAND TOTAL"
            elif ci in (3,5,7): val=f"=SUM({C(ci)}2:{C(ci)}{row-1})"
            elif ci in (4,6,8): val=f"=SUM({C(ci)}2:{C(ci)}{row-1})"
            elif ci in (9,10):  val=f"=SUM({C(ci)}2:{C(ci)}{row-1})"
            cell=ws.cell(row=row,column=ci,value=val)
            cell.fill=GRAND_FILL; cell.font=GRAND_FONT; cell.border=BORDER
            cell.alignment=Alignment(horizontal="left" if ci==1 else "center",vertical="center")
            if ci in (4,6,8): cell.number_format=MONEY_FMT
            elif ci in (3,5,7): cell.number_format=INT_FMT

    for ci,w in enumerate([24,12,12,18,14,18,14,18,12,14,12],1):  # type: ignore[assignment]
        ws.column_dimensions[C(ci)].width=w
    ws.freeze_panes="B2"

    # Per-project sheets
    for proj in sorted(projects):
        ws_p = wb.create_sheet(proj[:28])  # Excel sheet name limit
        items=load_items(proj); meta=load_meta(proj)
        p_hdrs=["Type","Description","PO Qty","CO Qty","PO+CO Qty",
                "Unit Cost","Total Cost+Tax","Total Delivered","Delivered Cost+Tax",
                "Remaining","Remaining Cost+Tax","% Delivered"]
        for ci,h in enumerate(p_hdrs,1):
            cell=ws_p.cell(row=1,column=ci,value=h)
            cell.font=HEADER_FONT; cell.fill=HEADER_FILL
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            cell.border=BORDER
        pr=2
        for item in items:
            typ=item.get("type","")
            fill=TYPE_FILLS.get(typ,TYPE_FILLS["Unknown"])
            pq=n(item.get("po_qty",0)); cq=n(item.get("co_qty",0)); tq=pq+cq
            td=sum(item.get("deliveries",{}).values())
            _,_,tc=item_fields(item,tq); _,_,dc=item_fields(item,td)
            rem=tq-td; _,_,rc=item_fields(item,rem)
            pct=round(td/tq*100,1) if tq else 0
            row_vals=[typ,item.get("description",""),pq,cq,tq,
                      n(item.get("unit_cost",0)),round(tc*(1+TAX_RATE),2),
                      td,round(dc*(1+TAX_RATE),2),rem,round(rc*(1+TAX_RATE),2),pct]
            for ci,val in enumerate(row_vals,1):
                cell=ws_p.cell(row=pr,column=ci,value=val)
                cell.fill=fill; cell.font=NORMAL_FONT; cell.border=BORDER
                cell.alignment=Alignment(horizontal="left" if ci==2 else "center",vertical="center")
                if ci in (7,9,11): cell.number_format=MONEY_FMT
                elif ci in (3,4,5,8,10): cell.number_format=INT_FMT
                elif ci==12: cell.number_format='0.0"%"'
            pr+=1
        for ci,w in enumerate([10,34,10,10,10,12,16,14,16,12,16,12],1):  # type: ignore[assignment]
            ws_p.column_dimensions[C(ci)].width=w
        ws_p.freeze_panes="B2"

    path = BASE_DIR / "summary_all_projects.xlsx"
    wb.save(str(path))
    return path


# ── API Routes ────────────────────────────────────────────────────────────────
@app.get("/projects/poll")
def poll_projects(since: str = ""):
    """
    Lightweight endpoint for the frontend's background sync.
    Returns the last-updated timestamp for every project (or one project).
    The frontend calls this every 15 seconds and only calls silentRefresh()
    if the timestamp for the current project has advanced since last check.
    This means: 1 tiny DB query every 15s per browser tab — no wasted work.
    """
    timestamps = _db.all_projects_last_updated()
    return {"timestamps": timestamps, "server_time": datetime.utcnow().isoformat()}

@app.get("/projects/{project}/poll")
def poll_single_project(project: str):
    """Poll a single project's last-updated time."""
    ts = _db.project_last_updated(project)
    return {"project": project, "last_updated": ts, "server_time": datetime.utcnow().isoformat()}

@app.get("/projects")
def list_projects():
    return sorted(_db.list_projects())

@app.put("/projects/{project}/meta")
def update_project_meta(project: str, body: dict):
    """Update project metadata: gsf, state, building_type, architect, completed"""
    meta = load_meta(project)
    for field in ["gsf", "state", "building_type", "architect", "completed"]:
        if field in body:
            meta[field] = body[field]
    save_meta(project, meta)
    return {"status": "ok", "meta": meta}

@app.post("/projects/{project}/create")
def create_project(project: str, body: dict | None = None):
    _db.create_project(project)
    if body:
        meta = load_meta(project)
        meta["gsf"]           = body.get("gsf", 0)
        meta["state"]         = body.get("state", "")
        meta["building_type"] = body.get("building_type", "")
        meta["architect"]     = body.get("architect", "")
        meta["project_name"]  = body.get("project_name", project)
        save_meta(project, meta)
    return {"status": "created", "project": project}

@app.get("/projects/{project}/status")
def project_status(project: str):
    items=load_items(project); meta=load_meta(project)
    return {"project":project,"has_po":len(items)>0,"item_count":len(items),
            "invoice_count":len(meta.get("invoices",[])),"co_count":meta.get("co_count",0),
            "has_excel":len(items)>0,"invoices":meta.get("invoices",[])}

@app.delete("/projects/{project}")
def delete_project(project: str):
    if not _db.project_exists(project):
        return {"status": "deleted", "message": "Project not found (already deleted)"}
    _db.delete_project(project)
    return {"status": "deleted", "message": f"Project '{project}' deleted"}

@app.delete("/projects/{project}/po")
def delete_po(project: str):
    save_items(project,[])
    save_meta(project,{"invoices":[],"co_count":0,"change_orders":[],"unmatched_items":[]})
    xlsx=project_xlsx(project)
    if xlsx.exists(): xlsx.unlink()
    return {"status":"po_deleted","message":"All data cleared. Upload a new PO."}

@app.delete("/projects/{project}/co/{co_no}")
def delete_co(project: str, co_no: str):
    items=load_items(project); meta=load_meta(project)
    cos=meta.get("change_orders",[])
    co=next((c for c in cos if c.get("co_no")==co_no), None)
    if not co: raise HTTPException(404,f"CO #{co_no} not found.")
    for ci_item in co.get("items",[]):
        mp=ci_item.get("matched_po","")
        qty_change=n(ci_item.get("qty_change",0))
        action=ci_item.get("action","")
        if action=="update_po_qty" and mp:
            for item in items:
                if item.get("description")==mp:
                    item["co_qty"]=n(item.get("co_qty",0)) - qty_change
                    break
        else:
            # Remove CO-added items
            desc_base=ci_item.get("description","")
            items=[i for i in items if not (
                i.get("source")=="co" and
                i.get("description","").startswith(desc_base.split(f" (CO#{co_no})")[0]) and
                i.get("description","").endswith(f"(CO#{co_no})")
            )]
    meta["change_orders"]=[c for c in cos if c.get("co_no")!=co_no]
    meta["co_count"]=max(0,meta.get("co_count",0)-1)
    save_items(project,items); save_meta(project,meta); rebuild_excel(project)
    return {"status":"co_deleted","message":f"CO #{co_no} removed and quantities reversed."}

@app.delete("/projects/{project}/invoice/{invoice_no}")
def delete_invoice(project: str, invoice_no: str):
    items=load_items(project); meta=load_meta(project)
    invoices=meta.get("invoices",[])
    if invoice_no not in [i["invoice_no"] for i in invoices]:
        raise HTTPException(404,f"Invoice #{invoice_no} not found.")
    for item in items:
        item.get("deliveries",{}).pop(invoice_no,None)
    meta["invoices"]=[i for i in invoices if i["invoice_no"]!=invoice_no]
    # Also remove unmatched items for this invoice
    meta["unmatched_items"]=[u for u in meta.get("unmatched_items",[]) if u.get("invoice_no")!=invoice_no]
    save_items(project,items); save_meta(project,meta); rebuild_excel(project)
    return {"status":"invoice_deleted","message":f"Invoice #{invoice_no} deleted."}

# ── PO upload ─────────────────────────────────────────────────────────────────
@app.delete("/projects/{project}/po/{category}")
def delete_po_category(project: str, category: str):
    """Remove all items from a specific PO category (lumber/housewrap/siding)."""
    items   = load_items(project)
    meta    = load_meta(project)
    before  = len(items)
    # Keep items NOT from this category (or items with no category tag = legacy)
    # For "lumber" category, also keep items with no po_category tag (backward compat)
    if category == "lumber":
        # Remove items tagged as lumber, keep others
        remaining = [it for it in items if it.get("po_category","lumber") not in ("lumber",)]
        # Also remove untagged items only if we're deleting lumber
        remaining = [it for it in items if it.get("po_category") not in (category, None, "")]
    else:
        remaining = [it for it in items if it.get("po_category") != category]
    removed = before - len(remaining)
    save_items(project, remaining)
    # Clear the po_files entry for this category
    po_files = meta.get("po_files", {})
    po_files.pop(category, None)
    meta["po_files"] = po_files
    save_meta(project, meta)
    try: rebuild_excel(project)
    except Exception as e: print(f"[WARN] rebuild after PO delete: {e}")
    return {"status":"ok","removed":removed,
            "message":f"Removed {removed} items from {category} PO"}


@app.post("/projects/{project}/complete")
def toggle_project_complete(project: str, body: dict | None = None):
    """Toggle project complete/active. Completed projects are used for estimation."""
    meta = load_meta(project)
    currently = meta.get("completed", False)
    new_state  = not currently
    if body and "completed" in body:
        new_state = bool(body["completed"])
    meta["completed"] = new_state
    if new_state and not meta.get("completed_date"):
        from datetime import date
        meta["completed_date"] = str(date.today())
    save_meta(project, meta)
    return {"status": "ok", "project": project, "completed": new_state,
            "message": f"Project marked {'✅ Complete' if new_state else '🔄 Active'}"}


@app.post("/projects/{project}/upload-po")
async def upload_po(project: str, file: UploadFile = File(...),
                   po_category: str = Form("lumber")):
    """Upload a PO and APPEND its items to existing items (does not overwrite).
    po_category: 'lumber' | 'housewrap' | 'siding'
    Items are deduplicated by type+description+unit_cost so re-uploading is safe.
    """
    _db.create_project(project)
    pdf_path = UPLOAD_DIR / f"{project}_PO_{po_category}_{file.filename}"
    pdf_path.write_bytes(await file.read())

    text = pdf_to_text(pdf_path)
    if not text.strip():
        raise HTTPException(400, "Could not extract text from this PDF.")

    items_raw = parse_po_with_claude(text, po_category)
    if not items_raw:
        raise HTTPException(422, "No items found in this PO PDF.")

    # Build new items list from parsed data
    new_items = []
    for it in items_raw:
        it["po_qty"]      = n(it.get("qty", 0))
        it["co_qty"]      = 0.0
        it["deliveries"]  = {}
        it["leftover_lf"] = 0.0
        it["source"]      = "po"
        it["po_category"] = po_category   # tag which PO this came from
        new_items.append(normalize_item(it))

    # Load existing items
    existing = load_items(project)
    src_filename = file.filename

    # ADDITIVE UPLOAD: Only clear items from the SAME source file (same filename).
    # Items from OTHER PO files (even in same category) are PRESERVED.
    # This allows: Lumber PO + LVL PO + Each PO to all coexist.
    # Re-uploading the same file replaces only that file's items.
    existing = [it for it in existing if it.get("po_file_src","") != src_filename]

    # Tag each new item with its source filename
    for item in new_items:
        item["po_file_src"] = src_filename

    for item in new_items:
        existing.append(item)
    added = len(new_items); skipped = 0

    save_items(project, existing)
    meta = load_meta(project)
    # Track which PO files have been uploaded
    po_files = meta.get("po_files", {})
    po_files[po_category] = file.filename
    meta["po_files"] = po_files
    meta["po_file"]  = file.filename   # backward compat
    save_meta(project, meta)
    try: rebuild_excel(project)
    except Exception as e: print(f"[WARN] rebuild after PO: {e}")

    skip_msg = f" ({skipped} duplicates skipped)" if skipped else ""
    return {"status": "ok", "items_parsed": len(new_items), "items_added": added,
            "message": f"{po_category.capitalize()} PO uploaded: {added} items added{skip_msg}"}

# ── CO: single-step upload (mirrors invoice upload logic exactly) ─────────────
@app.post("/projects/{project}/upload-co")
async def upload_co(project: str, file: UploadFile = File(...),
                    co_no: str = Form(""), co_date: str = Form("")):
    items = load_items(project)
    if not items:
        raise HTTPException(400, "Upload a PO first before adding Change Orders.")

    meta      = load_meta(project)
    co_number = co_no.strip() or str(meta.get("co_count", 0) + 1).zfill(3)
    co_date_  = co_date or datetime.now().strftime("%m/%d/%Y")

    # Save PDF
    pdf_path = UPLOAD_DIR / f"{project}_CO_{co_number}_{file.filename}"
    pdf_path.write_bytes(await file.read())

    # Extract text
    text = pdf_to_text(pdf_path)
    if not text.strip():
        raise HTTPException(400, "Could not extract text from the PDF. Is it a scanned image?")

    # Parse with Claude (extraction only)
    co_items_raw = parse_co_with_claude(text, items)
    if not co_items_raw:
        raise HTTPException(422,
            "Claude could not find any line items in this CO PDF. "
            "Please verify it is a Matheus Lumber Change Order document.")

    matched_c = 0
    new_c     = 0
    logged    = []

    for co in co_items_raw:
        qty_change = n(co.get("qty_change", 0))
        if qty_change == 0:
            continue                          # skip zero-qty lines

        co_uc   = n(co.get("unit_cost", 0))
        co_t    = n(co.get("t_num",    0))
        co_w    = n(co.get("w_num",    0))
        co_len  = n(co.get("length_num",0))
        co_type = co.get("type", "Unknown")
        co_desc = co.get("description", "")
        co_uom  = str(co.get("uom", "PC")).upper().strip()

        # Fix panel dimensions — Claude often reads "4X8" as T=4, W=4, L=8
        # Correct: t=fractional thickness, w=8, length=0
        if co_type == "Panels":
            co_w   = 8
            co_len = 0
            if co_t >= 1.0 or co_t == 0.0:
                import re as _re
                TMAP = {"7/16":0.4375,"15/32":0.46875,"1/2":0.5,
                        "19/32":0.59375,"5/8":0.625,"23/32":0.71875,"3/4":0.75}
                co_t = next((v for k,v in TMAP.items() if k in co_desc), 0.46875)

        # Fix panel dimensions at the matching stage too
        if co_type == "Panels":
            co_w   = 8    # always 8 for standard 4x8 sheets
            co_len = 0    # panels have no "length"
            if co_t >= 1.0 or co_t == 0.0:
                # Parse thickness from description
                import re as _re
                TMAP = {"7/16":0.4375,"15/32":0.46875,"1/2":0.5,
                        "19/32":0.59375,"5/8":0.625,"23/32":0.71875,"3/4":0.75}
                co_t = next((v for k,v in TMAP.items() if k in co_desc), 0.46875)
        # For LF-ordered items, qty_change is already in LF
        # Convert to piece count for matching if length_num > 1
        # but store original LF value in co_qty_pieces for the item
        is_lf_ordered = co_uom == "LF"

        # ── Matching rules ────────────────────────────────────────────────
        # Match CO item to PO item when:
        #   1. Same type
        #   2. For PANELS: t_num is NOT used (panels t_num=4 in PO but may be
        #      fractional from CO PDF). Match on w_num + length_num + unit_cost.
        #      The fractional thickness (15/32, 23/32) is encoded in unit_cost:
        #      different spec = different price, so cost match is sufficient.
        #   3. For LUMBER/LVL: t_num + w_num + length_num must all match.
        #   4. Unit cost within $1.00 (prevents same-dimension diff-spec merges)
        #   Also search CO-added items (source=="co") so re-uploads don't duplicate.

        COST_TOLERANCE = 1.00   # max $1 difference to be "same item"

        matched_item = None
        # Search PO items AND previously-added CO items
        all_candidates = items  # items already includes previous CO additions

        for item in all_candidates:
            item_type = item.get("type", "")
            type_ok = (item_type == co_type) or co_type == "Unknown"
            if not type_ok:
                continue

            cost_ok = abs(n(item.get("unit_cost", 0)) - co_uc) <= COST_TOLERANCE

            if item_type == "Panels":
                # For panels: match on unit_cost only (each unique panel spec has unique price)
                # w_num and length_num also checked for safety
                w_ok   = abs(n(item.get("w_num",      0)) - co_w)   < 0.002
                len_ok = abs(n(item.get("length_num", 0)) - co_len) < 0.002
                if cost_ok and w_ok and len_ok:
                    matched_item = item
                    break
            else:
                # For Lumber/LVL/Each: full dimension match required
                t_ok   = abs(n(item.get("t_num",      0)) - co_t)   < 0.002
                w_ok   = abs(n(item.get("w_num",      0)) - co_w)   < 0.002
                len_ok = abs(n(item.get("length_num", 0)) - co_len) < 0.002
                if type_ok and t_ok and w_ok and len_ok and cost_ok:
                    matched_item = item
                    break

        if matched_item:
            # ── Update existing PO item's co_qty ──────────────────────────
            matched_item["co_qty"] = n(matched_item.get("co_qty", 0)) + qty_change
            logged.append({
                **co,
                "action":     "update_po_qty",
                "matched_po": matched_item["description"],
            })
            matched_c += 1
        else:
            # ── Add as new CO line item ────────────────────────────────────
            new_item = normalize_item({
                "type":        co_type if co_type != "Unknown" else "Lumber",
                "description": co_desc + f" (CO#{co_number})",
                "t_num":       co_t,
                "w_num":       co_w,
                "length_num":  co_len,
                "po_qty":      0.0,
                "co_qty":      qty_change,
                "unit_cost":   co_uc,
                "deliveries":  {},
                "leftover_lf": 0.0,
                "source":      "co",
            })
            items.append(new_item)
            logged.append({
                **co,
                "action":     "add_new",
                "matched_po": "NEW",
            })
            new_c += 1

    if not logged:
        raise HTTPException(422, "All CO items had zero quantity — nothing to save.")

    # Save everything — same pattern as invoice upload
    meta.setdefault("change_orders", []).append({
        "co_no":  co_number,
        "date":   co_date_,
        "file":   file.filename,
        "items":  logged,
    })
    meta["co_count"] = meta.get("co_count", 0) + 1
    save_items(project, items)
    save_meta(project, meta)
    try:
        rebuild_excel(project)
    except Exception as e:
        print(f"[WARN] rebuild_excel after CO upload failed: {e}")

    return {
        "status":    "ok",
        "co_no":     co_number,
        "matched":   matched_c,
        "new_items": new_c,
        "items":     logged,
        "message":   (f"CO #{co_number}: {matched_c} PO item(s) updated, "
                      f"{new_c} new item(s) added."),
    }



@app.post("/projects/{project}/preview-invoice")
async def preview_invoice(project: str, file: UploadFile = File(...)):
    """Read invoice PDF and extract invoice number and date."""
    data = await file.read()
    tmp_path = UPLOAD_DIR / f"inv_preview_{file.filename}"
    tmp_path.write_bytes(data)
    text = pdf_to_text(tmp_path)
    inv_no, inv_date = extract_invoice_header(text)
    return {"status":"ok","invoice_no":inv_no,"invoice_date":inv_date,
            "filename":file.filename}

# ── Invoice upload ────────────────────────────────────────────────────────────
@app.post("/projects/{project}/upload-invoice")
async def upload_invoice(project: str, file: UploadFile=File(...),
                         invoice_no: str=Form(""), invoice_date: str=Form("")):
    items=load_items(project)
    if not items: raise HTTPException(400,"Upload a PO first.")
    if not invoice_no.strip():
        raise HTTPException(400,"Invoice number is required. Please enter it manually or wait for auto-read.")
    invoice_no = invoice_no.strip()
    meta=load_meta(project); invoices=meta.get("invoices",[])
    if invoice_no in [i["invoice_no"] for i in invoices]:
        raise HTTPException(400,f"Invoice #{invoice_no} already recorded.")
    pdf_path=UPLOAD_DIR/f"{project}_INV_{invoice_no}_{file.filename}"
    pdf_path.write_bytes(await file.read())
    inv_date_str=invoice_date or datetime.now().strftime("%m/%d/%Y")
    deliveries=parse_invoice_with_claude(pdf_to_text(pdf_path),items)
    matched,unmatched_count=0,0
    new_unmatched=[]

    for d in deliveries:
        desc    = d.get("matched_description", "UNMATCHED")
        qty_pcs = n(d.get("qty_delivered_pcs", d.get("qty_delivered", 0)))  # pieces
        inv_len = n(d.get("invoice_length", 0))
        inv_t   = n(d.get("invoice_t", 0))
        inv_w   = n(d.get("invoice_w", 0))
        inv_uc  = n(d.get("invoice_unit_cost", 0))
        inv_uom = str(d.get("invoice_uom", "PC")).upper().strip()
        is_rl   = bool(d.get("rl_match", False))

        if qty_pcs == 0:
            continue

        # ── Determine what value to store in deliveries ──────────────────────
        # For R/L lumber items (length_num==1), the PO tracks LINEAR FEET.
        # So when an invoice delivers 192 PC of 2X8X14, we store 192*14 = 2,688 LF.
        # For Panels, Each, Siding, HouseWrap: store piece count.

        def get_store_qty(po_item, pcs, length, uom="PC"):
            """Return the quantity to store in deliveries dict.
            For LVL and R/L Lumber, deliveries store LINEAR FEET.
            
            UOM=LF → qty already in LF → store directly (no conversion)
            UOM=PC + LVL → convert: pcs × po_length = LF stored
            UOM=PC + R/L Lumber (length_num=1) + inv has specific length → pcs × inv_length
            UOM=PC + exact-length Lumber/Panels/Each → store as pieces
            """
            typ    = po_item.get("type", "")
            po_len = n(po_item.get("length_num", 0))
            uom_up = uom.upper().strip()

            if typ in ("Lumber", "LVL"):
                if uom_up == "LF":
                    # Invoice already gives LF — store directly
                    return round(pcs, 4)
                elif po_len == 1.0 and length > 1:
                    # R/L Lumber: convert pieces × invoice_length → LF
                    return round(pcs * length, 4)
                elif typ == "LVL" and po_len > 1:
                    # Specific-length LVL: pcs × po_length → LF
                    # (invoice might say "120LF" already handled above)
                    return round(pcs * po_len, 4)
            return pcs   # Panels, Each, Siding, HouseWrap, exact lumber → pieces

        if desc == "UNMATCHED":
            # ── Python-side fallback matching ─────────────────────────────────
            fallback_match = None

            # 1. R/L Lumber: same T+W+cost, length_num=1
            for item in items:
                cost_diff = abs(n(item.get("unit_cost", 0)) - inv_uc) if inv_uc > 0 else 0
                if (item.get("type") in ("Lumber", "LVL") and
                    abs(n(item.get("t_num", 0)) - inv_t) < 0.001 and
                    abs(n(item.get("w_num", 0)) - inv_w) < 0.001 and
                    n(item.get("length_num", 0)) == 1.0 and
                    cost_diff <= 1.5):
                    fallback_match = item; break

            # 2. Each items: match by unit_cost within $0.50 (e.g. sill seals same price)
            if not fallback_match:
                inv_desc_lower = d.get("invoice_description","").lower()
                for item in items:
                    if item.get("type") != "Each": continue
                    cost_diff = abs(n(item.get("unit_cost", 0)) - inv_uc) if inv_uc > 0 else 99
                    if cost_diff <= 0.50:
                        fallback_match = item; break

            if fallback_match:
                store_qty = get_store_qty(fallback_match, qty_pcs, inv_len, inv_uom)
                fallback_match.setdefault("deliveries", {})[invoice_no] =                     n(fallback_match.get("deliveries", {}).get(invoice_no, 0)) + store_qty
                matched += 1
            else:
                # Truly unmatched — ALWAYS log for user to see and assign
                lf_val = round(qty_pcs * inv_len, 4) if inv_len > 1 else qty_pcs
                new_unmatched.append({
                    "invoice_no": invoice_no, "invoice_date": inv_date_str,
                    "description": d.get("invoice_description", "Not Identified"),
                    "t": inv_t, "w": inv_w, "length": inv_len,
                    "qty": qty_pcs, "unit_cost": inv_uc,
                    "computed_lf": lf_val,
                    "notes": f"No PO match found. Unit cost on invoice: ${inv_uc}. "
                             f"Please assign to a PO line item or add as new.",
                    "status": "Pending"
                })
                unmatched_count += 1
        else:
            # Claude gave us a matched description — find the item using multi-level matching:
            # Level 1: exact description + unit cost within $2
            # Level 2: normalized description (strip CO# suffixes) + unit cost within $2
            # Level 3: dimension match (T+W+Length) + unit cost within $2
            # Level 4: dimension match only (unit cost within $5 for close-enough prices)
            found_item = None

            # Build normalized invoice key from dimensions
            inv_dim_key = (round(inv_t,3), round(inv_w,3), round(inv_len,3))
            inv_desc_norm = norm_desc_key(d.get("invoice_description", desc))

            def uc_match(item, tolerance=2.0):
                po_uc = n(item.get("unit_cost",0))
                return inv_uc == 0 or abs(po_uc - inv_uc) <= tolerance

            # Level 1: exact desc match
            for item in items:
                if item["description"] == desc and uc_match(item, 2.0):
                    found_item = item; break

            # Level 2: normalized desc match (strips CO# etc.)
            if not found_item:
                for item in items:
                    if norm_desc_key(item.get("description","")) == inv_desc_norm and uc_match(item, 2.0):
                        found_item = item; break

            # Level 3: dimension match + unit cost within $2
            if not found_item and (inv_t or inv_w or inv_len):
                for item in items:
                    it_key = (round(n(item.get("t_num",0)),3),
                              round(n(item.get("w_num",0)),3),
                              round(n(item.get("length_num",0)),3))
                    if it_key == inv_dim_key and uc_match(item, 2.0):
                        found_item = item; break

            # Level 4: dimension match + unit cost within $5 (catches small price changes)
            if not found_item and (inv_t or inv_w or inv_len):
                for item in items:
                    it_key = (round(n(item.get("t_num",0)),3),
                              round(n(item.get("w_num",0)),3),
                              round(n(item.get("length_num",0)),3))
                    if it_key == inv_dim_key and uc_match(item, 5.0):
                        found_item = item; break

            if found_item:
                store_qty = get_store_qty(found_item, qty_pcs, inv_len, inv_uom)
                found_item.setdefault("deliveries", {})[invoice_no] =                     n(found_item.get("deliveries", {}).get(invoice_no, 0)) + store_qty
                matched += 1
            else:
                # Truly no match — log for manual assignment
                po_uc_found = n(found_item.get("unit_cost",0)) if found_item else 0
                lf_val = round(qty_pcs * inv_len, 4) if inv_len > 1 else qty_pcs
                # Check if a dim match exists but cost differs (to give better error message)
                dim_match_item = None
                if inv_t or inv_w or inv_len:
                    for item in items:
                        it_key=(round(n(item.get("t_num",0)),3),
                                round(n(item.get("w_num",0)),3),
                                round(n(item.get("length_num",0)),3))
                        if it_key == inv_dim_key:
                            dim_match_item = item; break
                if dim_match_item:
                    note = (f"Dimension match found (T={inv_t} W={inv_w} L={inv_len}) but "
                            f"unit cost differs: invoice ${inv_uc} vs PO "
                            f"${n(dim_match_item.get('unit_cost',0))}. "
                            f"Assign to correct line item.")
                else:
                    note = (f"No PO match: T={inv_t} W={inv_w} L={inv_len} "
                            f"@${inv_uc}/unit. Please assign manually.")
                new_unmatched.append({
                    "invoice_no": invoice_no, "invoice_date": inv_date_str,
                    "description": d.get("invoice_description", desc),
                    "t": inv_t, "w": inv_w, "length": inv_len,
                    "qty": qty_pcs, "unit_cost": inv_uc,
                    "computed_lf": lf_val,
                    "notes": note,
                    "status": "Pending"
                })
                unmatched_count += 1

    meta.setdefault("unmatched_items",[]).extend(new_unmatched)
    meta["invoices"]=invoices+[{"invoice_no":invoice_no,"date":inv_date_str,
        "file":file.filename,"matched":matched,"unmatched":unmatched_count}]
    save_meta(project,meta); save_items(project,items)
    # Rebuild Excel — wrapped so a build error never loses the data already saved
    try:
        rebuild_excel(project)
    except Exception as e:
        # Data is saved — just log the Excel error, don't fail the whole request
        print(f"[WARN] rebuild_excel after invoice upload failed: {e}")
    msg=f"Invoice #{invoice_no}: {matched} item(s) matched"
    if unmatched_count: msg+=f", ⚠ {unmatched_count} item(s) could not be matched — see Unmatched Items panel"
    return {"status":"ok","matched":matched,"unmatched":unmatched_count,"message":msg,
            "unmatched_items":new_unmatched}

# ── Assign unmatched item to a PO line ───────────────────────────────────────
@app.post("/projects/{project}/assign-unmatched")
async def assign_unmatched(project: str, body: dict):
    """
    Assign an unmatched invoice item to an existing PO line,
    OR add it as a brand-new invoice-based line item.
    """
    items=load_items(project); meta=load_meta(project)
    um_idx=body.get("unmatched_idx")
    invoice_no=body.get("invoice_no","")
    add_as_new=bool(body.get("add_as_new", False))
    unmatched=meta.get("unmatched_items",[])
    if um_idx is None or um_idx>=len(unmatched):
        raise HTTPException(400,"Invalid unmatched item index.")
    um=unmatched[um_idx]
    qty=n(um.get("qty",0))
    inv_len=n(um.get("length",0))
    inv_t=n(um.get("t",0)); inv_w=n(um.get("w",0)); inv_uc=n(um.get("unit_cost",0))

    if add_as_new:
        typ_guess="Lumber" if (inv_t and inv_w) else "Each"
        desc=um.get("description","Unknown Invoice Item")
        new_item=normalize_item({
            "type":       typ_guess,
            "description":desc+f" (INV#{invoice_no})",
            "t_num":inv_t,"w_num":inv_w,"length_num":inv_len,
            "po_qty":0.0,"co_qty":0.0,"unit_cost":inv_uc,
            "deliveries":{invoice_no:qty},
            "leftover_lf":0.0,"source":"invoice",
        })
        items.append(new_item)
        unmatched[um_idx]["status"]=f"Added as new: {desc}"
        meta["unmatched_items"]=unmatched
        save_items(project,items); save_meta(project,meta); rebuild_excel(project)
        return {"status":"ok","message":f"Added '{desc}' as new line with {qty} units."}

    target_desc=body.get("target_description","")
    if not target_desc:
        raise HTTPException(400,"Provide target_description or set add_as_new=true.")
    for item in items:
        if item.get("description")==target_desc:
            po_len=n(item.get("length_num",0))
            is_rl_po=(item.get("type") in ("Lumber","LVL") and abs(po_len-1.0)<0.001)
            store_qty=round(qty*inv_len,4) if (is_rl_po and inv_len>1) else qty
            item.setdefault("deliveries",{})[invoice_no]=\
                n(item.get("deliveries",{}).get(invoice_no,0))+store_qty
            unmatched[um_idx]["status"]=f"Assigned to: {target_desc}"
            meta["unmatched_items"]=unmatched
            save_items(project,items); save_meta(project,meta); rebuild_excel(project)
            return {"status":"ok","message":f"Assigned {store_qty} to '{target_desc}'"}
    raise HTTPException(404,"Target PO item not found.")

# ── Item CRUD (edit/add/delete) ───────────────────────────────────────────────
def is_rl_lumber(item):
    """True ONLY for Lumber items with length=1 (Random Length).
    Deliveries for R/L Lumber are stored in LF not pieces.
    LVL is NEVER included here — LVL always uses the else/item_fields path."""
    return (item.get("type") == "Lumber" and
            abs(n(item.get("length_num",0)) - 1.0) < 0.001)

def norm_desc_key(desc):
    """Normalize description for lookup: strip (CO#xxx)/(INV#xxx) suffixes."""
    import re
    d = (desc or "").strip().upper()
    d = re.sub(r'\s*\(CO#?\s*\d+\)\s*$', '', d)
    d = re.sub(r'\s*\(INV#?\s*\S+\)\s*$', '', d)
    d = re.sub(r'\s+', ' ', d).strip()
    return d


@app.get("/projects/{project}/items")
def get_items(project: str):
    items=load_items(project); meta=load_meta(project)
    po_tax   = n(meta.get("po_tax_rate",       meta.get("tax_rate", TAX_RATE)))
    del_tax  = n(meta.get("delivery_tax_rate", meta.get("tax_rate", TAX_RATE)))
    tax_rate = po_tax   # backward compat alias
    result=[]
    for idx,item in enumerate(items):
        deliveries = item.get("deliveries",{})
        typ = item.get("type","")
        tv  = n(item.get("t_num",0))
        wv  = n(item.get("w_num",0))
        lv  = n(item.get("length_num",0))
        uc  = n(item.get("unit_cost",0))
        total_qty = n(item.get("po_qty",0)) + n(item.get("co_qty",0))
        lf,bfsf,cost = item_fields(item, total_qty)

        # ── Delivery calculations ─────────────────────────────────────────────
        # For R/L items: deliveries[inv_no] stores LINEAR FEET (pcs × length).
        # For all others: deliveries[inv_no] stores PIECES.
        rl = is_rl_lumber(item)

        if rl:
            # deliveries values = LF already
            del_lf_total = sum(deliveries.values())          # total LF delivered
            # Pieces = LF / avg length — we don't know actual lengths, so we
            # track "delivered" in LF units for R/L (total_delivered = LF)
            total_del = del_lf_total                         # show LF in delivered col
            dlf       = del_lf_total
            # BF from delivered LF: BF = LF × T × W / 12
            dbfsf = round((dlf * tv * wv) / 12, 4) if (tv and wv) else 0
            dcost = round((dbfsf * uc) / 1000, 2)
            rem   = lf - dlf                                 # remaining LF
            rlf   = rem
            rbfsf = round((rlf * tv * wv) / 12, 4) if (tv and wv and rlf > 0) else 0
            rcost = round((rbfsf * uc) / 1000, 2)
        else:
            # Piece-based items
            total_del       = sum(deliveries.values())
            dlf,dbfsf,dcost = item_fields(item, total_del)
            rem             = total_qty - total_del
            rlf,rbfsf,rcost = item_fields(item, rem)

        # ── Progress % ───────────────────────────────────────────────────────
        # For R/L: compare delivered LF vs total LF
        denom = lf if rl and lf else total_qty
        pct   = round(total_del / denom * 100, 1) if denom else 0

        # ── Used on site ─────────────────────────────────────────────────────
        lo_lf    = n(item.get("leftover_lf",0))
        used_lf  = max(0.0, dlf - lo_lf)
        if typ == "Lumber" and tv and wv:
            # used_lf is in linear feet; convert to BF then cost
            used_bf   = round((used_lf * tv * wv) / 12, 4)
            used_cost = round((used_bf * uc) / 1000, 2)
        elif typ == "LVL":
            used_cost = round(used_lf * uc, 2)
        elif typ in ("Panels","Siding","HouseWrap","Each") and total_del > 0:
            # For piece-based items: used proportion of delivered cost
            used_ratio = (total_del - lo_lf) / total_del if total_del else 0
            used_ratio = max(0.0, min(1.0, used_ratio))
            used_cost  = round(dcost * used_ratio, 2)
        else:
            used_cost = 0.0

        result.append({**item,"idx":idx,
            "total_qty":total_qty,"total_delivered":total_del,
            "remaining": rem,
            "lf":lf,"bf_sf":bfsf,"total_cost":cost,"total_cost_tax":round(cost*(1+po_tax),2),
            "delivered_lf":dlf,"delivered_bf_sf":dbfsf,"delivered_cost":dcost,
            "delivered_cost_tax":round(dcost*(1+del_tax),2),
            "remaining_lf":rlf,"remaining_bf_sf":rbfsf,"remaining_cost":rcost,
            "remaining_cost_tax":round(rcost*(1+po_tax),2),
            "leftover_lf":lo_lf,"used_lf":used_lf,"used_cost":used_cost,
            "delivered_pct":pct,
            "is_rl": rl,
            "tax_rate": po_tax,
            "po_tax_rate": po_tax,
            "delivery_tax_rate": del_tax})
    return {"items":result,"meta":meta,"tax_rate":po_tax,
            "po_tax_rate":po_tax,"delivery_tax_rate":del_tax}

@app.put("/projects/{project}/items/{item_idx}")
def update_item(project: str, item_idx: int, body: dict):
    """Update any field of a line item."""
    items=load_items(project)
    if item_idx<0 or item_idx>=len(items): raise HTTPException(404,"Item not found.")
    item=items[item_idx]
    # Allowed fields to update
    allowed=["type","description","t_num","w_num","length_num","unit_cost",
              "po_qty","co_qty","leftover_lf","cost_formula"]
    for k,v in body.items():
        if k in allowed:
            if k in ("t_num","w_num","length_num","unit_cost","po_qty","co_qty","leftover_lf"):
                item[k]=n(v)
            elif k == "cost_formula":
                # Store None (auto) or a valid formula string
                item[k] = v if v else None
            else:
                item[k]=v
    save_items(project, items)

    # Propagate ALL edits (qty, dimensions, formula) to CO items in meta
    meta = load_meta(project)
    po_tax = n(meta.get("po_tax_rate", meta.get("tax_rate", TAX_RATE)))
    updated_desc      = (item.get("description") or "").strip().upper()
    updated_desc_norm = norm_desc_key(item.get("description",""))
    cf_now  = item.get("cost_formula")
    new_tv  = n(item.get("t_num",    0))
    new_wv  = n(item.get("w_num",    0))
    new_lv  = n(item.get("length_num",0))
    new_coq = n(item.get("co_qty",   0))

    # Collect all CO items that match this inventory item
    matching_co_items = []
    for co in meta.get("change_orders", []):
        for ci in co.get("items", []):
            ci_desc_norm = norm_desc_key(ci.get("description",""))
            ci_desc_raw  = (ci.get("description") or "").strip().upper()
            if (ci_desc_raw  == updated_desc or
                ci_desc_raw  == updated_desc_norm or
                ci_desc_norm == updated_desc or
                ci_desc_norm == updated_desc_norm):
                matching_co_items.append((co, ci))

    changed_cos = bool(matching_co_items)

    # If co_qty was edited AND there are matching CO items, redistribute the new
    # total co_qty across all matching CO items proportionally (preserve sign).
    if matching_co_items:
        # Sum of existing qty_change across all matching CO items
        existing_total_qty = sum(abs(n(ci.get("qty_change",0)))
                                 for _, ci in matching_co_items)

        for co, ci in matching_co_items:
            typ = ci.get("type",""); uc = n(ci.get("unit_cost",0))
            old_qty = n(ci.get("qty_change", 0))
            sign    = -1 if old_qty < 0 else 1

            # ── Update dimensions from the item if they changed ─────────────
            tv = new_tv if new_tv else n(ci.get("t_num",0))
            wv = new_wv if new_wv else n(ci.get("w_num",0))
            lv = new_lv if new_lv else n(ci.get("length_num",0))

            # ── Fix panel dimensions ─────────────────────────────────────────
            if typ == "Panels":
                if tv >= 1.0 or tv == 0.0:
                    TMAP = {"7/16":0.4375,"15/32":0.46875,"1/2":0.5,
                            "19/32":0.59375,"5/8":0.625,"23/32":0.71875,"3/4":0.75}
                    ci_d = (ci.get("description") or "").upper()
                    tv = next((v for k,v in TMAP.items() if k in ci_d), 0.46875)
                wv = 8; lv = 0
                ci["t_num"] = tv; ci["w_num"] = 8; ci["length_num"] = 0

            # ── Update qty_change if co_qty was changed ─────────────────────
            # Distribute new_coq across CO items proportionally
            if existing_total_qty > 0:
                proportion  = abs(old_qty) / existing_total_qty
                new_ci_qty  = round(new_coq * proportion, 4) * sign
            else:
                new_ci_qty  = new_coq * sign
            ci["qty_change"] = new_ci_qty
            ci["t_num"] = tv; ci["w_num"] = wv; ci["length_num"] = lv

            # ── Update formula ───────────────────────────────────────────────
            if cf_now is not None:
                ci["cost_formula"] = cf_now
            cf = ci.get("cost_formula")

            # ── Recompute cost with updated qty + dims + formula ────────────
            abs_new_qty = abs(new_ci_qty)
            new_cost = compute_cost(typ, abs_new_qty, tv, wv, lv, uc, cf) * sign
            ci["cost"]     = new_cost
            ci["cost_tax"] = round(new_cost*(1+po_tax),2)

        # Recalculate CO totals
        for co in meta.get("change_orders", []):
            co["total_cost"]     = round(sum(n(ci.get("cost",0))
                                             for ci in co.get("items",[])), 2)
            co["total_cost_tax"] = round(co["total_cost"]*(1+po_tax),2)

    if changed_cos:
        save_meta(project, meta)

    try: rebuild_excel(project)
    except Exception as e: print(f"[WARN] rebuild after item update: {e}")
    return {"status":"ok","message":"Item updated."}

@app.post("/projects/{project}/items")
def add_item(project: str, body: dict):
    """Add a new line item manually."""
    items=load_items(project)
    new_item=normalize_item({
        "type":body.get("type","Unknown"),
        "description":body.get("description",""),
        "t_num":n(body.get("t_num",0)),
        "w_num":n(body.get("w_num",0)),
        "length_num":n(body.get("length_num",0)),
        "po_qty":n(body.get("po_qty",0)),
        "co_qty":n(body.get("co_qty",0)),
        "unit_cost":n(body.get("unit_cost",0)),
        "deliveries":{},"leftover_lf":0.0,"source":"manual",
    })
    items.append(new_item)
    save_items(project,items); rebuild_excel(project)
    return {"status":"ok","message":"Item added.","idx":len(items)-1}

@app.delete("/projects/{project}/items/{item_idx}")
def delete_item(project: str, item_idx: int):
    """Delete a line item."""
    items=load_items(project)
    if item_idx<0 or item_idx>=len(items): raise HTTPException(404,"Item not found.")
    removed=items.pop(item_idx)
    save_items(project,items); rebuild_excel(project)
    return {"status":"ok","message":f"Item '{removed.get('description','')}' deleted."}


@app.patch("/projects/{project}/items/{item_idx}/delivery")
def edit_delivery(project: str, item_idx: int, body: dict):
    """Edit the delivered quantity for a specific invoice on an item.
    body: {invoice_no: str, qty: number, unit: "LF"|"PC"|"BF"|"SF"}
    - For LF unit (LVL, R/L Lumber): qty is stored as LF directly
    - For PC unit: qty is stored as pieces (converted to LF via item dimensions)
    Setting qty to 0 removes that invoice entry.
    """
    items = load_items(project)
    if item_idx < 0 or item_idx >= len(items):
        raise HTTPException(404, "Item not found.")
    item = items[item_idx]
    inv_no  = str(body.get("invoice_no", "")).strip()
    new_qty = n(body.get("qty", 0))
    unit    = str(body.get("unit", "")).upper()
    if not inv_no:
        raise HTTPException(400, "invoice_no is required.")

    typ = item.get("type", "")
    is_lf_item = (typ in ("Lumber", "LVL") or is_rl_lumber(item))

    # Convert to the correct storage unit
    # If user selected LF explicitly → store as LF
    # If user selected PC for an LF item → convert to LF via item length
    if unit == "LF" and is_lf_item:
        store_qty = round(new_qty, 4)        # already LF
    elif unit == "PC" and is_lf_item:
        # Convert pieces to LF: LF = pieces × length
        lv = n(item.get("length_num", 0))
        store_qty = round(new_qty * lv, 4) if lv else round(new_qty, 4)
    else:
        # PC-based item: store as pieces
        store_qty = round(new_qty, 4)

    deliveries = item.setdefault("deliveries", {})
    old_qty = n(deliveries.get(inv_no, 0))
    if new_qty == 0:
        deliveries.pop(inv_no, None)
    else:
        deliveries[inv_no] = store_qty

    save_items(project, items)
    try: rebuild_excel(project)
    except Exception as e: print(f"[WARN] rebuild after delivery edit: {e}")
    display_unit = "LF" if is_lf_item else "PC"
    return {"status": "ok",
            "message": f"Delivery for {inv_no} updated: {old_qty:.4f} → {store_qty:.4f} {display_unit}",
            "item_idx": item_idx, "invoice_no": inv_no, "qty": store_qty}


@app.patch("/projects/{project}/items/{item_idx}/leftover")
def set_leftover(project: str, item_idx: int, body: dict):
    """Set leftover LF for an item."""
    items=load_items(project)
    if item_idx<0 or item_idx>=len(items): raise HTTPException(404,"Item not found.")
    items[item_idx]["leftover_lf"]=n(body.get("leftover_lf",0))
    save_items(project,items); rebuild_excel(project)
    return {"status":"ok","message":"Leftover updated."}

@app.post("/projects/{project}/items/{item_idx}/set-type")
def set_item_type(project: str, item_idx: int, body: dict):
    items=load_items(project)
    if item_idx<0 or item_idx>=len(items): raise HTTPException(404,"Item not found.")
    new_type=body.get("type","")
    if new_type not in ["Lumber","Panels","LVL","Each"]: raise HTTPException(400,"Invalid type.")
    items[item_idx]["type"]=new_type
    save_items(project,items); rebuild_excel(project)
    return {"status":"ok","message":f"Item type set to {new_type}"}



@app.get("/projects/{project}/tax-rate")
def get_tax_rate(project: str):
    meta = load_meta(project)
    return {
        "tax_rate":          meta.get("tax_rate",          TAX_RATE),
        "po_tax_rate":       meta.get("po_tax_rate",       meta.get("tax_rate", TAX_RATE)),
        "delivery_tax_rate": meta.get("delivery_tax_rate", meta.get("tax_rate", TAX_RATE)),
    }

@app.post("/projects/{project}/tax-rate")
def set_tax_rate(project: str, body: dict):
    meta = load_meta(project)
    # Accept either a single rate or separate po/delivery rates
    po_rate  = float(body.get("po_tax_rate",       body.get("tax_rate", TAX_RATE)))
    del_rate = float(body.get("delivery_tax_rate",  body.get("tax_rate", TAX_RATE)))
    for rate in [po_rate, del_rate]:
        if not (0 <= rate <= 1):
            raise HTTPException(400, "Tax rate must be 0–1 (e.g. 0.06 = 6%)")
    meta["po_tax_rate"]       = po_rate
    meta["delivery_tax_rate"] = del_rate
    meta["tax_rate"]          = po_rate   # backward compat
    save_meta(project, meta)
    rebuild_excel(project)
    return {"status": "ok",
            "po_tax_rate": po_rate, "delivery_tax_rate": del_rate,
            "message": f"PO tax: {po_rate*100:.1f}%  |  Invoice/Delivery tax: {del_rate*100:.1f}%"}


@app.get("/projects/{project}/change-orders")
def get_change_orders(project: str):
    """Return full CO list with item details for a project."""
    meta  = load_meta(project)
    items = load_items(project)
    cos   = meta.get("change_orders", [])
    result = []
    # Build lookup: description → cost_formula from current inventory items
    # Use normalized key (strips CO#xxx suffixes) so "19/32 CDX (CO#001)" matches "19/32 CDX"
    cf_lookup = {}
    for it in items:
        cf = it.get("cost_formula")
        if cf:
            # Store under both original and normalized key
            desc_raw  = (it.get("description") or "").strip().upper()
            desc_norm = norm_desc_key(it.get("description",""))
            cf_lookup[desc_raw]  = cf
            cf_lookup[desc_norm] = cf

    for co in cos:
        co_items_detail = []
        for ci in co.get("items", []):
            typ = ci.get("type","")
            uc  = n(ci.get("unit_cost",0))
            qty = n(ci.get("qty_change",0))
            tv  = n(ci.get("t_num",0)); wv=n(ci.get("w_num",0)); lv=n(ci.get("length_num",0))
            abs_qty = abs(qty)
            sign    = -1 if qty < 0 else 1
            # Look up cost_formula from current item (user may have changed it)
            # Normalize to strip (CO#001) suffixes before lookup
            desc_key      = (ci.get("description") or "").strip().upper()
            desc_key_norm = norm_desc_key(ci.get("description",""))
            cf = cf_lookup.get(desc_key) or cf_lookup.get(desc_key_norm) or ci.get("cost_formula")
            lf   = compute_lf(typ, abs_qty, tv, wv, lv) * sign
            bfsf = compute_bf_sf(typ, abs_qty, tv, wv, lv) * sign
            cost = compute_cost(typ, abs_qty, tv, wv, lv, uc, cf) * sign
            po_tax = n(meta.get("po_tax_rate", meta.get("tax_rate", TAX_RATE)))
            co_items_detail.append({
                **ci,
                "abs_qty": abs_qty,
                "lf": lf, "bfsf": bfsf,
                "cost": cost,
                "cost_tax": round(cost*(1+po_tax),2),
                "cost_formula": cf,
            })
        total_cost = sum(i["cost"] for i in co_items_detail)
        po_tax_    = n(meta.get("po_tax_rate", meta.get("tax_rate", TAX_RATE)))
        result.append({
            "co_no":   co["co_no"],
            "date":    co.get("date",""),
            "file":    co.get("file",""),
            "items":   co_items_detail,
            "total_cost":     round(total_cost, 2),
            "total_cost_tax": round(total_cost*(1+po_tax_),2),
            "item_count": len(co_items_detail),
            "po_tax_rate": po_tax_,
        })
    return {"project": project, "change_orders": result}

@app.get("/all-change-orders")
def get_all_change_orders():
    """Return CO list for ALL projects."""
    projects = _db.list_projects()
    result = {}
    for proj in sorted(projects):
        meta  = load_meta(proj)
        pitems = load_items(proj)
        cos   = meta.get("change_orders", [])
        po_tax = n(meta.get("po_tax_rate", meta.get("tax_rate", TAX_RATE)))
        # Build formula lookup for this project
        cf_lkp = {}
        for it in pitems:
            if it.get("cost_formula"):
                dk  = (it.get("description") or "").strip().upper()
                dkn = norm_desc_key(it.get("description",""))
                cf_lkp[dk]  = it["cost_formula"]
                cf_lkp[dkn] = it["cost_formula"]
        proj_cos = []
        for co in cos:
            total = 0
            for ci in co.get("items",[]):
                typ=ci.get("type",""); uc=n(ci.get("unit_cost",0)); qty=n(ci.get("qty_change",0))
                tv=n(ci.get("t_num",0)); wv=n(ci.get("w_num",0)); lv=n(ci.get("length_num",0))
                dk2  = (ci.get("description") or "").strip().upper()
                dk2n = norm_desc_key(ci.get("description",""))
                cf2  = cf_lkp.get(dk2) or cf_lkp.get(dk2n) or ci.get("cost_formula")
                sign2 = -1 if qty < 0 else 1
                cost=compute_cost(typ,abs(qty),tv,wv,lv,uc,cf2)*sign2; total+=cost
            proj_cos.append({
                "co_no": co["co_no"], "date": co.get("date",""),
                "item_count": len(co.get("items",[])),
                "total_cost": round(total,2),
                "total_cost_tax": round(total*(1+po_tax),2),
            })
        result[proj] = proj_cos
    return {"projects": result}


# ── Schedule / BT Estimate ────────────────────────────────────────────────────
import zipfile, io as _io

# Old /schedule/generate removed — use /projects/{project}/schedule/parse-bt instead

@app.get("/schedule/download/{filename}")
def download_schedule(filename: str):
    path = BASE_DIR / filename
    if not path.exists(): raise HTTPException(404, "File not found.")
    return FileResponse(str(path), filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Schedule storage endpoints ────────────────────────────────────────────────
def load_schedule(project):
    return _db.load_schedule(project) or {"activities":[], "generated_at":"", "project_name":"", "baseline_locked":False}

def save_schedule(project, data):
    _db.save_schedule(project, data)

@app.post("/projects/{project}/schedule/generate")
async def generate_project_schedule(project: str, body: dict):
    """Parse BT estimate data and save schedule to project."""
    acts = body.get("activities", [])
    proj_name = body.get("project_name", project)
    baseline_locked = body.get("baseline_locked", False)
    data = {
        "activities": acts,
        "project_name": proj_name,
        "generated_at": datetime.now().isoformat(),
        "baseline_locked": baseline_locked,
    }
    _db.create_project(project)
    save_schedule(project, data)
    return {"status":"ok","activities":len(acts),"message":f"Schedule saved with {len(acts)} activities."}

@app.get("/projects/{project}/schedule")
def get_schedule(project: str):
    return load_schedule(project)

@app.post("/projects/{project}/schedule/lock-baseline")
def lock_baseline(project: str):
    """Lock the current V2 schedule as a baseline (creates baseline #N + locks it).
    Frontend uses this single button — internally it calls the V2 baseline helper."""
    try:
        bl = _create_baseline_internal(project, f"Baseline {datetime.now():%Y-%m-%d %H:%M}", locked=True)
        return {"status":"ok","message":"Baseline locked successfully.","baseline":bl}
    except Exception as e:
        raise HTTPException(400, f"Lock baseline failed: {e}")


@app.delete("/projects/{project}/schedule/delete")
def delete_schedule(project: str):
    """Delete the entire schedule for a project (clears DB rows)."""
    save_schedule(project, {"activities": [], "generated_at": "", "project_name": "", "baseline_locked": False})
    save_sched_v2(project, {"activities": [], "relationships": [], "next_id": 1000})
    save_baselines(project, {"baselines": [], "next_bl_id": 1})
    return {"status": "ok", "message": f"Schedule for '{project}' deleted."}

@app.get("/projects/{project}/schedule/excel")
def download_schedule_excel(project: str):
    """Generate and download schedule Excel for a project.
    Reads V2 schedule first (current), falls back to legacy if V2 is empty."""
    # Try V2 first
    v2 = load_sched_v2(project)
    if v2 and v2.get("activities"):
        # Build relationship maps for predecessor/successor display
        rels = v2.get("relationships", [])
        pred_map = {}  # act_id → list of pred_ids
        succ_map = {}  # act_id → list of succ_ids
        for r in rels:
            pred_map.setdefault(r.get("succ_id",""), []).append(r.get("pred_id",""))
            succ_map.setdefault(r.get("pred_id",""), []).append(r.get("succ_id",""))

        acts = []
        for a in sorted(v2["activities"],
                        key=lambda x: (x.get("bldg_seq",9999), x.get("seq_no",9999))):
            acts.append({
                "id":             a.get("id",""),
                "p6_code":        a.get("p6_code","") or a.get("id",""),
                "name":           a.get("name",""),
                "building":       a.get("wbs",""),
                "level":          a.get("level",""),
                "seq_no":         a.get("seq_no",""),
                "bldg_seq":       a.get("bldg_seq",""),
                "activity_type":  a.get("cost_code","") or a.get("activity_type",""),
                "duration":       a.get("duration",0),
                "planned_start":  a.get("start","") or a.get("planned_start",""),
                "planned_finish": a.get("finish","") or a.get("planned_finish",""),
                "baseline_start": a.get("baseline_start",""),
                "baseline_finish":a.get("baseline_finish",""),
                "actual_start":   a.get("actual_start",""),
                "actual_finish":  a.get("actual_finish",""),
                "pct_complete":   a.get("pct_complete",0),
                "status":         a.get("status","Not Started"),
                "critical":       a.get("critical", False),
                "builder_cost":   a.get("builder_cost",0),
                "client_price":   a.get("client_price",0),
                "profit":         round((a.get("client_price",0) - a.get("builder_cost",0)), 2),
                "notes":          a.get("notes",""),
                "predecessor_ids": pred_map.get(a.get("id",""), []),
                "successor_ids":   succ_map.get(a.get("id",""), []),
            })
        data = {"activities": acts, "project_name": project}
    else:
        data = load_schedule(project)
        acts = data.get("activities",[])
    proj_name = data.get("project_name", project)
    if not acts:
        raise HTTPException(400, "No schedule data. Generate a schedule first.")
    
    import openpyxl as _xl
    from openpyxl.styles import Font as _F, PatternFill as _PF, Alignment as _A, Border as _B, Side as _S
    from openpyxl.utils import get_column_letter as _gcl
    
    wb = _xl.Workbook()
    ws = wb.active
    assert ws is not None  # type: ignore
    ws.title="Schedule"
    thin = _S(border_style="thin",color="CCCCCC")
    bdr = _B(left=thin,right=thin,top=thin,bottom=thin)
    hfill = _PF("solid",start_color="1B3A5C")
    hfont = _F(bold=True,color="FFFFFF",name="Arial",size=10)
    
    # Title
    ws.merge_cells("A1:X1")
    tc = ws.cell(row=1,column=1,value=f"PROJECT SCHEDULE — {proj_name}")
    tc.font=_F(bold=True,name="Arial",size=14,color="FFFFFF"); tc.fill=hfill
    tc.alignment=_A(horizontal="center",vertical="center"); ws.row_dimensions[1].height=30
    
    hdrs=["Seq","Activity ID","Activity Name","Building","Level","Activity Type",
          "Duration (Days)","Planned Start","Planned Finish",
          "Baseline Start","Baseline Finish","Actual Start","Actual Finish",
          "% Complete","Float (Days)","Status",
          "Builder Cost","Client Price","Profit","Profit %",
          "Predecessor IDs","Successor IDs","Material Types","Notes"]
    ws.row_dimensions[2].height=36
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(row=2,column=ci,value=h)
        c.font=hfont; c.fill=hfill; c.border=bdr
        c.alignment=_A(horizontal="center",vertical="center",wrap_text=True)
    
    BLDG_COLORS={"Building VA (North)":"DDEEFF","Building VA (South)":"D4EDDA",
                 "Club house VB (North)":"FFF3CD","Mail Kiosk (North)":"F8D7DA",
                 "Mail Kiosk (South)":"FCE4EC","Trash Enclosure (North)":"E8F5E9",
                 "Trash Enclosure (South)":"F3E5F5"}
    
    row=3
    for act in acts:
        bg=BLDG_COLORS.get(act.get("building",""),"F8F8F8")
        fill=_PF("solid",start_color=bg)
        ps=act.get("planned_start",""); pf=act.get("planned_finish","")
        bs=act.get("baseline_start",""); bf=act.get("baseline_finish","")
        float_days=""
        try:
            from datetime import datetime as _dt
            if pf and bf:
                _pfd=_dt.fromisoformat(pf.split("T")[0]); _bfd=_dt.fromisoformat(bf.split("T")[0])
                float_days=(_bfd-_pfd).days
        except: pass
        
        profit_pct=act.get("profit",0)/act.get("client_price",1) if act.get("client_price") else 0
        vals=[act.get("seq_no",""),
              act.get("id",""),act.get("name",""),act.get("building",""),act.get("level",""),
              act.get("activity_type",""),act.get("duration",0),
              ps[:10] if ps else "",pf[:10] if pf else "",
              bs[:10] if bs else "",bf[:10] if bf else "",
              act.get("actual_start","")[:10] if act.get("actual_start") else "",
              act.get("actual_finish","")[:10] if act.get("actual_finish") else "",
              act.get("pct_complete",0)/100,float_days,act.get("status",""),
              act.get("builder_cost",0),act.get("client_price",0),act.get("profit",0),
              profit_pct,
              ", ".join(act.get("predecessor_ids",[])),
              ", ".join(act.get("successor_ids",[])),
              ", ".join(act.get("material_types",[])),act.get("notes","")]
        fmts=["0",None,None,None,None,None,"0",
              "MM/DD/YYYY","MM/DD/YYYY","MM/DD/YYYY","MM/DD/YYYY","MM/DD/YYYY","MM/DD/YYYY",
              "0%","0",None,'"$"#,##0.00','"$"#,##0.00','"$"#,##0.00',"0.0%",None,None,None,None]
        aligns=["center","center","left","left","center","left","center",
                "center","center","center","center","center","center",
                "center","center","center","right","right","right","center","left","left","left","left"]
        for ci,(val,fmt,ha) in enumerate(zip(vals,fmts,aligns),1):
            c=ws.cell(row=row,column=ci,value=val)
            c.font=_F(name="Arial",size=9); c.fill=fill; c.border=bdr
            c.alignment=_A(horizontal=ha,vertical="center")
            if fmt: c.number_format=fmt
        row+=1
    
    # Grand total row
    if row>3:
        gtf=_PF("solid",start_color="1B3A5C"); gtft=_F(bold=True,color="FFFFFF",name="Arial",size=10)
        for ci in range(1,25):
            c=ws.cell(row=row,column=ci); c.fill=gtf; c.font=gtft; c.border=bdr
            c.alignment=_A(horizontal="center" if ci!=3 else "left",vertical="center")
        ws.cell(row=row,column=3,value="GRAND TOTAL")
        ws.cell(row=row,column=17,value=f"=SUM(Q3:Q{row-1})").number_format='"$"#,##0.00'
        ws.cell(row=row,column=18,value=f"=SUM(R3:R{row-1})").number_format='"$"#,##0.00'
        ws.cell(row=row,column=19,value=f"=SUM(S3:S{row-1})").number_format='"$"#,##0.00'
        ws.cell(row=row,column=20,value=f"=S{row}/R{row}").number_format="0.0%"
        for ci in (17,18,19,20):
            ws.cell(row=row,column=ci).fill=gtf; ws.cell(row=row,column=ci).font=gtft
    
    # col widths: Seq, ID, Name, Building, Level, Type, Dur, PlnSt, PlnFin,
    #             BlSt, BlFin, ActSt, ActFin, Pct, Float, Status,
    #             BldCost, CliPrice, Profit, Profit%, Preds, Succs, MatTypes, Notes
    col_w=[6,12,40,22,7,26,9,13,13,13,13,13,13,11,10,14,14,14,12,10,18,18,18,24]
    for ci,w in enumerate(col_w,1): ws.column_dimensions[_gcl(ci)].width=w  # type: ignore[assignment]
    ws.freeze_panes="D3"  # freeze Seq+ID+Name columns
    
    out=BASE_DIR/f"schedule_{project.replace(' ','_')}_export.xlsx"
    wb.save(str(out))
    return FileResponse(str(out),
        filename=f"{project}_Schedule.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/projects/{project}/schedule/parse-bt")
async def parse_bt_estimate(
    project: str,
    file: UploadFile = File(...),
    start_date: str = Form(""),
    days_layout: int = Form(3),
    days_wall_framing: int = Form(10),
    days_sheathing: int = Form(6),
    days_floor_truss: int = Form(5),
    days_corridor_floor_truss: int = Form(3),
    days_roof_truss: int = Form(7),
    days_hardware: int = Form(4),
    days_zip_sheathing: int = Form(4),
    days_punchout: int = Form(3),
    days_stairs: int = Form(2),
    days_corridor_plywood: int = Form(3),
    days_stair_plywood: int = Form(2),
    days_nails: int = Form(5),
):
    """Parse BT estimate XLS/XLSX using only built-in libs (openpyxl+xlrd) — no pandas needed."""
    import re as _re, io as _io
    from datetime import datetime as _dt, timedelta as _td

    raw = await file.read()
    fname = (file.filename or "").lower()

    # ── Read spreadsheet ──────────────────────────────────────────────────────
    data_rows = []
    try:
        if fname.endswith(".xlsx"):
            import openpyxl as _xl
            wb = _xl.load_workbook(_io.BytesIO(raw), data_only=True)
            ws = wb.active
            assert ws is not None  # type: ignore
            all_rows = list(ws.values)
            headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
            for r in all_rows[1:]:
                data_rows.append({headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))})
        else:
            try:
                import xlrd as _xl
                wb = _xl.open_workbook(file_contents=raw)
                ws = wb.sheet_by_index(0)
                headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
                for r in range(1, ws.nrows):
                    data_rows.append({headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)})
            except ImportError:
                raise HTTPException(500,
                    "xlrd not installed in your venv. Run: pip install xlrd  "
                    "OR save the BT estimate as .xlsx and upload that.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Cannot read file: {e}")

    def nf(v):
        try: return float(v or 0)
        except: return 0.0

    labor_rows = [r for r in data_rows if "labor" in str(r.get("Cost Type","")).lower()]
    if not labor_rows:
        raise HTTPException(400, "No labor rows found. Ensure the file has a 'Cost Type' column with 'Labor' values.")

    # ── Activity config ───────────────────────────────────────────────────────
    ACTIVITY_ORDER = [
        "Layout","Wall Framing","Sheathing","Stair Plywood Sheathing",
        "Corridor seperation Plywood Sheathing","Unit Floor Truss",
        "Corridor Floor Truss","Roof Truss","Hardware","Zip Sheathing",
        "Stair-landings","Punchout","Nails Framing","House Wrap",
    ]
    DAYS_MAP = {
        "Layout": days_layout, "Wall Framing": days_wall_framing,
        "Sheathing": days_sheathing, "Unit Floor Truss": days_floor_truss,
        "Corridor Floor Truss": days_corridor_floor_truss,
        "Roof Truss": days_roof_truss, "Hardware": days_hardware,
        "Zip Sheathing": days_zip_sheathing, "Punchout": days_punchout,
        "Stair-landings": days_stairs, "Stair Plywood Sheathing": days_stair_plywood,
        "Corridor seperation Plywood Sheathing": days_corridor_plywood,
        "Nails Framing": days_nails, "House Wrap": 2,
    }
    MAT_MAP = {
        "Wall Framing":["Lumber","LVL"],"Sheathing":["Panels"],
        "Stair Plywood Sheathing":["Panels"],
        "Corridor seperation Plywood Sheathing":["Panels"],
        "Zip Sheathing":["Panels"],"Hardware":["Each"],
        "Roof Truss":["LVL"],"Unit Floor Truss":["LVL"],"Corridor Floor Truss":["LVL"],
    }
    BLDG_ORDER = [
        "Club house VB (North)","Building VA (North)","Building VA (South)",
        "Mail Kiosk (North)","Mail Kiosk (South)",
        "Trash Enclosure (North)","Trash Enclosure (South)",
    ]

    def parse_title(title):
        t = str(title).strip()
        lv_m = _re.search(r"[-\s]+(L\d)\s*$", t)
        level = lv_m.group(1) if lv_m else ""
        building = next((b for b in BLDG_ORDER if b.lower() in t.lower()), "")
        if not building and "nails" in t.lower(): building = "All Buildings"
        act = next((a for a in ACTIVITY_ORDER if a.lower() in t.lower()), "")
        return building, act, level

    # ── Build activity list ───────────────────────────────────────────────────
    raw_acts = []
    for r in labor_rows:
        title = str(r.get("Title","")).strip()
        if not title: continue
        building, act, level = parse_title(title)
        if not act: continue
        raw_acts.append({
            "name":title,"building":building,"activity_type":act,"level":level,
            "duration":DAYS_MAP.get(act,5),
            "builder_cost":nf(r.get("Builder Cost",0)),
            "client_price":nf(r.get("Client Price",0)),
            "profit":nf(r.get("Profit",0)),
            "quantity":nf(r.get("Quantity",0)),
            "unit":str(r.get("Unit","")).strip(),
        })

    if not raw_acts:
        raise HTTPException(400, "No recognizable activities found. Check Title column contains names like 'Layout', 'Wall Framing', etc.")

    def sort_key(a):
        bi=BLDG_ORDER.index(a["building"]) if a["building"] in BLDG_ORDER else 99
        li=int(a["level"][1]) if a["level"] and a["level"][1:].isdigit() else 0
        ai=ACTIVITY_ORDER.index(a["activity_type"]) if a["activity_type"] in ACTIVITY_ORDER else 99
        return (bi,li,ai)
    raw_acts.sort(key=sort_key)

    # ── Date calculation ──────────────────────────────────────────────────────
    try:
        start_dt = _dt.strptime(start_date,"%Y-%m-%d") if start_date else _dt.now().replace(hour=0,minute=0,second=0,microsecond=0)
    except ValueError:
        start_dt = _dt.now().replace(hour=0,minute=0,second=0,microsecond=0)

    bl_end = {}  # (building,level) -> last end date
    acts = []
    for idx, a in enumerate(raw_acts):
        b=a["building"]; lv=a["level"]
        prev = bl_end.get((b,lv), start_dt)
        if lv=="L2": prev=max(prev, bl_end.get((b,"L1"),start_dt))
        elif lv=="L3": prev=max(prev, bl_end.get((b,"L2"),start_dt))
        ps=prev; pf=ps+_td(days=a["duration"])
        bl_end[(b,lv)]=pf
        code=f"A{(idx+1)*10:04d}"
        acts.append({
            "id":code,"name":a["name"],"building":a["building"],
            "level":a["level"],"activity_type":a["activity_type"],
            "duration":a["duration"],
            "planned_start":ps.strftime("%Y-%m-%d"),
            "planned_finish":pf.strftime("%Y-%m-%d"),
            "baseline_start":ps.strftime("%Y-%m-%d"),
            "baseline_finish":pf.strftime("%Y-%m-%d"),
            "baseline_duration":a["duration"],"baseline_cost":a["builder_cost"],
            "actual_start":"","actual_finish":"","pct_complete":0.0,"status":"Not Started",
            "builder_cost":a["builder_cost"],"client_price":a["client_price"],
            "profit":a["profit"],"quantity":a["quantity"],"unit":a["unit"],
            "material_types":MAT_MAP.get(a["activity_type"],[]),
            "predecessor_ids":[],"notes":"",
        })

    _db.create_project(project)
    sch_data={
        "activities":acts,"project_name":project,
        "generated_at":_dt.now().isoformat(),
        "baseline_locked":True,"baseline_date":_dt.now().isoformat(),
        "start_date":start_dt.strftime("%Y-%m-%d"),
    }
    save_schedule(project,sch_data)
    return {"status":"ok","activities":len(acts),"schedule":sch_data,
            "message":f"Schedule built: {len(acts)} activities from BT estimate"}


@app.post("/projects/{project}/schedule/import-xer")
async def import_xer_schedule(project: str, file: UploadFile = File(...)):
    """Import a Primavera P6 XER file and save as project schedule."""
    from datetime import datetime as _dt
    import re as _re
    raw = await file.read()
    text = raw.decode('latin-1', errors='replace')
    lines = text.split('\n')

    in_task=False; task_fields=[]; tasks=[]
    in_rel=False; rel_fields=[]; rels=[]
    in_wbs=False; wbs_fields=[]; wbs_list=[]

    for line in lines:
        line=line.strip()
        if line=='%T\tTASK': in_task=True;in_rel=False;in_wbs=False;continue
        if line=='%T\tTASKPRED': in_rel=True;in_task=False;in_wbs=False;continue
        if line=='%T\tPROJWBS': in_wbs=True;in_task=False;in_rel=False;continue
        if line.startswith('%T\t'): in_task=False;in_rel=False;in_wbs=False;continue
        if in_task:
            if line.startswith('%F\t'): task_fields=line[3:].split('\t')
            elif line.startswith('%R\t'):
                vals=line[3:].split('\t'); tasks.append(dict(zip(task_fields,vals)))
        if in_rel:
            if line.startswith('%F\t'): rel_fields=line[3:].split('\t')
            elif line.startswith('%R\t'):
                vals=line[3:].split('\t'); rels.append(dict(zip(rel_fields,vals)))
        if in_wbs:
            if line.startswith('%F\t'): wbs_fields=line[3:].split('\t')
            elif line.startswith('%R\t'):
                vals=line[3:].split('\t'); wbs_list.append(dict(zip(wbs_fields,vals)))

    wbs_map = {w.get('wbs_id',''): w.get('wbs_name','') for w in wbs_list}
    id_to_code = {t['task_id']: t.get('task_code','') for t in tasks}
    code_to_id = {t.get('task_code',''): t['task_id'] for t in tasks}

    # Build predecessor map: task_code → list of (pred_code, rel_type, lag_days)
    pred_map = {}
    for r in rels:
        succ_id = r.get('task_id',''); pred_id = r.get('pred_task_id','')
        succ_code = id_to_code.get(succ_id,''); pred_code = id_to_code.get(pred_id,'')
        if succ_code and pred_code:
            lag_days = round(float(r.get('lag_hr_cnt',0) or 0) / 22, 1)
            pred_map.setdefault(succ_code,[]).append({
                "pred_id": pred_code,
                "rel_type": r.get('pred_type','PR_FS'),
                "lag_days": lag_days
            })

    BLDG_COLORS = {
        "Club House":"Club house VB (North)",
        "Buildind VA (North)":"Building VA (North)",
        "Buildind VA (South)":"Building VA (South)",
        "North":"Mail Kiosk (North)",  # context-dependent
        "South":"Mail Kiosk (South)",
    }

    acts = []
    for t in tasks:
        code = t.get('task_code','')
        name = t.get('task_name','')
        if not code or not name: continue
        # Skip milestones
        if t.get('task_type','') == 'TT_Mile': continue

        wbs_id  = t.get('wbs_id','')
        wbs_name = wbs_map.get(wbs_id,'')
        # Normalize building name
        building = BLDG_COLORS.get(wbs_name, wbs_name)
        # If "North" or "South", check name for context
        if wbs_name in ("North","South"):
            if "mail kiosk" in name.lower():
                building = f"Mail Kiosk ({wbs_name})"
            elif "trash" in name.lower():
                building = f"Trash Enclosure ({wbs_name})"

        # Duration in hours → days (22hr calendar)
        dur_hrs  = float(t.get('target_drtn_hr_cnt',0) or 0)
        dur_days = max(1, round(dur_hrs / 22))

        # Dates
        def parse_date(s):
            s = str(s or '')[:10]
            try: return _dt.strptime(s,'%Y-%m-%d').strftime('%Y-%m-%d')
            except: return ''

        ps = parse_date(t.get('target_start_date',''))
        pf = parse_date(t.get('target_end_date',''))
        pct = float(t.get('phys_complete_pct',0) or 0)

        # Extract level from name
        import re as _re
        lv_m = _re.search(r'[-\s]+(L\d)\s*$', name)
        level = lv_m.group(1) if lv_m else ""

        # Activity type
        ACT_TYPES = ["Layout","Wall Framing","Sheathing","Stair Plywood Sheathing",
                     "Corridor seperation Plywood Sheathing","Unit Floor Truss",
                     "Corridor Floor Truss","Roof Truss","Hardware","Zip Sheathing",
                     "Stair-landings","Punchout","House Wrap"]
        activity_type = next((a for a in ACT_TYPES if a.lower() in name.lower()),"")

        status = "Completed" if pct>=100 else "In Progress" if pct>0 else "Not Started"

        preds = [p["pred_id"] for p in pred_map.get(code,[])]

        acts.append({
            "id": code,
            "name": name,
            "building": building,
            "level": level,
            "activity_type": activity_type,
            "duration": dur_days,
            "planned_start": ps, "planned_finish": pf,
            "baseline_start": ps, "baseline_finish": pf,
            "baseline_duration": dur_days, "baseline_cost": 0,
            "actual_start": ps if pct>0 else "",
            "actual_finish": pf if pct>=100 else "",
            "pct_complete": float(pct),
            "status": status,
            "builder_cost": 0, "client_price": 0, "profit": 0,
            "quantity": 0, "unit": "",
            "material_types": [],
            "predecessor_ids": preds,
            "predecessor_details": pred_map.get(code,[]),
            "notes": "",
        })

    if not acts:
        raise HTTPException(400, "No activities found in XER file.")

    _db.create_project(project)
    sch_data = {
        "activities": acts,
        "project_name": project,
        "generated_at": _dt.now().isoformat(),
        "baseline_locked": True,
        "baseline_date": _dt.now().isoformat(),
        "start_date": acts[0]["planned_start"] if acts else "",
        "source": "xer",
        "relationships": [
            {"pred": id_to_code.get(r.get("pred_task_id",""),""),
             "succ": id_to_code.get(r.get("task_id",""),""),
             "type": r.get("pred_type","PR_FS"),
             "lag":  round(float(r.get("lag_hr_cnt",0) or 0)/22,1)}
            for r in rels
        ]
    }
    save_schedule(project, sch_data)
    return {"status":"ok","activities":len(acts),
            "relationships":len(rels),"schedule":sch_data,
            "message":f"XER imported: {len(acts)} activities, {len(rels)} relationships"}


# ── Backup & Restore ──────────────────────────────────────────────────────────
import zipfile, io, time

@app.get("/backup/all")
def backup_all_projects():
    """Download a ZIP containing all project data from Postgres."""
    projects = _db.list_projects()
    buf = io.BytesIO()
    domain_loaders = {
        "meta.json":        load_meta,
        "items.json":       load_items,
        "schedule.json":    load_schedule,
        "schedule_v2.json": load_sched_v2,
        "baselines.json":   load_baselines,
        "calendar.json":    load_calendar,
        "labor.json":       load_labor,
        "bt_estimate.json": lambda p: _db.load_bt_estimate(p),
        "bt_pos.json":      load_bt_pos,
    }
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        manifest = {"created": datetime.now().isoformat(), "projects": projects}
        zf.writestr("manifest.json", json.dumps(manifest, indent=2))
        for proj in projects:
            for fname, loader in domain_loaders.items():
                try:
                    data = loader(proj)
                    if data:
                        zf.writestr(f"{proj}/{fname}", json.dumps(data, indent=2))
                except Exception:
                    pass
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=MatInv_Backup_{ts}.zip"})

@app.get("/backup/{project}")
def backup_single_project(project: str):
    """Download a ZIP for a single project from Postgres."""
    domain_loaders = {
        "meta.json":        load_meta,
        "items.json":       load_items,
        "schedule.json":    load_schedule,
        "schedule_v2.json": load_sched_v2,
        "baselines.json":   load_baselines,
        "calendar.json":    load_calendar,
        "labor.json":       load_labor,
        "bt_estimate.json": lambda p: _db.load_bt_estimate(p),
        "bt_pos.json":      load_bt_pos,
    }
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        manifest = {"created": datetime.now().isoformat(), "projects": [project]}
        zf.writestr("manifest.json", json.dumps(manifest, indent=2))
        for fname, loader in domain_loaders.items():
            try:
                data = loader(project)
                if data:
                    zf.writestr(f"{project}/{fname}", json.dumps(data, indent=2))
            except Exception:
                pass
    buf.seek(0)
    safe = project.replace(" ", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=MatInv_{safe}_{ts}.zip"})

@app.post("/restore")
async def restore_backup(file: UploadFile = File(...)):
    """Restore projects from a backup ZIP into Postgres."""
    data = await file.read()
    restored = []; skipped = []
    domain_savers = {
        "meta.json":        save_meta,
        "items.json":       save_items,
        "schedule.json":    save_schedule,
        "schedule_v2.json": save_sched_v2,
        "baselines.json":   save_baselines,
        "calendar.json":    save_calendar,
        "labor.json":       save_labor,
        "bt_estimate.json": save_bt_estimate,
        "bt_pos.json":      save_bt_pos,
    }
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        names = zf.namelist()
        project_names = set()
        for name in names:
            parts = name.split("/")
            if len(parts) == 2 and parts[1] in domain_savers:
                project_names.add(parts[0])
        for proj in project_names:
            try:
                _db.create_project(proj)
                for fname, saver in domain_savers.items():
                    arc_key = f"{proj}/{fname}"
                    if arc_key in names:
                        blob = json.loads(zf.read(arc_key))
                        saver(proj, blob)
                try:
                    rebuild_excel(proj)
                except Exception:
                    pass
                restored.append(proj)
            except Exception as e:
                skipped.append(f"{proj} ({e})")
    return {"status": "ok", "restored": restored, "skipped": skipped,
            "message": f"Restored {len(restored)} project(s): {', '.join(restored)}"}

# ── GSF Material Estimator ────────────────────────────────────────────────────
@app.post("/estimate-materials")
def estimate_materials(body: dict):
    """
    Estimate materials for a new project based on ACTUAL DELIVERED quantities
    from COMPLETED projects with matching state + building_type.

    Logic: delivered_qty / project_gsf = consumption_rate
           consumption_rate × target_gsf = estimated_qty_needed
           estimated_cost = estimated_qty × avg_unit_cost

    This gives real construction cost based on what was actually used,
    not what was ordered (which includes excess/waste).

    body: { gsf, state, building_type }
    """
    target_gsf   = n(body.get("gsf", 0))
    target_state = body.get("state", "").strip().upper()
    target_btype = body.get("building_type", "").strip().upper()
    target_arch  = body.get("architect", "").strip().upper()

    if not target_gsf:
        raise HTTPException(400, "GSF is required.")

    all_projects = _db.list_projects()

    def _get_project_gsf(proj, items_p, meta):
        """Get GSF from meta or estimate from delivered LF."""
        p_gsf = n(meta.get("gsf", 0))
        if p_gsf == 0:
            # Estimate from total delivered LF (more accurate than ordered)
            del_lf = sum(
                compute_lf(it.get("type",""),
                           sum(it.get("deliveries",{}).values()),
                           n(it.get("t_num",0)), n(it.get("w_num",0)), n(it.get("length_num",0)))
                for it in items_p if it.get("type") in ("Lumber","LVL")
            )
            if del_lf > 0:
                p_gsf = del_lf * 0.15
            else:
                # Fall back to ordered LF
                ord_lf = sum(
                    compute_lf(it.get("type",""),
                               n(it.get("po_qty",0))+n(it.get("co_qty",0)),
                               n(it.get("t_num",0)), n(it.get("w_num",0)), n(it.get("length_num",0)))
                    for it in items_p if it.get("type") in ("Lumber","LVL")
                )
                p_gsf = max(1.0, ord_lf * 0.15)
        return p_gsf

    # Step 1: Find COMPLETED projects matching state + building_type
    matching = []
    for proj in all_projects:
        items_p = load_items(proj)
        if not items_p: continue
        meta = load_meta(proj)
        # Only completed projects — they have actual delivered data
        if not meta.get("completed", False):
            continue
        p_gsf   = _get_project_gsf(proj, items_p, meta)
        p_state = meta.get("state", "").strip().upper()
        p_btype = meta.get("building_type", "").strip().upper()
        p_arch  = meta.get("architect", "").strip().upper()
        state_ok = (not target_state) or (not p_state) or (p_state == target_state)
        btype_ok = (not target_btype) or (not p_btype) or (p_btype == target_btype)
        arch_ok  = (not target_arch) or (target_arch in p_arch) or (p_arch in target_arch)
        if state_ok and btype_ok and arch_ok:
            matching.append((proj, p_gsf, items_p, meta))

    fallback_used = False

    # STRICT: Only completed projects are used. Never fall back to incomplete.
    if not matching:
        raise HTTPException(404,
            "No completed projects found matching your criteria. "
            "Please mark at least one project as Complete (✓ Mark Complete button) before using the estimator.")

    # Build per-description averages based on DELIVERED quantities
    from collections import defaultdict
    desc_data = defaultdict(lambda: {
        "type":"", "unit_costs":[], "del_qty_per_gsf":[], "ord_qty_per_gsf":[],
        "length_num":0, "t_num":0, "w_num":0, "projects":set()
    })

    for proj, p_gsf, items, meta in matching:
        del_tax = n(meta.get("delivery_tax_rate", meta.get("tax_rate", TAX_RATE)))
        po_tax  = n(meta.get("po_tax_rate",       meta.get("tax_rate", TAX_RATE)))
        for item in items:
            desc = item.get("description","")
            if not desc: continue
            typ  = item.get("type","")
            tv   = n(item.get("t_num",0)); wv=n(item.get("w_num",0)); lv=n(item.get("length_num",0))
            uc   = n(item.get("unit_cost",0))
            cf   = item.get("cost_formula")

            # DELIVERED quantity (actual consumption)
            del_raw = sum(item.get("deliveries",{}).values())
            # deliveries store LF for LVL and R/L Lumber, pieces for specific Lumber/Panels/Each
            if typ == "LVL":
                del_lf  = del_raw        # already LF (get_store_qty stores LF for LVL)
            elif typ == "Lumber" and lv == 1.0:
                del_lf  = del_raw        # R/L: already LF
            elif typ == "Lumber" and lv > 1:
                del_lf  = del_raw * lv   # specific length: pieces × length = LF
            else:
                del_lf  = 0.0
            del_qty = del_lf if typ in ("Lumber","LVL") and del_lf > 0 else del_raw
            if del_qty <= 0:
                # Fall back to ordered qty if no deliveries
                ord_raw = n(item.get("po_qty",0)) + n(item.get("co_qty",0))
                ord_lf  = compute_lf(typ, ord_raw, tv, wv, lv)
                ord_raw2 = n(item.get("po_qty",0)) + n(item.get("co_qty",0))
                if typ == "LVL":
                    ord_lf2 = ord_raw2  # po_qty for LVL might be pieces×length already
                elif typ == "Lumber" and lv == 1.0:
                    ord_lf2 = ord_raw2
                elif typ == "Lumber" and lv > 1:
                    ord_lf2 = ord_raw2 * lv
                else:
                    ord_lf2 = 0.0
                del_qty = ord_lf2 if typ in ("Lumber","LVL") and ord_lf2 > 0 else ord_raw2
            if del_qty <= 0: continue

            del_per_gsf = del_qty / p_gsf
            dd = desc_data[desc]
            dd["type"]         = typ
            dd["length_num"]   = lv
            dd["t_num"]        = tv
            dd["w_num"]        = wv
            dd["del_qty_per_gsf"].append(del_per_gsf)
            dd["projects"].add(proj)
            if uc > 0: dd["unit_costs"].append(uc)

    result = []
    for desc, dd in desc_data.items():
        rates = dd["del_qty_per_gsf"]
        if not rates: continue
        avg_per_gsf = sum(rates) / len(rates)
        est_qty     = round(avg_per_gsf * target_gsf, 2)
        if est_qty <= 0: continue
        avg_uc      = round(sum(dd["unit_costs"])/len(dd["unit_costs"]),2) if dd["unit_costs"] else 0
        typ         = dd["type"]
        tv          = dd["t_num"]; wv=dd["w_num"]; lv=dd["length_num"]
        cf          = dd.get("cost_formula")
        # For LVL/Lumber: est_qty IS the LF (since del_qty was already in LF)
        # est_qty is in LF for Lumber/LVL (that's what del_qty_per_gsf tracks)
        est_lf = est_qty if typ in ("Lumber","LVL") else round(compute_lf(typ,est_qty,tv,wv,lv),1)
        
        if typ == "Lumber" and lv > 1:
            # Convert LF back to pieces for compute_cost (avoids double-multiply by length)
            est_pieces = round(est_qty / lv, 4)
            est_cost   = compute_cost(typ, est_pieces, tv, wv, lv, avg_uc, cf)
        elif typ == "LVL":
            # LVL cost = LF × unit_cost — pass LF directly, length=0 so compute_lf returns qty
            est_cost = round(est_qty * avg_uc, 2)
        else:
            # R/L Lumber (l=1) and all others: est_qty is correct as-is
            est_cost = compute_cost(typ, est_qty, tv, wv, lv, avg_uc, cf)
        n_proj   = len(dd["projects"])
        result.append({
            "description":        desc,
            "type":               typ,
            "t_num": tv, "w_num": wv, "length_num": lv,
            "avg_del_per_gsf":    round(avg_per_gsf, 6),
            "estimated_qty":      est_qty,
            "estimated_lf":       est_lf,
            "avg_unit_cost":      avg_uc,
            "estimated_cost":     round(est_cost, 2),
            "estimated_cost_tax": round(est_cost*(1+TAX_RATE), 2),
            "projects_used":      n_proj,
        })

    TYPE_ORDER = ["Lumber","LVL","Panels","Siding","HouseWrap","Each","Unknown"]
    result.sort(key=lambda x: (TYPE_ORDER.index(x["type"]) if x["type"] in TYPE_ORDER else 99, x["description"]))

    proj_names = [p for p,_,__,___ in matching]
    return {
        "status":        "ok",
        "target_gsf":    target_gsf,
        "projects_used": len(matching),
        "project_names": proj_names,
        "state":         target_state,
        "building_type": target_btype,
        "architect":     target_arch,
        "fallback_used": fallback_used,
        "warning":       ("⚠ No completed projects found — using active projects with delivery data. "
                          "Mark projects complete for more accurate estimates.") if fallback_used else "",
        "estimates":     result,
        "total_est_cost_tax": round(sum(r["estimated_cost_tax"] for r in result), 2),
    }


@app.get("/estimate-materials/excel")
def estimate_excel(gsf: float = 0, state: str = "", building_type: str = "", architect: str = ""):
    """Download GSF estimate as Excel."""
    from fastapi import Query
    # Re-use the estimate logic
    target_gsf   = n(gsf)
    target_state = state.strip().upper()
    target_btype = building_type.strip().upper()
    target_arch  = architect.strip().upper()
    if not target_gsf:
        raise HTTPException(400, "GSF is required.")

    projects = _db.list_projects()
    matching = []
    for proj in projects:
        meta  = load_meta(proj); p_gsf = n(meta.get("gsf", 0))
        if p_gsf == 0: continue
        p_state = meta.get("state","").strip().upper()
        p_btype = meta.get("building_type","").strip().upper()
        p_arch  = meta.get("architect","").strip().upper()
        state_ok = (not target_state) or (not p_state) or (p_state == target_state)
        btype_ok = (not target_btype) or (not p_btype) or (p_btype == target_btype)
        arch_ok  = (not target_arch) or (target_arch in p_arch) or (p_arch in target_arch)
        if state_ok and btype_ok and arch_ok:
            matching.append((proj, p_gsf, load_items(proj)))

    from collections import defaultdict
    desc_data = defaultdict(lambda: {"type":"","unit_costs":[],"qty_per_gsf":[],"length_num":0,"t_num":0,"w_num":0})
    for proj, p_gsf, items in matching:
        for item in items:
            desc = item.get("description","")
            if not desc: continue
            tq = n(item.get("po_qty",0))+n(item.get("co_qty",0))
            if tq==0: continue
            dd=desc_data[desc]; dd["type"]=item.get("type","")
            dd["length_num"]=item.get("length_num",0); dd["t_num"]=item.get("t_num",0); dd["w_num"]=item.get("w_num",0)
            dd["qty_per_gsf"].append(tq/p_gsf)
            uc=n(item.get("unit_cost",0))
            if uc>0: dd["unit_costs"].append(uc)

    C = get_column_letter
    wb = openpyxl.Workbook(); ws = wb.active
    assert ws is not None  # type: ignore
    ws.title = "Material Estimate"
    title_row = [f"MATERIAL ESTIMATE — GSF: {int(target_gsf):,}",
                 f"State: {target_state or 'All'}","","","","","",
                 f"Building Type: {target_btype or 'All'}","",
                 f"Based on {len(matching)} project(s)"]
    for ci,v in enumerate(title_row,1):
        cell=ws.cell(row=1,column=ci,value=v)
        cell.font=Font(bold=True,name="Arial",size=11,color="FFFFFF")
        cell.fill=HEADER_FILL
    hdrs=["Type","Description","T","W","Length","Est. Qty (PCS)","Est. LF","Avg Unit Cost","Est. Cost","Est. Cost+Tax","Projects Used"]
    for ci,h in enumerate(hdrs,1):
        cell=ws.cell(row=2,column=ci,value=h)
        cell.font=HEADER_FONT; cell.fill=HEADER_FILL
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=BORDER
    TYPE_ORDER=["Lumber","LVL","Each","Panels","Siding","HouseWrap","Unknown"]
    rows=[]
    for desc,dd in desc_data.items():
        if not dd["qty_per_gsf"]: continue
        avg=sum(dd["qty_per_gsf"])/len(dd["qty_per_gsf"]); est_qty=round(avg*target_gsf,1)
        avg_uc=round(sum(dd["unit_costs"])/len(dd["unit_costs"]),2) if dd["unit_costs"] else 0
        typ=dd["type"]; est_lf=round(compute_lf(typ,est_qty,dd["t_num"],dd["w_num"],dd["length_num"]),1)
        est_cost=compute_cost(typ,est_qty,dd["t_num"],dd["w_num"],dd["length_num"],avg_uc)
        rows.append((typ,desc,dd["t_num"],dd["w_num"],dd["length_num"],est_qty,est_lf,avg_uc,est_cost,round(est_cost*(1+TAX_RATE),2),len(dd["qty_per_gsf"])))
    rows.sort(key=lambda x:(TYPE_ORDER.index(x[0]) if x[0] in TYPE_ORDER else 99,x[1]))
    for ri,row_data in enumerate(rows,3):
        typ=row_data[0]; fill=TYPE_FILLS.get(typ,TYPE_FILLS["Unknown"])
        for ci,val in enumerate(row_data,1):
            cell=ws.cell(row=ri,column=ci,value=val)
            cell.fill=fill; cell.font=NORMAL_FONT; cell.border=BORDER
            cell.alignment=Alignment(horizontal="left" if ci==2 else "center",vertical="center")
            if ci in (9,10): cell.number_format=MONEY_FMT
            elif ci==8: cell.number_format=DEC_FMT
    for ci,w in enumerate([10,34,6,6,8,12,12,12,14,14,12],1):  # type: ignore[assignment]
        ws.column_dimensions[C(ci)].width=w
    ws.freeze_panes="B3"
    path=BASE_DIR/"estimate_output.xlsx"; wb.save(str(path))
    state_slug = target_state or "All"
    return FileResponse(str(path),filename=f"Material_Estimate_{int(target_gsf)}GSF_{state_slug}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Summary endpoint ──────────────────────────────────────────────────────────
@app.get("/summary")
def get_summary():
    """Get summary data for all projects."""
    projects=_db.list_projects()
    result=[]
    for proj in sorted(projects):
        items=load_items(proj); meta=load_meta(proj)
        total_qty=sum(n(i.get("po_qty",0))+n(i.get("co_qty",0)) for i in items)
        total_del=sum(sum(i.get("deliveries",{}).values()) for i in items)
        total_cost=sum(item_fields(i)[2] for i in items)
        total_del_cost=sum(item_fields(i,sum(i.get("deliveries",{}).values()))[2] for i in items)
        rem_cost=total_cost-total_del_cost
        pct=round(total_del/total_qty*100,1) if total_qty else 0

        # NEW: pull client cost / builder cost / profit / margin from BT estimate
        bt_est = load_bt_estimate(proj)
        builder_cost_bt = sum(r.get("builder_cost", 0) for r in bt_est)
        client_cost_bt  = sum(r.get("client_price", 0) for r in bt_est)
        profit_bt = client_cost_bt - builder_cost_bt
        margin_pct = round(profit_bt / client_cost_bt * 100, 2) if client_cost_bt else 0

        # PO totals (Labor/GC/Material BT POs)
        bt_pos = load_bt_pos(proj)
        po_total = sum(r.get("cost", 0) for r in bt_pos if not r.get("is_vpo"))
        vpo_total = sum(r.get("cost", 0) for r in bt_pos if r.get("is_vpo"))
        billed_total = sum(r.get("cost",0) * r.get("pct_billed",0)/100 for r in bt_pos)

        result.append({
            "project":proj,
            "item_count":len(items),
            "total_qty":total_qty,
            "total_delivered":total_del,
            "remaining":total_qty-total_del,
            "total_cost":total_cost,
            "total_cost_tax":round(total_cost*(1+TAX_RATE),2),
            "delivered_cost":total_del_cost,
            "delivered_cost_tax":round(total_del_cost*(1+TAX_RATE),2),
            "remaining_cost":rem_cost,
            "remaining_cost_tax":round(rem_cost*(1+TAX_RATE),2),
            "pct_delivered":pct,
            "invoice_count":len(meta.get("invoices",[])),
            "co_count":meta.get("co_count",0),
            "change_orders":meta.get("change_orders",[]),
            "invoices":meta.get("invoices",[]),
            # NEW BT-driven fields
            "bt_builder_cost":   round(builder_cost_bt, 2),
            "bt_client_cost":    round(client_cost_bt, 2),
            "bt_profit":         round(profit_bt, 2),
            "bt_margin_pct":     margin_pct,
            "bt_po_total":       round(po_total, 2),
            "bt_vpo_total":      round(vpo_total, 2),
            "bt_po_vpo_total":   round(po_total + vpo_total, 2),
            "bt_billed_total":   round(billed_total, 2),
            "has_bt":            bool(bt_est),
        })
    return {"projects":result}

# ── Downloads ─────────────────────────────────────────────────────────────────
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _make_streaming_response(wb, filename: str):
    """Save openpyxl workbook to BytesIO and stream it as a download response."""
    import io
    from fastapi.responses import StreamingResponse
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = filename.replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{safe}"',
                 "Cache-Control": "no-cache, no-store, must-revalidate"}
    )


@app.get("/projects/{project}/download")
def download_excel(project: str):
    """Full inventory Excel: Inventory + Change Orders + Invoices + Audit sheets."""
    items = load_items(project)
    meta  = load_meta(project)
    if not items:
        raise HTTPException(400, "No items yet — upload a PO first.")
    try:
        wb = _build_workbook_for_project(project, items, meta)
    except Exception as e:
        raise HTTPException(500, f"Excel build failed: {e}")
    safe = project.replace(" ","_")
    return _make_streaming_response(wb, f"{safe}_inventory.xlsx")


@app.get("/projects/{project}/download-po")
def download_po_excel(project: str):
    items = load_items(project)
    meta  = load_meta(project)
    if not items:
        raise HTTPException(400, "No PO data — upload a PO first.")
    try:
        wb = _build_po_workbook(project, items, meta)
    except Exception as e:
        raise HTTPException(500, f"PO Excel build failed: {e}")
    safe = project.replace(" ","_")
    return _make_streaming_response(wb, f"{safe}_PO_report.xlsx")


@app.get("/projects/{project}/download-co")
def download_co_excel(project: str):
    meta = load_meta(project)
    if not meta.get("change_orders"):
        raise HTTPException(400, "No change orders yet.")
    try:
        wb = _build_co_workbook(project, meta)
    except Exception as e:
        raise HTTPException(500, f"CO Excel build failed: {e}")
    safe = project.replace(" ","_")
    return _make_streaming_response(wb, f"{safe}_CO_report.xlsx")


@app.get("/download-all-projects")
def download_all_projects_excel():
    """One Excel with All Projects Summary + one Inventory sheet per project."""
    projects = sorted(_db.list_projects())
    if not projects:
        raise HTTPException(400, "No projects found.")
    try:
        wb = _build_all_projects_workbook(projects)
    except Exception as e:
        raise HTTPException(500, f"All-projects Excel build failed: {e}")
    return _make_streaming_response(wb, "All_Projects_Report.xlsx")



# ══════════════════════════════════════════════════════════════════════════════
# WORKBOOK BUILDERS — return openpyxl Workbook objects (no disk write)
# ══════════════════════════════════════════════════════════════════════════════

def _xl_styles():
    """Return common style objects."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(border_style="thin", color="CCCCCC")
    bdr  = lambda: Border(left=thin, right=thin, top=thin, bottom=thin)
    HFILL = PatternFill("solid", start_color="1B3A5C")
    HFONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    NFONT = Font(name="Arial", size=9)
    SFONT = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    SFILL = PatternFill("solid", start_color="2C3E50")
    GFONT = Font(bold=True, color="FFD700",  name="Arial", size=10)
    GFILL = PatternFill("solid", start_color="1A252F")
    TYPE_FILLS = {
        "Lumber":    PatternFill("solid", start_color="DDEEFF"),
        "LVL":       PatternFill("solid", start_color="E8D5FF"),
        "Each":      PatternFill("solid", start_color="FFF3CD"),
        "Panels":    PatternFill("solid", start_color="D4EDDA"),
        "Siding":    PatternFill("solid", start_color="FCE4EC"),
        "HouseWrap": PatternFill("solid", start_color="E0F7FA"),
        "Unknown":   PatternFill("solid", start_color="F8F8F8"),
    }
    MFMT = '"$"#,##0.00'
    NFMT = '#,##0.0##'
    IFMT = '#,##0'
    return dict(bdr=bdr, HFILL=HFILL, HFONT=HFONT, NFONT=NFONT,
                SFONT=SFONT, SFILL=SFILL, GFONT=GFONT, GFILL=GFILL,
                TYPE_FILLS=TYPE_FILLS, MFMT=MFMT, NFMT=NFMT, IFMT=IFMT)


def _write_hdr(ws, r, c, v, s):
    from openpyxl.styles import Alignment
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = s["HFONT"]; cell.fill = s["HFILL"]
    cell.border = s["bdr"]()  # type: ignore[operator]
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _write_val(ws, r, c, v, fill, font, bdr, fmt=None, ha="center"):
    from openpyxl.styles import Alignment
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font; cell.fill = fill; cell.border = bdr()
    cell.alignment = Alignment(horizontal=ha, vertical="center")
    if fmt: cell.number_format = fmt
    return cell


def _build_inventory_ws(ws, items_list, meta, s):
    """Write Inventory sheet into ws."""
    from openpyxl.utils import get_column_letter as gcl
    from openpyxl.styles import Alignment, PatternFill, Font
    po_tax  = n(meta.get("po_tax_rate",  meta.get("tax_rate", TAX_RATE)))
    del_tax = n(meta.get("delivery_tax_rate", meta.get("tax_rate", TAX_RATE)))
    invoices = meta.get("invoices", [])
    n_inv = len(invoices)
    ws.row_dimensions[1].height = 44
    ws.freeze_panes = "B2"
    MFMT=s["MFMT"]; NFMT=s["NFMT"]; IFMT=s["IFMT"]

    # Column indices
    INV0=14; INC=INV0+n_inv
    DQ=INC+1; DLF=INC+2; DBF=INC+3; DC=INC+4; DCX=INC+5
    RQ=INC+6; RLF=INC+7; RBF=INC+8; RC=INC+9; RCX=INC+10
    LO=INC+11; ULF=INC+12; UC_=INC+13; LAST=UC_

    fixed=["Type","Description","T","W","Length","PO Qty","CO Qty","PO+CO Qty",
           "Unit Cost","LF (Linear Ft)","BF/SF","Total Cost","Total Cost+Tax"]
    for ci,h in enumerate(fixed,1): _write_hdr(ws,1,ci,h,s)
    for i,inv in enumerate(invoices): _write_hdr(ws,1,INV0+i,inv.get("date",""),s)
    for ci,h in {INC:"Invoice #",DQ:"Total Delivered",DLF:"Delivered LF",DBF:"Delivered BF/SF",
                 DC:"Delivered Cost",DCX:"Delivered Cost+Tax",RQ:"Remaining Qty",
                 RLF:"Remaining LF",RBF:"Remaining BF/SF",RC:"Remaining Cost",
                 RCX:"Remaining Cost+Tax",LO:"Leftover LF (Site)",ULF:"Material Used LF",
                 UC_:"Material Used Cost"}.items():
        _write_hdr(ws,1,ci,h,s)

    TYPE_ORDER=["Lumber","LVL","Each","Panels","Siding","HouseWrap","Unknown"]
    row=2; grand={k:0.0 for k in ["lf","bf","cost","ctax","dlf","dbf","dc","dcx","rlf","rbf","rc","rcx","ulf","uc"]}

    for typ in TYPE_ORDER:
        grp=[it for it in items_list if it.get("type")==typ]
        if not grp: continue
        fill=(s["TYPE_FILLS"] or {}).get(typ, PatternFill("solid",start_color="F8F8F8"))
        sfill=PatternFill("solid",start_color="A8B8C8") 
        g={k:0.0 for k in grand}

        for item in grp:
            tv=n(item.get("t_num",0)); wv=n(item.get("w_num",0)); lv=n(item.get("length_num",0))
            uc=n(item.get("unit_cost",0)); typ2=item.get("type","")
            poq=n(item.get("po_qty",0)); coq=n(item.get("co_qty",0)); tq=poq+coq
            cf2 = item.get("cost_formula", None)
            lf=compute_lf(typ2,tq,tv,wv,lv); bf=compute_bf_sf(typ2,tq,tv,wv,lv)
            cost=compute_cost(typ2,tq,tv,wv,lv,uc,cf2); ctax=round(cost*(1+po_tax),2)
            invd=item.get("deliveries",{})
            inv_vals=[n(invd.get(inv["invoice_no"],0)) for inv in invoices]
            del_t=sum(inv_vals)
            inv_nos="\n".join(inv["invoice_no"] for inv in invoices if invd.get(inv["invoice_no"],0))
            dlf=compute_lf(typ2,del_t,tv,wv,lv); dbf=compute_bf_sf(typ2,del_t,tv,wv,lv)
            dc=compute_cost(typ2,del_t,tv,wv,lv,uc,cf2); dcx=round(dc*(1+del_tax),2)
            rem=tq-del_t
            rlf=compute_lf(typ2,max(0,rem),tv,wv,lv); rbf=compute_bf_sf(typ2,max(0,rem),tv,wv,lv)
            rc=compute_cost(typ2,max(0,rem),tv,wv,lv,uc,cf2); rcx=round(rc*(1+po_tax),2)
            lo=n(item.get("leftover_lf",0)); ulf=max(0.0,dlf-lo)
            if typ2=="Lumber" and tv and wv: ucc=round((ulf*tv*wv/12*uc)/1000,2)
            elif typ2=="LVL": ucc=round(ulf*uc,2)
            elif del_t>0: ucc=round(dc*max(0,(del_t-lo))/del_t,2)
            else: ucc=0.0
            for k,v2 in [("lf",lf),("bf",bf),("cost",cost),("ctax",ctax),
                         ("dlf",dlf),("dbf",dbf),("dc",dc),("dcx",dcx),
                         ("rlf",rlf),("rbf",rbf),("rc",rc),("rcx",rcx),
                         ("ulf",ulf),("uc",ucc)]:
                g[k]+=v2; grand[k]+=v2

            def wv2(c,v2,fmt=None,ha="center"):
                _write_val(ws,row,c,v2,fill,s["NFONT"],s["bdr"],fmt,ha)
            wv2(1,typ2); wv2(2,item.get("description",""),ha="left")
            for ci2,vv in enumerate([tv or "",wv or "",lv or "",poq,coq,tq,uc],3):
                fmt2=MFMT if ci2==9 else (IFMT if ci2 in (6,7,8) else None)
                wv2(ci2,vv,fmt2)
            wv2(10,lf,NFMT); wv2(11,bf,NFMT); wv2(12,cost,MFMT); wv2(13,ctax,MFMT)
            for i2,vv in enumerate(inv_vals): wv2(INV0+i2,vv,IFMT)
            wv2(INC,inv_nos,ha="left")
            wv2(DQ,del_t,IFMT); wv2(DLF,dlf,NFMT); wv2(DBF,dbf,NFMT)
            wv2(DC,dc,MFMT); wv2(DCX,dcx,MFMT)
            wv2(RQ,rem,IFMT); wv2(RLF,rlf,NFMT); wv2(RBF,rbf,NFMT)
            wv2(RC,rc,MFMT); wv2(RCX,rcx,MFMT)
            wv2(LO,lo,NFMT); wv2(ULF,ulf,NFMT); wv2(UC_,ucc,MFMT)
            row+=1

        # Subtotal
        for ci2 in range(1,LAST+1):
            c=ws.cell(row=row,column=ci2); c.font=s["SFONT"]; c.fill=sfill; c.border=s["bdr"]()  # type: ignore[assignment]
            c.alignment=Alignment(horizontal="left" if ci2==1 else "center",vertical="center")
        ws.cell(row=row,column=1).value=f"{typ.upper()} SUBTOTAL"
        sub={10:g["lf"],11:g["bf"],12:g["cost"],13:g["ctax"],
             DQ:sum(sum(n(it.get("deliveries",{}).get(inv["invoice_no"],0)) for inv in invoices) for it in grp),
             DLF:g["dlf"],DBF:g["dbf"],DC:g["dc"],DCX:g["dcx"],
             RLF:g["rlf"],RBF:g["rbf"],RC:g["rc"],RCX:g["rcx"],
             ULF:g["ulf"],UC_:g["uc"]}
        for ci2,vv in sub.items():
            c=ws.cell(row=row,column=ci2); c.value=vv; c.font=s["SFONT"]; c.fill=sfill; c.border=s["bdr"]()  # type: ignore[assignment]
            c.alignment=Alignment(horizontal="center",vertical="center")
            if ci2 in (12,13,DC,DCX,RC,RCX,UC_): c.number_format=MFMT
            elif ci2 in (10,11,DLF,DBF,RLF,RBF,ULF): c.number_format=NFMT
            elif ci2 in (DQ,RQ): c.number_format=IFMT
        row+=1; row+=1  # spacer

    # Grand Total
    for ci2 in range(1,LAST+1):
        c=ws.cell(row=row,column=ci2); c.font=s["GFONT"]; c.fill=s["GFILL"]; c.border=s["bdr"]()  # type: ignore[assignment]
        c.alignment=Alignment(horizontal="left" if ci2==1 else "center",vertical="center")
    ws.cell(row=row,column=1).value="GRAND TOTAL"  # type: ignore[union-attr]
    for ci2,vv in {10:grand["lf"],11:grand["bf"],12:grand["cost"],13:grand["ctax"],
                   DLF:grand["dlf"],DBF:grand["dbf"],DC:grand["dc"],DCX:grand["dcx"],
                   RLF:grand["rlf"],RBF:grand["rbf"],RC:grand["rc"],RCX:grand["rcx"],
                   ULF:grand["ulf"],UC_:grand["uc"]}.items():
        c=ws.cell(row=row,column=ci2); c.value=vv; c.font=s["GFONT"]; c.fill=s["GFILL"]; c.border=s["bdr"]()  # type: ignore[assignment]
        c.alignment=Alignment(horizontal="center",vertical="center")
        if ci2 in (12,13,DC,DCX,RC,RCX,UC_): c.number_format=MFMT
        elif ci2 in (10,11,DLF,DBF,RLF,RBF,ULF): c.number_format=NFMT

    # Column widths
    from openpyxl.utils import get_column_letter as gcl
    ws.column_dimensions["A"].width=10; ws.column_dimensions["B"].width=36
    for cc in "CDEFGHI": ws.column_dimensions[cc].width=9
    ws.column_dimensions["J"].width=12; ws.column_dimensions["K"].width=12
    ws.column_dimensions["L"].width=14; ws.column_dimensions["M"].width=14
    for i2 in range(n_inv): ws.column_dimensions[gcl(INV0+i2)].width=12
    for ci2,ww in zip(range(INC,LAST+1),[18,13,13,13,14,15,12,12,12,14,15,14,14,14]):
        ws.column_dimensions[gcl(ci2)].width=ww


def _build_co_ws(ws, meta, s):
    """Write Change Orders sheet."""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter as gcl
    po_tax=n(meta.get("po_tax_rate", meta.get("tax_rate",TAX_RATE)))
    ws.row_dimensions[1].height=36
    co_hdrs=["CO #","CO Date","Type","Description","T","W","Len",
             "Qty Change","Unit Cost","Footage/SF","LF","BF/SF",
             "Amount","Amount+Tax","Action","Matched PO Item"]
    for ci,h in enumerate(co_hdrs,1): _write_hdr(ws,1,ci,h,s)
    cr=2
    for co in meta.get("change_orders",[]):
        for ci_item in co.get("items",[]):
            typ=ci_item.get("type","")
            fill=(s["TYPE_FILLS"] or {}).get(typ, PatternFill("solid",start_color="F8F8F8"))
            qty=n(ci_item.get("qty_change",0))
            tv=n(ci_item.get("t_num",0)); wv=n(ci_item.get("w_num",0)); lv=n(ci_item.get("length_num",0))
            uc=n(ci_item.get("unit_cost",0)); sign=-1 if qty<0 else 1; aq=abs(qty)
            ftg=compute_bf_sf(typ,aq,tv,wv,lv)*sign
            lf_=compute_lf(typ,aq,tv,wv,lv)*sign
            bf_=compute_bf_sf(typ,aq,tv,wv,lv)*sign
            amt=compute_cost(typ,aq,tv,wv,lv,uc)*sign
            amtt=round(amt*(1+po_tax),2)
            rv=[co.get("co_no",""),co.get("date",""),typ,ci_item.get("description",""),
                tv or "",wv or "",lv or "",qty,uc,ftg,lf_,bf_,amt,amtt,
                ci_item.get("action",""),ci_item.get("matched_po","")]
            fmts=[None,None,None,None,None,None,None,s["IFMT"],s["MFMT"],s["NFMT"],s["NFMT"],s["NFMT"],s["MFMT"],s["MFMT"],None,None]
            for ci,(v2,f2) in enumerate(zip(rv,fmts),1):
                c=ws.cell(row=cr,column=ci,value=v2); c.font=s["NFONT"]; c.fill=fill; c.border=s["bdr"]()  # type: ignore[assignment]
                c.alignment=Alignment(horizontal="left" if ci in (4,15,16) else "center",vertical="center")
                if f2: c.number_format=f2
            cr+=1
    for ci,ww in enumerate([8,12,10,34,6,6,6,10,11,12,10,10,14,14,16,28],1):  # type: ignore[assignment]
        ws.column_dimensions[gcl(ci)].width=ww
    ws.freeze_panes="A2"


def _build_invoices_ws(ws, items_list, meta, s):
    """Write Invoices sheet."""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter as gcl
    del_tax=n(meta.get("delivery_tax_rate",meta.get("tax_rate",TAX_RATE)))
    ws.row_dimensions[1].height=36
    inv_hdrs=["Invoice #","Date","Description","Type","T","W","Length",
              "Qty Delivered","Delivered LF","Delivered BF/SF","Delivered Cost","Delivered Cost+Tax"]
    for ci,h in enumerate(inv_hdrs,1): _write_hdr(ws,1,ci,h,s)
    ir=2
    for inv in meta.get("invoices",[]):
        inv_no=inv["invoice_no"]; inv_date=inv.get("date","")
        for item in items_list:
            qty_del=n(item.get("deliveries",{}).get(inv_no,0))
            if qty_del==0: continue
            typ=item.get("type","")
            fill=(s["TYPE_FILLS"] or {}).get(typ, PatternFill("solid",start_color="F8F8F8"))
            tv=n(item.get("t_num",0)); wv=n(item.get("w_num",0)); lv=n(item.get("length_num",0))
            uc=n(item.get("unit_cost",0))
            dlf=compute_lf(typ,qty_del,tv,wv,lv); dbf=compute_bf_sf(typ,qty_del,tv,wv,lv)
            dcost=compute_cost(typ,qty_del,tv,ww,lv,uc) if False else compute_cost(typ,qty_del,tv,wv,lv,uc)
            dctax=round(dcost*(1+del_tax),2)
            rv=[inv_no,inv_date,item.get("description",""),typ,tv or "",wv or "",lv or "",
                qty_del,dlf,dbf,dcost,dctax]
            fmts=[None,None,None,None,None,None,None,s["IFMT"],s["NFMT"],s["NFMT"],s["MFMT"],s["MFMT"]]
            for ci,(v2,f2) in enumerate(zip(rv,fmts),1):
                c=ws.cell(row=ir,column=ci,value=v2); c.font=s["NFONT"]; c.fill=fill; c.border=s["bdr"]()  # type: ignore[assignment]
                c.alignment=Alignment(horizontal="left" if ci==3 else "center",vertical="center")
                if f2: c.number_format=f2
            ir+=1
    for ci,ww in enumerate([14,12,36,10,6,6,8,12,12,12,14,14],1):  # type: ignore[assignment]
        ws.column_dimensions[gcl(ci)].width=ww
    ws.freeze_panes="A2"


def _build_audit_ws(ws, items_list, meta, s):
    """Audit sheet: Serial#, Type, Material Name, PO+CO LF/BF/PC, Delivered, Remaining."""
    from openpyxl.styles import Alignment, PatternFill
    ws.row_dimensions[1].height=36
    hdrs=["#","Type","Material Name","PO+CO Qty (LF/BF/PC)","Delivered (LF/BF/PC)","Remaining (LF/BF/PC)"]
    for ci,h in enumerate(hdrs,1): _write_hdr(ws,1,ci,h,s)
    row=2
    for sno,item in enumerate(items_list,1):
        typ=item.get("type","")
        fill=(s["TYPE_FILLS"] or {}).get(typ, PatternFill("solid",start_color="F8F8F8"))
        tv=n(item.get("t_num",0)); wv=n(item.get("w_num",0)); lv=n(item.get("length_num",0))
        poq=n(item.get("po_qty",0)); coq=n(item.get("co_qty",0)); tq=poq+coq
        del_t=sum(item.get("deliveries",{}).values())
        rem_t=tq-del_t
        # Show LF for lumber, BF/SF for panels, PC for each
        def qty_str(t,q):
            lf=compute_lf(t,q,tv,wv,lv)
            bf=compute_bf_sf(t,q,tv,wv,lv)
            if t in ("Lumber","LVL") and lf: return lf
            if t=="Panels" and bf: return bf
            return q
        tq_v=qty_str(typ,tq); del_v=qty_str(typ,del_t); rem_v=qty_str(typ,max(0,rem_t))
        rv=[sno,typ,item.get("description",""),tq_v,del_v,rem_v]
        from openpyxl.styles import Alignment
        for ci,v2 in enumerate(rv,1):
            c=ws.cell(row=row,column=ci,value=v2); c.font=s["NFONT"]; c.fill=fill; c.border=s["bdr"]()  # type: ignore[assignment]
            c.alignment=Alignment(horizontal="left" if ci==3 else "center",vertical="center")
            if ci in (4,5,6) and isinstance(v2,float): c.number_format='#,##0.0'
        row+=1
    ws.column_dimensions["A"].width=6; ws.column_dimensions["B"].width=10
    ws.column_dimensions["C"].width=36
    for cc in "DEF": ws.column_dimensions[cc].width=18
    ws.freeze_panes="A2"


def _build_workbook_for_project(project, items_list, meta):
    """Build full 4-sheet workbook for one project."""
    import openpyxl
    s=_xl_styles()
    wb=openpyxl.Workbook()
    ws=wb.active
    assert ws is not None  # type: ignore
    ws.title="Inventory"
    _build_inventory_ws(ws, items_list, meta, s)
    ws2=wb.create_sheet("Change Orders")
    _build_co_ws(ws2, meta, s)
    ws3=wb.create_sheet("Invoices")
    _build_invoices_ws(ws3, items_list, meta, s)
    ws4=wb.create_sheet("Audit Report")
    _build_audit_ws(ws4, items_list, meta, s)
    return wb


def _build_po_workbook(project, items_list, meta):
    """Build PO report workbook."""
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter as gcl
    s=_xl_styles()
    po_tax=n(meta.get("po_tax_rate",meta.get("tax_rate",TAX_RATE)))
    wb=openpyxl.Workbook()
    ws=wb.active
    assert ws is not None  # type: ignore
    ws.title="PO Report"
    hdrs=["Type","Description","T","W","Length","PO Qty","Unit Cost","LF","BF/SF","Total Cost","Total Cost+Tax"]
    ws.row_dimensions[1].height=30
    for ci,h in enumerate(hdrs,1): _write_hdr(ws,1,ci,h,s)
    row=2
    for typ in ["Lumber","LVL","Each","Panels","Siding","HouseWrap","Unknown"]:
        grp=[it for it in items_list if it.get("type")==typ and n(it.get("po_qty",0))>0]
        if not grp: continue
        fill=(s["TYPE_FILLS"] or {}).get(typ, PatternFill("solid",start_color="F8F8F8"))
        ds=row
        for item in grp:
            tv=n(item.get("t_num",0)); wv=n(item.get("w_num",0)); lv=n(item.get("length_num",0))
            uc=n(item.get("unit_cost",0)); qty=n(item.get("po_qty",0))
            lf=compute_lf(typ,qty,tv,wv,lv); bf=compute_bf_sf(typ,qty,tv,wv,lv)
            cost=compute_cost(typ,qty,tv,wv,lv,uc)
            rv=[typ,item.get("description",""),tv or "",wv or "",lv or "",qty,uc,lf,bf,cost,round(cost*(1+po_tax),2)]
            for ci,v2 in enumerate(rv,1):
                c=ws.cell(row=row,column=ci,value=v2); c.font=s["NFONT"]; c.fill=fill; c.border=s["bdr"]()  # type: ignore[assignment]
                c.alignment=Alignment(horizontal="left" if ci==2 else "center",vertical="center")
                if ci in (10,11): c.number_format=s["MFMT"]  # type: ignore[assignment]
                elif ci==7: c.number_format=s["MFMT"]  # type: ignore[assignment]
                elif ci in (6,8,9): c.number_format=s["NFMT"]  # type: ignore[assignment]
            row+=1
        de=row-1
        sfill=PatternFill("solid",start_color="A8B8C8")
        for ci in range(1,12):
            c=ws.cell(row=row,column=ci); c.font=s["SFONT"]; c.fill=sfill; c.border=s["bdr"]()  # type: ignore[assignment]
            c.alignment=Alignment(horizontal="left" if ci==1 else "center",vertical="center")
        ws.cell(row=row,column=1).value=f"{typ.upper()} SUBTOTAL"
        for ci2 in [6,8,9,10,11]:
            c=ws.cell(row=row,column=ci2); c.value=f"=SUM({gcl(ci2)}{ds}:{gcl(ci2)}{de})"
            c.font=s["SFONT"]; c.fill=sfill; c.border=s["bdr"]()  # type: ignore[assignment]
            c.alignment=Alignment(horizontal="center",vertical="center")
            if ci2 in (10,11): c.number_format=s["MFMT"]  # type: ignore[assignment]
            else: c.number_format=s["NFMT"]  # type: ignore[assignment]
        row+=1
    for ci,ww in enumerate([10,34,7,6,7,9,11,12,12,14,14],1):  # type: ignore[assignment]
        ws.column_dimensions[gcl(ci)].width=ww
    ws.freeze_panes="B2"
    return wb


def _build_co_workbook(project, meta):
    """Build CO report workbook."""
    import openpyxl
    s=_xl_styles()
    wb=openpyxl.Workbook()
    ws=wb.active
    assert ws is not None  # type: ignore
    ws.title="CO Report"
    _build_co_ws(ws, meta, s)
    return wb


def _build_all_projects_workbook(projects):
    """Build All Projects workbook: Summary + one sheet per project."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter as gcl
    s=_xl_styles()
    wb=openpyxl.Workbook(); wb.remove(wb.active)  # type: ignore[arg-type]
    ws_sum=wb.create_sheet("All Projects Summary")
    ws_sum.row_dimensions[1].height=44
    # Headers: PO and CO split for both Qty and Cost
    sum_hdrs=["Project","Total Items",
              "PO Qty","CO Qty","PO+CO Qty",
              "PO Cost+Tax","CO Cost+Tax","Total Cost+Tax",
              "Delivered Cost+Tax","Remaining Cost+Tax",
              "Change Orders","Invoices"]
    for ci,h in enumerate(sum_hdrs,1): _write_hdr(ws_sum,1,ci,h,s)
    sr=2
    for proj in projects:
        items_list=load_items(proj); meta=load_meta(proj)
        po_tax=n(meta.get("po_tax_rate",meta.get("tax_rate",TAX_RATE)))
        del_tax=n(meta.get("delivery_tax_rate",meta.get("tax_rate",TAX_RATE)))
        invoices=meta.get("invoices",[]); cos=meta.get("change_orders",[])
        # Separate PO and CO accumulators
        po_qty=0; co_qty=0
        po_cost=0; co_cost=0; tc=0; dc=0; rc=0
        for item in items_list:
            tv=n(item.get("t_num",0)); wv=n(item.get("w_num",0)); lv=n(item.get("length_num",0))
            uc=n(item.get("unit_cost",0)); typ=item.get("type","")
            cf_item=item.get("cost_formula",None)
            # PO portion
            pq=n(item.get("po_qty",0)); po_qty+=pq
            pc=compute_cost(typ,pq,tv,wv,lv,uc,cf_item)
            po_cost+=round(pc*(1+po_tax),2)
            # CO portion
            cq=n(item.get("co_qty",0)); co_qty+=cq
            cc=compute_cost(typ,abs(cq),tv,wv,lv,uc,cf_item)*(1 if cq>=0 else -1)
            co_cost+=round(cc*(1+po_tax),2)
            # Total
            q=pq+cq; tc+=round(compute_cost(typ,q,tv,wv,lv,uc,cf_item)*(1+po_tax),2)
            dt=sum(item.get("deliveries",{}).values())
            dc+=round(compute_cost(typ,dt,tv,wv,lv,uc,cf_item)*(1+del_tax),2)
            rm=max(0,q-dt); rc+=round(compute_cost(typ,rm,tv,wv,lv,uc,cf_item)*(1+po_tax),2)
        rv=[proj, len(items_list),
            po_qty, co_qty, po_qty+co_qty,
            po_cost, co_cost, tc,
            dc, rc,
            len(cos), len(invoices)]
        MONEY_COLS={6,7,8,9,10}
        for ci,v2 in enumerate(rv,1):
            c3=ws_sum.cell(row=sr,column=ci,value=v2)
            c3.font=s["NFONT"]; c3.border=s["bdr"]()  # type: ignore[operator]
            c3.alignment=Alignment(horizontal="left" if ci==1 else "center",vertical="center")
            if ci in MONEY_COLS: c3.number_format=s["MFMT"]
            elif ci==7 and v2<0: c3.font=Font(name="Arial",size=9,color="C0392B")  # CO cost red if negative
        sr+=1
        # Per-project inventory sheet
        sname=proj[:28]+"..." if len(proj)>31 else proj
        ws_p=wb.create_sheet(sname)
        _build_inventory_ws(ws_p, items_list, meta, s)
    for ci,ww in enumerate([28,10,11,11,12,14,14,14,15,15,12,10],1):  # type: ignore[assignment]
        ws_sum.column_dimensions[gcl(ci)].width=ww
    ws_sum.freeze_panes="A2"
    return wb


# ══════════════════════════════════════════════════════════════════════════════
# LABOR MODULE — Building × Scope × % Complete tracking
# ══════════════════════════════════════════════════════════════════════════════

STANDARD_SCOPES = [
    "Layout", "Framing", "Sheathing", "Floor Truss", "Roof Truss",
    "Balcony", "Punchout", "Hardware", "Windows", "Insulation",
    "Drywall", "Roofing", "MEP Rough", "MEP Finish", "Painting", "Other"
]

def load_labor(project: str):
    return _db.load_labor(project) or {"buildings": {}, "subs": [], "last_updated": "", "upload_history": []}

def save_labor(project: str, data: dict):
    _db.save_labor(project, data)


@app.get("/projects/{project}/labor")
def get_labor(project: str):
    return load_labor(project)


@app.put("/projects/{project}/labor/{building}/{scope}")
def update_labor_cell(project: str, building: str, scope: str, body: dict):
    """Manually edit a building×scope labor cell."""
    data = load_labor(project)
    if building not in data["buildings"]:
        data["buildings"][building] = {}
    cell = data["buildings"][building].get(scope, {})
    for k in ["subcontractor", "contract_amount", "billed_pct", "paid_pct", "billed_amount", "paid_amount", "notes"]:
        if k in body:
            cell[k] = body[k]
    # Auto-compute completion from 90% rule
    bp = n(cell.get("billed_pct", 0))
    cell["completed"] = bp >= 90
    cell["status"] = "Complete" if bp >= 90 else ("In Progress" if bp > 0 else "Not Started")
    cell["retainage_pct"] = 10 if bp >= 90 else 0
    # Compute retainage amount
    ca = n(cell.get("contract_amount", 0))
    ba = n(cell.get("billed_amount", cell.get("billed_pct",0) * ca / 100))
    cell["retainage_amount"] = round(ba * 0.10, 2) if cell["completed"] else 0
    data["buildings"][building][scope] = cell
    _recalc_subs(data)
    save_labor(project, data)
    return {"status": "ok", "cell": cell}


def _recalc_subs(data: dict):
    """Recalculate sub payment summaries from building data."""
    sub_map = {}
    for bldg, scopes in data.get("buildings", {}).items():
        for scope, cell in scopes.items():
            sub = cell.get("subcontractor", "")
            if not sub: continue
            if sub not in sub_map:
                sub_map[sub] = {"name": sub, "scopes": [], "contract": 0,
                                "billed": 0, "paid": 0, "retainage": 0}
            s = sub_map[sub]
            if f"{bldg}|{scope}" not in s["scopes"]:
                s["scopes"].append(f"{bldg}|{scope}")
            s["contract"]   += n(cell.get("contract_amount", 0))
            s["billed"]     += n(cell.get("billed_amount", 0))
            s["paid"]       += n(cell.get("paid_amount", 0))
            s["retainage"]  += n(cell.get("retainage_amount", 0))
    data["subs"] = list(sub_map.values())


@app.post("/projects/{project}/upload-labor")
async def upload_labor(project: str, file: UploadFile = File(...),
                       source: str = Form("bt")):
    """Parse BT or Sage weekly Excel export and update labor data.
    source: 'bt' (BuilderTrend) or 'sage' (Sage 300)
    """
    import io, re
    _db.create_project(project)
    raw = await file.read()

    # Extract all sheets from the Excel
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Cannot read Excel: {e}")

    # Build a text representation for Claude to parse
    text_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert ws is not None  # type: ignore
        if ws.max_row < 2: continue
        rows = []
        for r in range(1, min(ws.max_row + 1, 300)):
            row = [str(ws.cell(r, c).value or "").strip()
                   for c in range(1, min(ws.max_column + 1, 20))]
            if any(v for v in row): rows.append(" | ".join(row))
        if rows:
            text_parts.append(f"=== SHEET: {sheet_name} ===\n" + "\n".join(rows))
    full_text = "\n\n".join(text_parts)[:25000]  # cap for token limit

    prompt = (
        f"You are parsing a {'BuilderTrend' if source == 'bt' else 'Sage 300'} "
        f"weekly export Excel for a construction project.\n"
        "Extract EVERY labor line item. Return ONLY a valid JSON array.\n\n"
        "Each object:\n"
        '{"building": string, "scope": string, "subcontractor": string, '
        '"contract_amount": number, "billed_pct": number, "billed_amount": number, '
        '"paid_amount": number, "notes": string}\n\n'
        "RULES:\n"
        "- building: building name/number (e.g. 'Building 1', 'Bldg A', '01')\n"
        "- scope: type of work (Layout/Framing/Sheathing/Floor Truss/Roof Truss/"
        "Balcony/Punchout/Hardware/Windows/Insulation/Drywall/Roofing/MEP Rough/"
        "MEP Finish/Painting/Other)\n"
        "- billed_pct: percentage billed 0-100\n"
        "- billed_amount: dollar amount billed\n"
        "- paid_amount: dollar amount already paid\n"
        "- If a column is missing, use 0\n"
        "- Include ALL rows, all buildings, all scopes\n\n"
        "EXPORT TEXT:\n"
    ) + full_text + "\n\nReturn ONLY the JSON array."

    msg = get_claude().messages.create(
        model="claude-sonnet-4-5", max_tokens=8000,
        messages=[{"role": "user", "content": prompt}])
    rows = safe_json_parse(next((getattr(b,"text","") for b in msg.content if hasattr(b,"text")),"").strip())

    if not rows:
        raise HTTPException(422, "No labor data found in this file.")

    # Merge into labor data
    data = load_labor(project)
    from datetime import date as _date
    today = str(_date.today())
    added = 0
    for row in rows:
        bldg  = str(row.get("building", "Unknown")).strip()
        scope = str(row.get("scope", "Other")).strip()
        if not bldg or not scope: continue
        if bldg not in data["buildings"]: data["buildings"][bldg] = {}
        bp = n(row.get("billed_pct", 0))
        ca = n(row.get("contract_amount", 0))
        ba = n(row.get("billed_amount", 0)) or round(ca * bp / 100, 2)
        pa = n(row.get("paid_amount", 0))
        cell = {
            "subcontractor":   str(row.get("subcontractor", "")).strip(),
            "contract_amount": ca,
            "billed_pct":      bp,
            "billed_amount":   ba,
            "paid_amount":     pa,
            "notes":           str(row.get("notes", "")).strip(),
            "completed":       bp >= 90,
            "status":          "Complete" if bp >= 90 else ("In Progress" if bp > 0 else "Not Started"),
            "retainage_pct":   10 if bp >= 90 else 0,
            "retainage_amount": round(ba * 0.10, 2) if bp >= 90 else 0,
            "last_updated":    today,
            "source":          source,
        }
        data["buildings"][bldg][scope] = cell
        added += 1

    _recalc_subs(data)
    data["last_updated"] = today
    data["upload_history"].append({"date": today, "file": file.filename,
                                   "source": source, "rows": added})
    save_labor(project, data)
    return {"status": "ok", "rows_loaded": added,
            "buildings": len(data["buildings"]),
            "message": f"Loaded {added} labor entries from {file.filename}"}


@app.get("/projects/{project}/dashboard")
def get_dashboard(project: str):
    """Compute full project dashboard: material + labor + schedule combined."""
    items    = load_items(project)
    meta     = load_meta(project)
    labor    = load_labor(project)
    po_tax   = n(meta.get("po_tax_rate",       meta.get("tax_rate", TAX_RATE)))
    del_tax  = n(meta.get("delivery_tax_rate", meta.get("tax_rate", TAX_RATE)))

    # ── Material stats ────────────────────────────────────────────────────
    mat_total = mat_del = mat_rem = 0.0
    for item in items:
        tv=n(item.get("t_num",0)); wv=n(item.get("w_num",0)); lv=n(item.get("length_num",0))
        uc=n(item.get("unit_cost",0)); typ=item.get("type",""); cf=item.get("cost_formula")
        q  = n(item.get("po_qty",0)) + n(item.get("co_qty",0))
        dt = sum(item.get("deliveries",{}).values())
        rm = max(0, q - dt)
        mat_total += round(compute_cost(typ,q, tv,wv,lv,uc,cf)*(1+po_tax), 2)
        mat_del   += round(compute_cost(typ,dt,tv,wv,lv,uc,cf)*(1+del_tax),2)
        mat_rem   += round(compute_cost(typ,rm,tv,wv,lv,uc,cf)*(1+po_tax), 2)
    mat_del_pct = round(mat_del / mat_total * 100, 1) if mat_total > 0 else 0

    # ── Labor stats ───────────────────────────────────────────────────────
    lab_contract = lab_billed = lab_paid = lab_retainage = 0.0
    completed_tasks = total_tasks = 0
    building_summary = {}
    for bldg, scopes in labor.get("buildings", {}).items():
        b_pct_sum = 0; b_count = 0
        for scope, cell in scopes.items():
            total_tasks += 1
            if cell.get("completed"): completed_tasks += 1
            lab_contract  += n(cell.get("contract_amount", 0))
            lab_billed    += n(cell.get("billed_amount", 0))
            lab_paid      += n(cell.get("paid_amount", 0))
            lab_retainage += n(cell.get("retainage_amount", 0))
            b_pct_sum     += n(cell.get("billed_pct", 0))
            b_count       += 1
        building_summary[bldg] = {
            "avg_pct": round(b_pct_sum / b_count, 1) if b_count else 0,
            "tasks_complete": sum(1 for s in scopes.values() if s.get("completed")),
            "tasks_total": b_count,
        }
    lab_billed_pct = round(lab_billed / lab_contract * 100, 1) if lab_contract > 0 else 0

    # ── Risk flags ────────────────────────────────────────────────────────
    risks = []
    if mat_del_pct < lab_billed_pct - 20:
        risks.append({"type":"warning","msg":
            f"Labor is {round(lab_billed_pct-mat_del_pct,0)}% ahead of material delivery — "
            "risk of framing delays. Check material orders."})
    if mat_rem > 0 and mat_del_pct > 80:
        risks.append({"type":"info","msg":
            f"${mat_rem:,.0f} material remaining at {100-mat_del_pct:.0f}% — "
            "verify final delivery schedule."})
    unstarted = [f"{b}:{s}" for b,sc in labor.get("buildings",{}).items()
                 for s,c in sc.items() if c.get("status")=="Not Started" and
                 building_summary.get(b,{}).get("avg_pct",0) > 0]
    if unstarted:
        risks.append({"type":"warning","msg":
            f"{len(unstarted)} scope(s) not started in active buildings: "
            + ", ".join(unstarted[:3]) + ("..." if len(unstarted)>3 else "")})

    # ── Payment due this draw ─────────────────────────────────────────────
    payment_due = []
    for sub in labor.get("subs", []):
        due = n(sub.get("billed",0)) - n(sub.get("paid",0)) - n(sub.get("retainage",0))
        if due > 0:
            payment_due.append({"sub": sub["name"],
                                 "billed": round(n(sub.get("billed",0)),2),
                                 "paid":   round(n(sub.get("paid",0)),2),
                                 "due":    round(due,2),
                                 "retainage": round(n(sub.get("retainage",0)),2)})
    payment_due.sort(key=lambda x: -x["due"])

    return {
        "project": project,
        "material": {
            "total_cost_tax": round(mat_total, 2),
            "delivered_cost_tax": round(mat_del, 2),
            "remaining_cost_tax": round(mat_rem, 2),
            "pct_delivered": mat_del_pct,
            "items": len(items),
            "invoices": len(meta.get("invoices", [])),
            "change_orders": len(meta.get("change_orders", [])),
        },
        "labor": {
            "contract_total": round(lab_contract, 2),
            "billed_total": round(lab_billed, 2),
            "paid_total": round(lab_paid, 2),
            "retainage_total": round(lab_retainage, 2),
            "pct_billed": lab_billed_pct,
            "tasks_complete": completed_tasks,
            "tasks_total": total_tasks,
            "buildings": building_summary,
        },
        "payment_due": payment_due,
        "risks": risks,
        "last_labor_update": labor.get("last_updated", ""),
        "gsf": n(meta.get("gsf", 0)),
        "state": meta.get("state", ""),
        "building_type": meta.get("building_type", ""),
        "completed": meta.get("completed", False),
    }


# ══════════════════════════════════════════════════════════════════════════════
# BT (BUILDERTREND) MODULE — Estimate, PO/VPO, Labor & GC Summaries
# ══════════════════════════════════════════════════════════════════════════════
import re as _re

def load_bt_estimate(project: str):
    data = _db.load_bt_estimate(project)
    if not data: return []
    if isinstance(data, dict):
        return data.get("rows", [])
    return data

def load_bt_pos(project: str):
    return _db.load_bt_pos(project) or []

def save_bt_estimate(project: str, data): _db.save_bt_estimate(project, data)
def save_bt_pos(project: str, data):      _db.save_bt_pos(project, data)

# ── Helpers ────────────────────────────────────────────────────────────────
def _clean_cost_code(cc: str) -> str:
    """Normalize cost code — strip newlines, take first line only."""
    if not cc: return ""
    return str(cc).split("\n")[0].strip()


# Known building names — extracted from BT estimate titles by exact prefix match
_KNOWN_BUILDINGS = [
    "Building VA (North)", "Building VA (South)", "Club house VB (North)",
    "Mail Kiosk (North)", "Mail Kiosk (South)",
    "Trash Enclosure (North)", "Trash Enclosure (South)",
]

# Generic building patterns — matched if no known building is found.
# Captures things like: "Bldg.1", "Bldg 2", "Building 3", "Building #4",
# "Block A", "Tower 2", "Unit 5", "House 1", "Villa 3", "Phase 1", "Lot 12".
_BUILDING_PATTERNS = [
    re.compile(r"^(Bldg\.?\s*[\w#-]+(?:\s*\([^)]+\))?)", re.IGNORECASE),
    re.compile(r"^(Building\s*#?\s*[\w-]+(?:\s*\([^)]+\))?)", re.IGNORECASE),
    re.compile(r"^(Block\s+[\w-]+(?:\s*\([^)]+\))?)", re.IGNORECASE),
    re.compile(r"^(Tower\s+[\w-]+(?:\s*\([^)]+\))?)", re.IGNORECASE),
    re.compile(r"^(Unit\s+[\w-]+(?:\s*\([^)]+\))?)", re.IGNORECASE),
    re.compile(r"^(House\s+[\w-]+(?:\s*\([^)]+\))?)", re.IGNORECASE),
    re.compile(r"^(Villa\s+[\w-]+(?:\s*\([^)]+\))?)", re.IGNORECASE),
    re.compile(r"^(Phase\s+[\w-]+(?:\s*\([^)]+\))?)", re.IGNORECASE),
    re.compile(r"^(Lot\s+[\w-]+(?:\s*\([^)]+\))?)", re.IGNORECASE),
    re.compile(r"^(Club\s*house\s+[\w()-]+(?:\s*\([^)]+\))?)", re.IGNORECASE),
    re.compile(r"^(Mail\s+Kiosk(?:\s*\([^)]+\))?)", re.IGNORECASE),
    re.compile(r"^(Trash\s+Enclosure(?:\s*\([^)]+\))?)", re.IGNORECASE),
]

def _normalize_building(name: str) -> str:
    """Canonicalize a building name so 'Bldg. 1', 'Bldg 1', 'Bldg.1', 'Building 1', 'Building #1'
    all map to the same string ('Bldg 1'). This is what fixes the screenshot bug where
    labor was bucketed under 'Bldg. 1' and material under 'Bldg 1'.
    Returns "" for empty input."""
    if not name: return ""
    s = str(name).strip()
    if not s: return ""
    # Strip surrounding whitespace + collapse internal whitespace
    s = re.sub(r"\s+", " ", s)
    # Bldg./Bldg./Building variants → canonical "Bldg N"
    # "Bldg.1" / "Bldg. 1" / "Bldg 1" / "Building 1" / "Building #1" → "Bldg 1"
    m = re.match(r"^(?:Bldg\.?|Building)\s*#?\s*([\w\-]+)(\s*\([^)]+\))?$", s, re.IGNORECASE)
    if m:
        num = m.group(1).strip().rstrip("-_.")  # strip stray trailing punct (e.g. "Bldg 3-" → "Bldg 3")
        suffix = (m.group(2) or "").strip()
        # If number is purely numeric/letter, prefix "Bldg "; else fall through
        canonical = f"Bldg {num}"
        if suffix: canonical += f" {suffix}"
        return canonical
    return s

def _extract_building(title: str) -> str:
    """Match building from a scope title. Works on:
       - Known buildings (Willow Way etc.):  'Building VA (North) Lumber...' -> 'Building VA (North)'
       - Generic patterns (Cobia Cove etc.): 'Bldg.1 Layout - L1'           -> 'Bldg 1'  (normalized)
                                              'Bldg. 1 Layout - L1'          -> 'Bldg 1'
                                              'Building 1 Lumber'            -> 'Bldg 1'
                                              'Building 3 Framing'           -> 'Bldg 3'
                                              'Bldg 2 (North) Trusses'       -> 'Bldg 2 (North)'
       Returns "" if no building can be identified (e.g. Nails Framing, GC items)."""
    t = str(title).strip()
    if not t: return ""
    # 1) Known full names (Willow Way) — keep exact spelling, no normalization
    for b in _KNOWN_BUILDINGS:
        if t.lower().startswith(b.lower()):
            return b
    # 2) Generic patterns → normalize so variants merge
    for pat in _BUILDING_PATTERNS:
        m = pat.match(t)
        if m:
            return _normalize_building(m.group(1).strip())
    return ""

def _extract_building_set_from_rows(rows) -> list:
    """Scan all rows once and return a sorted unique list of buildings found.
    Used to dynamically know which buildings exist in this project."""
    seen = []
    seen_lower = set()
    for r in rows:
        title = r.get("title", "") if isinstance(r, dict) else str(r)
        b = _extract_building(title)
        if b and b.lower() not in seen_lower:
            seen.append(b)
            seen_lower.add(b.lower())
    return seen

def _extract_level(title: str) -> str:
    m = _re.search(r"[-\u2013]\s*L(\d+)", str(title))
    return f"L{m.group(1)}" if m else ""

def _code_category(cost_code: str) -> str:
    num = _clean_cost_code(cost_code).split(" - ")[0].strip()
    try:
        n = int(num)
        return "gc" if n < 100 else "labor" if n < 200 else "material"
    except:
        return "other"


# ══════════════════════════════════════════════════════════════════════════════
# BT ESTIMATE + SOV GENERATOR  (v2 — matches the manual Willow Way workflow)
#
# Input: ONE workbook with two sheets:
#   • "Estimate"  — quantities laid out as Building (row 1) × Level (row 2) × Scope (col A)
#   • "Buyout" / "Rev buyout" — builder unit-cost rates per scope + GC items, AND a
#     "Scope - Rough Carpentry" section holding the client CONTRACT sell prices
#     (Labor+GC package, Lumber, Hardware, Truss, Material Credit).
#
# Output: ONE Excel with two sheets (BT Estimate + SOV) containing LIVE FORMULAS,
#   stored as the project's bt_estimate.json so the BT Report works.
#
# Column model (matches the user's sheet exactly):
#   E Unit Cost | F Quantity | H Builder=E*F | I Markup | J Owner=H*(1+I)
#   K Mobilization=J*3% | L New Value=J-K
#
# Nails line is the PLUG that makes Labor+GC owner-cost hit the Labor contract:
#   J(nails) = Labor_Contract - SUM(all GC+labor owner costs above)
#   I(nails) = J/H - 1   (must land 20%-40%; auto-solver nudges labor markups if not)
#
# SOV: Mobilization row = SUM(all K). Every other line = New Value(L) + dist,
#   where dist = (sum GC New Values + Nails New Value) / (number of SOV lines).
# ══════════════════════════════════════════════════════════════════════════════

def _norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())

def _gbt_find_buyout(buyout, *keywords):
    """First buyout row whose description contains ALL keywords (case-insensitive).
    Hyphens are treated as optional spaces so 'sheathing - demising' matches
    'Sheathing - Demising Plywood' and 'Sheathing Demising Plywood'."""
    import re as _re2
    def _nk(k): return _re2.sub(r'[\s\-]+', ' ', k.strip().lower())
    norm_keys = [_nk(k) for k in keywords if k.strip() and k.strip() != '-']
    for b in buyout:
        d = _re2.sub(r'[\s\-]+', ' ', _norm(b["desc"]))
        if all(k in d for k in norm_keys):
            return b
    return None

# Scope → cost code / category / markup-tier mapping.
# Markup defaults match the manual Willow Way sheet — front-loaded but low enough
# that GC+labor owner cost stays UNDER the labor contract, leaving a positive nails
# residual in the 20-40% range. (Layout 30%, Framing 30%, then tapering down.)
_GBT_LABOR_SCOPES = [
    ("corridor wall sheathing",  "sheathing - corridor",  "104 - Shear Wall - Labor", "01 - Framing", "Corridor Plywood Sheathing", 0.55),
    ("demising wall",            "sheathing - demising",  "104 - Shear Wall - Labor", "01 - Framing", "Demising Plywood Sheathing", 0.55),
    ("stair wall sheathing",     "sheathing - stair",     "104 - Shear Wall - Labor", "01 - Framing", "Stair Plywood Sheathing",    0.55),
    ("unit floor framing",       "floor truss - unit",    "106 - Floor Truss - Labor","01 - Framing", "Unit Floor Truss",           0.45),
    ("corridor floor framing",   "floor truss - corridor","106 - Floor Truss - Labor","01 - Framing", "Corridor Floor Truss",       0.40),
    ("balcony floor framing",    "balcony",               "108 - Balcony - Labor",    "01 - Framing", "Balcony",                    0.40),
    ("roof framing",             "roof truss",            "107 - Roof Truss - Labor", "01 - Framing", "Roof Truss",                 0.35),
]
# Per-building labor priced off GSF (Layout/Framing/Hardware/Punchout) — markup defaults per sheet
_GBT_PER_BLDG_LABOR = [
    ("hardware",  "hardware",  "109 - Hardware - Labor",  "Hardware", 0.15),
    ("punchout",  "punchout",  "115 - Punchout - Labor",  "Punchout", 0.20),
]
# GC items (priced per-month or per-SF). markup default = gc_markup (40%).
_GBT_GC_ITEMS = [
    ("site supervision", "001 - Supervision", "Site Supervision", "Allowance", "month"),
    ("site office",      "004 - Site Office", "Site Office",      "Allowance", "month"),
    ("hotels",           "005 - Housing",     "Housing - Management","Allowance","month"),
    ("forklift",         "006 - Forklift",    "Forklift",         "Equipment", "month"),
    ("manlift",          "007 - Manlift",     "Manlift",          "Equipment", "month"),
    ("crane",            "008 - Crane",       "Crane",            "Equipment", "sf"),
    ("fuel",             "013 - Fuel",        "Fuel",             "Allowance", "month"),
]
# Default markups for the GSF-priced scopes (Layout, Framing) — front-loaded
_GBT_LAYOUT_MARKUP = 0.30
_GBT_FRAMING_MARKUP = 0.30

# ── FRONT-LOAD MARKUP LADDER (by construction sequence) ──
# Higher markups on early scopes → collect more cash at the start of the job.
# These are the BASE markups; the nails solver scales the whole ladder by one
# factor so the Nails residual lands in 20-40%. Order = build sequence.
_GBT_FRONTLOAD_LADDER = {
    "Layout":                                    0.85,
    "Wall Framing":                              0.75,
    "Sheathing":                                 0.60,
    "Corridor Plywood Sheathing":                0.55,
    "Corridor seperation Plywood Sheathing":     0.55,
    "Demising Plywood Sheathing":                0.55,
    "Stair Plywood Sheathing":                   0.55,
    "Unit Floor Truss":                          0.45,
    "Corridor Floor Truss":                      0.40,
    "Balcony":                                   0.40,
    "Roof Truss":                                0.35,
    "House Wrap":                                0.30,
    "Hardware":                                  0.20,
    "Stair":                                     0.20,
    "Punchout":                                  0.05,
}
def _ladder_markup(scope_label, scale=1.0):
    """Front-loaded markup for a scope, scaled by the solver's factor."""
    base = _GBT_FRONTLOAD_LADDER.get(scope_label, 0.30)
    return max(0.05, round(base * scale, 4))


def _gbt_parse_buyout(ws_buy):
    """Parse a buyout sheet → (rate_rows, scope_sov). rate_rows = list of dicts with
    desc/type/unit/unit_cost. scope_sov = dict of package → {cost, sell} from the
    'Scope - Rough Carpentry' section."""
    rate_rows = []
    scope_sov = {}
    # Find the Buy-Out section header (Description / Unit cost). Prefer the LAST one
    # (the actual buy-out rates, not the higher internal-estimate rates).
    hdr_rows = []
    for r in range(1, ws_buy.max_row + 1):
        rv = [_norm(ws_buy.cell(r, c).value) for c in range(1, ws_buy.max_column + 1)]
        if "description" in rv and "unit cost" in rv:
            hdr_rows.append(r)
    # Use the LAST header (buy-out section). Build column map.
    if hdr_rows:
        hr = hdr_rows[-1]
        col = {}
        for c in range(1, ws_buy.max_column + 1):
            v = _norm(ws_buy.cell(hr, c).value)
            if v in ("description","company","type","unit","unit cost","quantity","total cost"):
                col[v] = c
        for r in range(hr + 1, ws_buy.max_row + 1):
            desc = ws_buy.cell(r, col.get("description", 1)).value
            if not desc: 
                # blank row likely ends the section
                # but keep scanning a couple rows in case of gaps
                continue
            cost = ws_buy.cell(r, col.get("unit cost", 0)).value if col.get("unit cost") else None
            qty  = ws_buy.cell(r, col.get("quantity", 0)).value if col.get("quantity") else None
            rate_rows.append({
                "desc": str(desc).strip(),
                "type": str(ws_buy.cell(r, col.get("type", 0)).value or "").strip() if col.get("type") else "",
                "unit": str(ws_buy.cell(r, col.get("unit", 0)).value or "").strip() if col.get("unit") else "",
                "unit_cost": float(cost) if isinstance(cost, (int, float)) else 0.0,
                "qty": float(qty) if isinstance(qty, (int, float)) else 0.0,
            })

    # Find the "Scope - Rough Carpentry" SOV section
    for r in range(1, ws_buy.max_row + 1):
        if "scope" in _norm(ws_buy.cell(r, 1).value) and "carpentry" in _norm(ws_buy.cell(r, 1).value):
            # header at row r; columns: Scope|Unit|Rate|Quantity|Cost|Markup|Sell Price...
            colmap = {}
            for c in range(1, ws_buy.max_column + 1):
                v = _norm(ws_buy.cell(r, c).value)
                if v in ("scope - rough carpentry","unit","rate","quantity","cost","markup","sell price"):
                    colmap[v] = c
            cost_c = colmap.get("cost"); sell_c = colmap.get("sell price")
            for rr in range(r + 1, ws_buy.max_row + 1):
                scope = ws_buy.cell(rr, 1).value
                if not scope: continue
                sn = _norm(scope)
                if sn.startswith("total"): break
                cost = ws_buy.cell(rr, cost_c).value if cost_c else None
                sell = ws_buy.cell(rr, sell_c).value if sell_c else None
                scope_sov[sn] = {
                    "label": str(scope).strip(),
                    "cost": float(cost) if isinstance(cost,(int,float)) else 0.0,
                    "sell": float(sell) if isinstance(sell,(int,float)) else 0.0,
                }
            break

    return rate_rows, scope_sov


@app.post("/projects/{project}/generate-bt-estimate")
async def generate_bt_estimate(project: str, file: UploadFile = File(...),
                               gc_markup: float = Form(0.30),
                               months: int = Form(3),
                               labor_contract: float = Form(0.0),
                               material_packages_json: str = Form(""),
                               markups_json: str = Form("")):
    """Generate BT Estimate + SOV from an Estimate + Buyout workbook.

    months          = project duration (drives per-month GC items)
    gc_markup       = default GC markup
    labor_contract  = the Labor+GC client contract price (e.g. 565925.56). If 0, the
                      tool reads it from the buyout 'Scope - Rough Carpentry' section,
                      else falls back to (sum of buyout labor+GC cost) × (1+gc_markup).
    material_packages_json = optional JSON list of material packages the user typed:
                      [{"label":"Lumber Package","cost":320623,"sell":473900.52},...]
    markups_json    = optional JSON of per-scope markup overrides keyed by cost code,
                      e.g. {"101 - Layout - Labor":0.70,...}, and per-line overrides
                      keyed by exact title."""
    import io
    _db.create_project(project)
    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Cannot read workbook: {e}")

    overrides = {}
    if markups_json:
        try: overrides = json.loads(markups_json)
        except Exception: overrides = {}
    user_materials = []
    if material_packages_json:
        try: user_materials = json.loads(material_packages_json)
        except Exception: user_materials = []

    def _sheet(*names):
        for n in names:
            for sn in wb.sheetnames:
                if _norm(sn) == _norm(n): return wb[sn]
        for n in names:
            for sn in wb.sheetnames:
                if _norm(n) in _norm(sn): return wb[sn]
        return None

    ws_est = _sheet("Estimate")
    ws_buy = _sheet("Rev buyout", "Buyout", "Buy-out", "(5) Buyout", "Bid buyout")
    if ws_est is None:
        raise HTTPException(422, "No 'Estimate' sheet found.")
    if ws_buy is None:
        raise HTTPException(422, "No 'Buyout' sheet found.")

    rate_rows, scope_sov = _gbt_parse_buyout(ws_buy)
    if not rate_rows:
        raise HTTPException(422, "Buyout sheet: could not find the Buy-Out rate section.")

    # Merge any material packages the user typed into scope_sov
    for mp in user_materials:
        key = _norm(mp.get("label",""))
        if key:
            scope_sov[key] = {"label": mp.get("label",""),
                              "cost": float(mp.get("cost",0) or 0),
                              "sell": float(mp.get("sell",0) or 0)}

    # Contract target: explicit param > Scope section > fallback (labor+GC cost ×1+gc)
    if labor_contract <= 0:
        for k, v in scope_sov.items():
            if "labor" in k:
                labor_contract = v["sell"]; break
    if labor_contract <= 0:
        # fallback: sum buyout labor+GC builder cost, apply gc_markup
        lg = sum(r["unit_cost"]*r["qty"] for r in rate_rows
                 if r["type"].lower() in ("labor","overhead","equipment") and "nail" not in _norm(r["desc"]))
        labor_contract = round(lg * (1 + gc_markup), 2)

    # ── Parse estimate building/level header ──
    col_meta = {}
    cur_b = ""

    # ── Smart header detection ──────────────────────────────────────────────────
    # The Estimate sheet can have either:
    #   2-row header: Row1 = Building name, Row2 = Level (old format)
    #   3-row header: Row1 = Building type (merged), Row2 = BLDG name (merged), Row3 = Level 1/2/3...
    # We detect by checking whether row 3 contains level labels.
    # Also handle merged cells: openpyxl only stores values in the top-left cell of a merge.
    # So "Building Type I" in C1:H1 means C1 has the value, D1-H1 are None.
    # We propagate the last seen non-None value rightward for merged building headers.

    def _row_vals(r):
        return [ws_est.cell(r, c).value for c in range(1, ws_est.max_column + 2)]

    row1 = _row_vals(1)
    row2 = _row_vals(2)
    row3 = _row_vals(3)

    # Detect if row 3 contains level labels (e.g. "Level 1", "L1", "Level 2"...)
    def _has_levels(rv):
        return any(rv[i] and re.search(r"level\s*\d", str(rv[i]).lower())
                   for i in range(len(rv)))

    use_3row = _has_levels(row3)
    level_row_num = 3 if use_3row else 2
    bldg_row_num  = 2 if use_3row else 1

    # Build col_meta: for each data column, track (building_name, level_label)
    # Propagate building name rightward across merged cells (None → carry last value)
    cur_b = ""
    cur_bldg_label = ""  # e.g. "BLDG 1"
    for c in range(3, ws_est.max_column + 1):
        b1v = ws_est.cell(1, c).value        # Building type row
        b2v = ws_est.cell(bldg_row_num, c).value  # Building name row
        lvv = ws_est.cell(level_row_num, c).value  # Level row

        # Propagate building type (handles C1:H1 merge)
        if b1v: cur_b = str(b1v).strip()
        # Propagate building name/label (handles C2:H2 merge)
        if b2v: cur_bldg_label = str(b2v).strip()

        lvl = str(lvv).strip() if lvv else ""
        if not lvl:
            continue  # column has no level label — skip

        # Normalize level: "Level 1" → "L1", "Roof" → "Roof", "BLDG 1" if only 1 col
        m = re.search(r"level\s*(\d+)", lvl.lower())
        if m:
            ln = f"L{m.group(1)}"
        elif "roof" in lvl.lower():
            ln = "Roof"
        else:
            # Could be a plain building column (no level breakdown) — use as-is
            ln = lvl

        # Building display name: prefer "Building I" style from "Building Type I"
        bld_display = re.sub(r"building type\s*", "Building ", cur_b, flags=re.I).strip()
        if not bld_display:
            bld_display = cur_bldg_label or "Building I"

        col_meta[c] = (bld_display, ln)

    def _bld_short(b):
        # Already normalized in col_meta — just return as-is
        # (handles "Building I", "Building VA (North)" etc.)
        return str(b).strip()

    # ── Parse estimate scope rows + find the GSF row ──
    # Layout / Framing / Hardware / Punchout all use the per-floor GSF (a dedicated
    # "GSF" row in the Estimate), NOT the exterior-wall-stud SF. We locate that row
    # by its label and read GSF per building/level column.
    est_rows = []
    gsf_row_idx = None
    for r in range(3, ws_est.max_row + 1):
        desc = ws_est.cell(r, 1).value
        unit = ws_est.cell(r, 2).value
        if desc and _norm(desc) == "gsf":
            gsf_row_idx = r
        if not desc: continue
        desc = str(desc).strip()
        if desc.endswith(":") and not unit: continue
        qbc = {}
        cellref = {}
        for c in col_meta:
            v = ws_est.cell(r, c).value
            if isinstance(v, (int, float)) and v:
                qbc[c] = float(v)
        if qbc:
            est_rows.append({"desc": desc, "unit": str(unit or "").strip(),
                             "qty_by_col": qbc, "row_idx": r})

    # GSF per (building, level) — the master quantity for Layout/Framing/Hardware/Punchout.
    # Track the source cell (e.g. "C51") so the download can link the qty with a formula.
    from openpyxl.utils import get_column_letter as _gcl
    gsf_by_bl = {}
    gsf_cell_by_bl = {}   # (bld,lvl) -> "C51" style ref into the Estimate sheet
    if gsf_row_idx:
        for c, (bld, lvl) in col_meta.items():
            v = ws_est.cell(gsf_row_idx, c).value
            if isinstance(v, (int, float)) and v:
                key = (_bld_short(bld), lvl)
                gsf_by_bl[key] = float(v)
                gsf_cell_by_bl[key] = f"{_gcl(c)}{gsf_row_idx}"
    else:
        # Fallback: if no explicit GSF row, derive from exterior wall stud SF
        for er in est_rows:
            if "exterior wood stud wall" in _norm(er["desc"]):
                for c, q in er["qty_by_col"].items():
                    bld, lvl = col_meta[c]
                    gsf_by_bl[(_bld_short(bld), lvl)] = gsf_by_bl.get((_bld_short(bld), lvl), 0) + q

    # wall_sf is now the GSF map (kept name for the rest of the function)
    wall_sf = gsf_by_bl
    total_gsf = sum(wall_sf.values())
    # Nails GSF = sum of the ROOF-level GSF columns (matches the sheet's J57 =SUM of Roof GSF)
    nails_gsf = sum(v for (b, l), v in wall_sf.items() if str(l).lower() == "roof")
    if nails_gsf <= 0:
        nails_gsf = total_gsf

    def _markup_for(cost_code, title, default):
        # per-line override wins, then per-scope override, then default
        if title in overrides: return float(overrides[title])
        if cost_code in overrides: return float(overrides[cost_code])
        return default

    bt_rows = []  # ordered list of dicts

    # 1) GENERAL CONDITIONS
    for bo_kw, cc, title, ctype, basis in _GBT_GC_ITEMS:
        b = _gbt_find_buyout(rate_rows, bo_kw)
        if not b: continue
        if basis == "sf":
            qty = b["qty"] or total_gsf
            unit = "SF"; uc = b["unit_cost"]
        else:
            qty = months; unit = "Months"; uc = b["unit_cost"]
        mk = _markup_for(cc, title, gc_markup)
        bt_rows.append({"category":"00 - General Conditions","cost_code":cc,"title":title,
                        "cost_type":ctype,"unit_cost":uc,"qty":round(qty,2),"unit":unit,
                        "builder":round(uc*qty,2),"markup":mk,"is_gc":True,"is_nails":False,
                        "scope":"GC"})

    # 2) LABOR — Layout + Wall Framing + Sheathing-Exterior all priced off GSF.
    #    These exist for FLOOR levels only (L1/L2/L3) — NOT the Roof level.
    #    Markups come from the front-load ladder (Layout highest, tapering down).
    for cc, scope_label, bo_kw in [
        ("101 - Layout - Labor",   "Layout",    "layout"),
        ("102 - Framing - Labor",  "Wall Framing", "framing"),
        ("103 - Sheathing - Labor","Sheathing", "sheathing - exterior"),
    ]:
        b = _gbt_find_buyout(rate_rows, *bo_kw.split()) or _gbt_find_buyout(rate_rows, bo_kw)
        if not b: continue
        for (bld, lvl), sf in sorted(wall_sf.items()):
            if str(lvl).lower() == "roof": continue   # no Layout/Framing/Sheathing-ext on roof
            title = f"{bld} {scope_label} - {lvl}"
            mk = _markup_for(cc, title, _ladder_markup(scope_label))
            bt_rows.append({"category":"01 - Framing","cost_code":cc,"title":title,
                            "cost_type":"Labor","unit_cost":b["unit_cost"],"qty":round(sf,2),
                            "unit":"SF","builder":round(b["unit_cost"]*sf,2),"markup":mk,
                            "is_gc":False,"is_nails":False,"scope":scope_label,
                            "qty_cell":gsf_cell_by_bl.get((bld,lvl))})

    # scope-mapped labor using its OWN estimate-row quantity (not GSF).
    # Sheathing panels: divide SF by 32 (4×8 sheet = 32 SF).
    # Buyout rates for these are per Sheet (plywood) or per SF (floor truss).
    _SCOPE_OWN_QTY = [
        # (estimate desc keyword,         buyout keyword,           cost_code,                   scope_label,                  default_mk, divide)
        ("corridor wall sheathing",        "sheathing - corridor",   "104 - Shear Wall - Labor",  "Corridor Plywood Sheathing", 0.55, 32),
        ("demising wall",                  "sheathing - demising",   "104 - Shear Wall - Labor",  "Demising Plywood Sheathing", 0.55, 32),
        ("stair wall sheathing",           "sheathing - stair",      "104 - Shear Wall - Labor",  "Stair Plywood Sheathing",    0.55, 32),
        ("unit floor framing",             "floor truss - unit",     "106 - Floor Truss - Labor", "Unit Floor Truss",           0.45, 1),
        ("corridor floor framing",         "floor truss - corridor", "106 - Floor Truss - Labor", "Corridor Floor Truss",       0.40, 1),
        ("balcony floor framing",          "balcony",                "108 - Balcony - Labor",     "Balcony",                    0.40, 1),
        ("roof framing",                   "roof truss",             "107 - Roof Truss - Labor",  "Roof Truss",                 0.35, 1),
    ]
    for er in est_rows:
        d = _norm(er["desc"])
        for kw, bo_kw, cc, scope_label, front, divide in _SCOPE_OWN_QTY:
            if kw in d:
                if "roof truss" in cc.lower() and "shingle" in d:
                    continue
                # Try split search first (handles "floor truss - unit" → ["floor truss", "unit"])
                parts = bo_kw.split(" - ")
                b = _gbt_find_buyout(rate_rows, *parts) if len(parts) > 1 else None
                if not b:
                    b = _gbt_find_buyout(rate_rows, *bo_kw.split())
                if not b:
                    b = _gbt_find_buyout(rate_rows, bo_kw)
                if not b:
                    break
                for c, q in er["qty_by_col"].items():
                    bld, lvl = col_meta[c]
                    qty = round(q / divide, 4)
                    unit = "Sheets" if divide == 32 else (b["unit"] or er["unit"])
                    title = f"{_bld_short(bld)} {scope_label} - {lvl}"
                    mk = _markup_for(cc, title, _ladder_markup(scope_label))
                    qcell = f"{_gcl(c)}{er['row_idx']}"
                    bt_rows.append({"category":"01 - Framing","cost_code":cc,"title":title,
                                    "cost_type":"Labor","unit_cost":b["unit_cost"],"qty":qty,
                                    "unit":unit,"builder":round(b["unit_cost"]*qty,2),"markup":mk,
                                    "is_gc":False,"is_nails":False,"scope":scope_label,
                                    "qty_cell":qcell,"qty_divide":divide})
                break

    # per-building labor (hardware, punchout) — priced off GSF, FLOOR levels only (no roof)
    for kw, bo_kw, cc, scope_label, front in _GBT_PER_BLDG_LABOR:
        b = _gbt_find_buyout(rate_rows, bo_kw)
        if not b: continue
        for (bld, lvl), sf in sorted(wall_sf.items()):
            if str(lvl).lower() == "roof": continue   # no hardware/punchout on roof
            title = f"{bld} {scope_label} - {lvl}"
            mk = _markup_for(cc, title, _ladder_markup(scope_label))
            bt_rows.append({"category":"01 - Framing","cost_code":cc,"title":title,
                            "cost_type":"Labor","unit_cost":b["unit_cost"],"qty":round(sf,2),
                            "unit":"SF","builder":round(b["unit_cost"]*sf,2),"markup":mk,
                            "is_gc":False,"is_nails":False,"scope":scope_label,
                            "qty_cell":gsf_cell_by_bl.get((bld,lvl))})

    # House Wrap (110) + Stair (113) — per FLOOR level, from buyout rate.
    # House Wrap is a per-floor allowance (rolls); Stair is per floor-level landing.
    # These only apply to multi-story main buildings (those with L2/L3).
    multistory = sorted({b for (b, l) in wall_sf if str(l).upper() in ("L2", "L3")})
    hw_bo = _gbt_find_buyout(rate_rows, "zip", "sheathing") or _gbt_find_buyout(rate_rows, "house", "wrap")
    if hw_bo:
        # House wrap priced per floor level off GSF (uses the zip/house-wrap rate)
        for (bld, lvl), sf in sorted(wall_sf.items()):
            if str(lvl).lower() == "roof": continue
            cc = "110 - House Wrap - Labor"
            title = f"{bld} House Wrap - {lvl}"
            mk = _markup_for(cc, title, _ladder_markup("House Wrap"))
            bt_rows.append({"category":"01 - Framing","cost_code":cc,"title":title,
                            "cost_type":"Labor","unit_cost":hw_bo["unit_cost"],"qty":round(sf,2),
                            "unit":"SF","builder":round(hw_bo["unit_cost"]*sf,2),"markup":mk,
                            "is_gc":False,"is_nails":False,"scope":"House Wrap"})
    stair_bo = _gbt_find_buyout(rate_rows, "stairs", "landings") or _gbt_find_buyout(rate_rows, "stair")
    if stair_bo:
        for bld in multistory:
            for lvl in ("L1", "L2", "L3"):
                if (bld, lvl) not in wall_sf: continue
                cc = "113 - Stair - Labor"
                title = f"{bld} Stair-landings - {lvl}"
                mk = _markup_for(cc, title, _ladder_markup("Stair"))
                bt_rows.append({"category":"01 - Framing","cost_code":cc,"title":title,
                                "cost_type":"Labor","unit_cost":stair_bo["unit_cost"],"qty":1,
                                "unit":"Each","builder":round(stair_bo["unit_cost"]*1,2),"markup":mk,
                                "is_gc":False,"is_nails":False,"scope":"Stair"})

    # 3) Compute Owner/Mob/NewValue for GC + labor (so far)
    def _recompute(row):
        row["owner"] = round(row["builder"]*(1+row["markup"]), 2)
        row["mob"]   = round(row["owner"]*0.03, 2)
        row["newval"]= round(row["owner"]-row["mob"], 2)
    for row in bt_rows: _recompute(row)

    # 4) NAILS PLUG — make (GC + labor) owner cost hit the labor contract.
    #    Nails owner = labor_contract - SUM(all GC+labor owner above).  With the
    #    sheet's front-loaded-but-modest markups this lands POSITIVE and in 20-40%.
    #    If it doesn't, scale the WHOLE front-load ladder by ONE factor (preserving the
    #    descending front-loaded shape — Layout stays highest) until nails lands at 30%.
    nails_bo = _gbt_find_buyout(rate_rows, "nails")
    nails_row = None
    if nails_bo and labor_contract > 0:
        nails_builder = round(nails_bo["unit_cost"] * nails_gsf, 2)

        # Store each labor line's BASE ladder markup so we can rescale proportionally.
        for r in bt_rows:
            if not r["is_gc"]:
                r["_base_mk"] = _GBT_FRONTLOAD_LADDER.get(r.get("scope"), r["markup"])

        def _apply_scale(scale):
            for r in bt_rows:
                if r["is_gc"]: continue
                # honor any explicit per-line/per-scope override (don't rescale those)
                ov_title = r["title"] in overrides
                ov_code  = r["cost_code"] in overrides
                if ov_title or ov_code: continue
                r["markup"] = max(0.05, round(r["_base_mk"] * scale, 4))
                _recompute(r)

        def _nails_markup_at(scale):
            _apply_scale(scale)
            owner_sum = sum(r["owner"] for r in bt_rows)  # GC + labor
            n_owner = labor_contract - owner_sum
            return (n_owner / nails_builder - 1) if nails_builder else 0, n_owner

        # Binary-search the ladder scale so nails markup → 0.30 (center of 20-40%).
        lo_s, hi_s = 0.05, 3.0
        best_scale = 1.0
        for _ in range(40):
            mid = (lo_s + hi_s) / 2
            nm, _no = _nails_markup_at(mid)
            if nm > 0.30:
                # labor markups too low → nails too high → raise labor markups
                lo_s = mid
            else:
                hi_s = mid
            best_scale = mid
            if abs(nm - 0.30) < 0.002:
                break
        nails_markup, nails_owner = _nails_markup_at(best_scale)

        # Clamp into [0.20, 0.40] band; never negative.
        if nails_markup < 0.20 or nails_markup > 0.40 or nails_owner < 0:
            # final guard: set nails to 30% and leave labor at last scale
            if nails_owner < 0:
                nails_owner = round(nails_builder * 1.20, 2); nails_markup = 0.20
            else:
                nails_markup = max(0.20, min(0.40, nails_markup))

        # clean up temp field
        for r in bt_rows: r.pop("_base_mk", None)

        nails_row = {"category":"01 - Framing","cost_code":"206 - Nails - Framing","title":"Nails Framing",
                     "cost_type":"Labor","unit_cost":nails_bo["unit_cost"],"qty":round(nails_gsf,2),
                     "unit":"SF","builder":nails_builder,"markup":round(nails_markup,4),
                     "is_gc":False,"is_nails":True,"scope":"Nails","owner":round(nails_owner,2)}
        nails_row["mob"]    = round(nails_row["owner"]*0.03, 2)
        nails_row["newval"] = round(nails_row["owner"]-nails_row["mob"], 2)
        bt_rows.append(nails_row)

    # 5) MATERIAL packages (Lumber, Hardware-material, Truss) — back-compute rate
    #    Rate = package cost ÷ material GSF, where material GSF = the building
    #    footprint GSF (the Roof-level GSF, same basis as nails). Material lines are
    #    priced per building using that building's footprint GSF. Markup = sell/cost − 1.
    # Build per-building footprint GSF (roof level), keyed by building name.
    # Material rate basis: package cost ÷ total floor GSF (sum of L1/L2/L3 across
    # all buildings). Material lines are then priced per building × floor level
    # using each level's GSF — matching the sheet (Lumber L1/L2/L3 per building).
    floor_levels = {(b, l): sf for (b, l), sf in wall_sf.items() if str(l).lower() != "roof"}
    material_gsf = sum(floor_levels.values()) or nails_gsf or total_gsf

    def _material_pkg(scope_keys, cc, scope_label, mat_markup_default):
        pkg = None
        for sk in scope_keys:
            for k, v in scope_sov.items():
                if sk in k: pkg = v; break
            if pkg: break
        if not pkg or material_gsf <= 0: return
        rate = round(pkg["cost"]/material_gsf, 4)
        mk = (pkg["sell"]/pkg["cost"]-1) if pkg["cost"] else mat_markup_default
        for (bld, lvl), sf in sorted(floor_levels.items()):
            title = f"{bld} {scope_label} - {lvl}"
            mkl = _markup_for(cc, title, round(mk,4))
            bt_rows.append({"category":"01 - Framing","cost_code":cc,"title":title,
                            "cost_type":"Material","unit_cost":rate,"qty":round(sf,2),
                            "unit":"SF","builder":round(rate*sf,2),"markup":mkl,
                            "is_gc":False,"is_nails":False,"scope":scope_label})
    _material_pkg(["lumber"], "201 - Lumber - Material", "Lumber Package", 0.40)
    _material_pkg(["hardware"], "205 - Hardware - Material", "Hardware Package", 0.25)
    _material_pkg(["truss"], "207 - Trusses - Material", "Truss Package", 0.15)
    # recompute material rows
    for row in bt_rows:
        if "owner" not in row: _recompute(row)

    if not bt_rows:
        raise HTTPException(422, "Could not generate any BT lines. Check the Estimate/Buyout layout.")

    # ── SOV distribution ──
    # The GC New-Values + Nails New-Value are spread across every NON-GC SOV line.
    # Nails itself is NOT a separate SOV line (it folds into the distribution), so the
    # divisor EXCLUDES the nails line: divisor = (#labor + #material lines, no nails).
    gc_newval = sum(r["newval"] for r in bt_rows if r["is_gc"])
    nails_newval = nails_row["newval"] if nails_row else 0
    sov_lines = [r for r in bt_rows if not r["is_gc"] and not r["is_nails"]]  # labor + material only
    n_sov = len(sov_lines)
    dist = round((gc_newval + nails_newval) / n_sov, 4) if n_sov else 0
    total_mob = round(sum(r["mob"] for r in bt_rows), 2)

    # Build SOV rows
    sov_rows = [{"n":1,"scope":"Mobilization","type":"","value":total_mob}]
    n = 2
    for r in sov_lines:
        sov_rows.append({"n":n,"scope":r["title"],"type":r["cost_type"],
                         "value":round(r["newval"]+dist, 2)})
        n += 1

    # ── Store as project bt_estimate.json (so BT Report populates) ──
    # Rows MUST match the canonical shape produced by upload_bt_estimate so that
    # get_bt_summary works: needs code_cat + markup_pct on every row.
    stored = []
    for r in bt_rows:
        bldg = _extract_building(r["title"])
        cc = r["cost_code"]
        stored.append({"category":r["category"],"cost_code":cc,"title":r["title"],
                       "building":bldg,"level":_extract_level(r["title"]) if bldg else "",
                       "qty":round(r["qty"],4),"unit":r["unit"],"unit_cost":round(r["unit_cost"],4),
                       "cost_type":r["cost_type"],"builder_cost":round(r["builder"],2),
                       "client_price":round(r["owner"],2),"markup_pct":round(r["markup"],4),
                       "code_cat":_code_category(cc),
                       "qty_cell":r.get("qty_cell"),"qty_divide":r.get("qty_divide",1)})
    save_bt_estimate(project, {"rows":stored,"generated":True,"source":file.filename,
                    "labor_contract":labor_contract,"distribution":dist})
    # Save the uploaded source workbook so the download can bundle the original
    # Estimate + Buyout sheets (lets the user verify the linking).
    src_path = PROJECTS / project / "bt_gen_source.xlsx"
    try: src_path.write_bytes(raw)
    except Exception: pass

    # ── Build downloadable Excel with LIVE FORMULAS (shared builder) ──
    proj_short = re.sub(r"[^A-Za-z0-9_-]", "_", project)[:20] or "Project"
    out_path = PROJECTS / project / f"{proj_short}_BT_Estimate_SOV.xlsx"
    _build_bt_sov_workbook(stored, labor_contract, dist, out_path, source_path=src_path)

    # ── VERIFICATION: check each scope's total qty + builder cost vs the buyout ──
    # The buyout's rate_rows hold the authoritative qty + total cost per scope.
    # We sum the generated BT lines by buyout scope and compare. Anything off by
    # more than 0.5% is flagged so the user knows the numbers tie out.
    def _scope_of_buyout(desc):
        d = _norm(desc)
        if "layout" in d: return "Layout"
        if "framing" in d and "floor" not in d and "roof" not in d: return "Wall Framing"
        if "sheathing - exterior" in d: return "Sheathing"
        if "sheathing - corridor" in d: return "Corridor seperation Plywood Sheathing"
        if "fire wall" in d: return "Stair Plywood Sheathing"
        if "floor truss - unit" in d: return "Unit Floor Truss"
        if "floor truss - corridor" in d: return "Corridor Floor Truss"
        if "roof truss" in d: return "Roof Truss"
        if "hardware" in d: return "Hardware"
        if "punchout" in d: return "Punchout"
        if "stairs" in d or "landings" in d: return "Stair"
        if "zip" in d or "house wrap" in d: return "House Wrap"
        if "nails" in d: return "Nails"
        return None

    # Sum generated builder cost per scope
    gen_by_scope = {}
    for r in bt_rows:
        s = r.get("scope")
        if not s: continue
        g = gen_by_scope.setdefault(s, {"qty":0,"builder":0})
        g["qty"] += r["qty"]; g["builder"] += r["builder"]

    verification = []
    for b in rate_rows:
        s = _scope_of_buyout(b["desc"])
        if not s or s not in gen_by_scope: continue
        exp_builder = round(b["unit_cost"] * b["qty"], 2) if b["qty"] else None
        gen = gen_by_scope[s]
        # only check cost where buyout has a quantity (skip GC months etc. handled elsewhere)
        if exp_builder and exp_builder > 0:
            diff = gen["builder"] - exp_builder
            pct = abs(diff) / exp_builder * 100 if exp_builder else 0
            verification.append({
                "scope": s, "buyout_qty": round(b["qty"],2),
                "gen_qty": round(gen["qty"],2),
                "buyout_builder": exp_builder, "gen_builder": round(gen["builder"],2),
                "diff": round(diff,2), "pct_off": round(pct,3),
                "ok": pct <= 0.5,
            })
    n_off = sum(1 for v in verification if not v["ok"])

    # Summary
    cat_totals = {}
    for r in bt_rows:
        cat_totals[r["category"]] = cat_totals.get(r["category"], 0) + r["builder"]
    return {
        "status":"ok",
        "bt_lines":len(bt_rows),
        "sov_lines":len(sov_rows),
        "labor_contract":round(labor_contract,2),
        "nails_markup": round(nails_row["markup"]*100,1) if nails_row else None,
        "nails_in_range": (0.20 <= nails_row["markup"] <= 0.40) if nails_row else None,
        "total_builder_cost":round(sum(r["builder"] for r in bt_rows),2),
        "total_owner_cost":round(sum(r["owner"] for r in bt_rows),2),
        "total_mobilization":total_mob,
        "distribution_per_line":dist,
        "category_totals":{k:round(v,2) for k,v in cat_totals.items()},
        "verification": verification,
        "verification_issues": n_off,
        "download_url":f"/projects/{project}/bt-estimate-download",
        "message":f"Generated {len(bt_rows)} BT lines + {len(sov_rows)} SOV lines"
                  + (f" • Nails markup {nails_row['markup']*100:.1f}%" if nails_row else "")
                  + (f" • ⚠ {n_off} scope cost mismatch(es)" if n_off else " • ✓ all scope costs verified"),
    }


@app.get("/projects/{project}/bt-estimate-download")
def download_generated_bt_estimate(project: str):
    """Download the generated BT Estimate + SOV workbook (with LIVE formulas).
    If the cached .xlsx is missing (e.g. server restarted), rebuild it on-the-fly
    from the stored bt_estimate.json so the download always works."""
    proj_short = re.sub(r"[^A-Za-z0-9_-]", "_", project)[:20] or "Project"
    out_path = PROJECTS / project / f"{proj_short}_BT_Estimate_SOV.xlsx"

    if not out_path.exists():
        # Rebuild from stored DB data
        data = _db.load_bt_estimate(project)
        if not data:
            raise HTTPException(404, "No generated BT Estimate found. Generate one first.")
        if not (isinstance(data, dict) and data.get("rows")):
            raise HTTPException(404, "No generated BT Estimate found. Generate one first.")
        rows = data["rows"]
        labor_contract = float(data.get("labor_contract", 0) or 0)
        dist = float(data.get("distribution", 0) or 0)
        _src = PROJECTS / project / "bt_gen_source.xlsx"
        _build_bt_sov_workbook(rows, labor_contract, dist, out_path,
                               source_path=_src if _src.exists() else None)

    return FileResponse(out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=out_path.name)


def _build_bt_sov_workbook(rows, labor_contract, dist, out_path, source_path=None):
    """Write a BT Estimate + SOV workbook with LIVE Excel formulas from stored rows.
    If source_path is given, the original Estimate + Buyout sheets are bundled in AND
    the BT Estimate Quantity cells are LINKED to the Estimate sheet with live formulas
    (e.g. ='SRC-Estimate'!C51), so the user can trace every quantity. Shared by the
    generator and the on-the-fly download rebuild."""
    out_wb = openpyxl.Workbook()
    ws1 = out_wb.active; ws1.title = "BT Estimate"
    hdr_font = Font(bold=True, color="FFFFFF"); hdr_fill = PatternFill("solid", fgColor="2454FF")

    # ── First, bundle the original Estimate + Buyout sheets so we can LINK to them ──
    est_sheet_name = None   # the SRC-Estimate sheet name (for qty links)
    if source_path and Path(source_path).exists():
        try:
            src_wb = openpyxl.load_workbook(source_path, data_only=True)
            for sn in src_wb.sheetnames:
                if _norm(sn) == "estimate" or "buyout" in _norm(sn):
                    src_ws = src_wb[sn]
                    new_name = f"SRC-{sn}"[:31]
                    new_ws = out_wb.create_sheet(new_name)
                    for row in src_ws.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                new_ws.cell(cell.row, cell.column, cell.value)
                    if _norm(sn) == "estimate":
                        est_sheet_name = new_name
        except Exception:
            est_sheet_name = None

    headers = ["Category","Cost Code","Title","Cost Type","Unit Cost","Quantity","Unit",
               "Builder Cost","Markup","Owner Cost","Mobilization","New Value","Source / Link"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(1, c, h); cell.font = hdr_font; cell.fill = hdr_fill
    # find nails row
    nails_excel_row = None
    for i, r in enumerate(rows):
        if "nail" in str(r.get("cost_code","")).lower() or "nail" in str(r.get("title","")).lower():
            nails_excel_row = i + 2
    for i, r in enumerate(rows):
        rr = i + 2
        ws1.cell(rr, 1, r.get("category",""))
        ws1.cell(rr, 2, r.get("cost_code",""))
        ws1.cell(rr, 3, r.get("title",""))
        ws1.cell(rr, 4, r.get("cost_type",""))
        ws1.cell(rr, 5, round(float(r.get("unit_cost",0) or 0), 3))   # 3-decimal unit cost
        # Quantity: LIVE link to the Estimate sheet when we know the source cell.
        qcell = r.get("qty_cell"); qdiv = r.get("qty_divide", 1) or 1
        if est_sheet_name and qcell:
            if qdiv and qdiv != 1:
                ws1.cell(rr, 6, f"='{est_sheet_name}'!{qcell}/{qdiv}")
            else:
                ws1.cell(rr, 6, f"='{est_sheet_name}'!{qcell}")
            src_note = f"Qty ='{est_sheet_name}'!{qcell}" + (f"/{qdiv}" if qdiv!=1 else "")
        else:
            ws1.cell(rr, 6, round(float(r.get("qty",0) or 0), 3))
            src_note = "Qty: package/derived"
        ws1.cell(rr, 7, r.get("unit",""))
        ws1.cell(rr, 8, f"=E{rr}*F{rr}")                       # Builder = Unit*Qty
        if rr == nails_excel_row:
            ws1.cell(rr, 9, f"=(J{rr}/H{rr})-1")               # Markup reverse-calc
            ws1.cell(rr, 10, f"={labor_contract:.3f}-SUM(J2:J{rr-1})")  # Owner = contract - sum above
            src_note = "Nails PLUG = contract − SUM(owner above)"
        else:
            ws1.cell(rr, 9, round(float(r.get("markup_pct", r.get("markup", 0)) or 0), 4))
            ws1.cell(rr, 10, f"=H{rr}*(1+I{rr})")              # Owner = Builder*(1+Markup)
        ws1.cell(rr, 11, f"=J{rr}*3%")                         # Mobilization = Owner*3%
        ws1.cell(rr, 12, f"=J{rr}-K{rr}")                      # New Value = Owner-Mob
        cc = str(r.get("cost_code",""))
        ws1.cell(rr, 13, f"{src_note} • Rate from Buyout [{cc.split(' - ')[0]}]")
    last_bt = len(rows) + 1
    for col_cells in ws1.columns:
        w = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws1.column_dimensions[col_cells[0].column_letter].width = min(w+2, 55)  # type: ignore[union-attr]

    # SOV sheet with live formulas (3-decimal distribution for exact contract match)
    ws2 = out_wb.create_sheet("SOV")
    ws2.cell(1,1,"Owners Invoices").font = Font(bold=True, size=13)
    for c,h in enumerate(["#","Scope","Type","Value"],1):
        cell = ws2.cell(2,c,h); cell.font = hdr_font; cell.fill = hdr_fill
    ws2.cell(1,6,"Dist/line:"); ws2.cell(1,7,round(float(dist or 0), 3))   # G1 holds distribution (3 dp)
    ws2.cell(3,1,1); ws2.cell(3,2,"Mobilization")
    ws2.cell(3,4,f"=SUM('BT Estimate'!K2:K{last_bt})")        # Mobilization = sum of all K
    rr = 4; idx = 2
    for r in rows:
        # GC lines are not separate SOV lines; nails folds into the distribution
        if str(r.get("code_cat","")) == "gc": continue
        if "nail" in str(r.get("cost_code","")).lower() or "nail" in str(r.get("title","")).lower(): continue
        ws2.cell(rr,1,idx); ws2.cell(rr,2,r.get("title","")); ws2.cell(rr,3,r.get("cost_type",""))
        ws2.cell(rr,4,f"=VLOOKUP(B{rr},'BT Estimate'!C:L,10,FALSE)+$G$1")  # NewVal + dist
        rr += 1; idx += 1
    tot_r = rr + 1
    ws2.cell(tot_r,3,"Total").font = Font(bold=True)
    ws2.cell(tot_r,4,f"=SUM(D3:D{rr-1})").font = Font(bold=True)
    for col_cells in ws2.columns:
        w = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws2.column_dimensions[col_cells[0].column_letter].width = min(w+2, 45)  # type: ignore[union-attr]

    # Move BT Estimate + SOV to the FRONT (before the SRC sheets)
    out_wb.move_sheet("BT Estimate", -(len(out_wb.sheetnames)-1))
    try:
        out_wb.move_sheet("SOV", -(out_wb.sheetnames.index("SOV")-1))
    except Exception:
        pass

    out_wb.save(out_path)

# ── Upload BT Estimate ────────────────────────────────────────────────────
@app.post("/projects/{project}/upload-bt-estimate")
async def upload_bt_estimate(project: str, file: UploadFile = File(...)):
    """Parse BT Estimate XLS/XLSX and store rows."""
    import io
    _db.create_project(project)
    raw = await file.read()
    ext = (file.filename or "").lower().rsplit(".", 1)[-1]
    try:
        import pandas as pd
        if ext == "xls":
            df = pd.read_excel(io.BytesIO(raw), engine="xlrd", header=0)
        else:
            df = pd.read_excel(io.BytesIO(raw), engine="openpyxl", header=0)
        df.columns = [str(c).strip() for c in df.columns]
    except Exception as e:
        raise HTTPException(400, f"Cannot read file: {e}")

    # Ensure required cols
    required = ["Category","Cost Code","Title","Builder Cost"]
    for rc in required:
        if rc not in df.columns:
            raise HTTPException(422, f"Missing column '{rc}'. Expected BT Estimate export.")

    rows = []
    for _, r in df.iterrows():
        cat  = str(r.get("Category","") or "").strip()
        cc   = _clean_cost_code(str(r.get("Cost Code","") or ""))
        if not cc: continue
        title = str(r.get("Title","") or "").strip()
        bc    = float(str(r.get("Builder Cost",0) or 0).replace(",","") or 0)
        cp    = float(str(r.get("Client Price",0) or 0).replace(",","") or 0)
        qty   = float(str(r.get("Quantity",0) or 0) or 0)
        unit  = str(r.get("Unit","") or "").strip()
        uc    = float(str(r.get("Unit Cost",0) or 0) or 0)
        ct    = str(r.get("Cost Type","") or "").strip()
        markup = float(str(r.get("Markup",0) or 0) or 0)
        bldg = _extract_building(title)
        rows.append({
            "category":    cat,
            "cost_code":   cc,
            "title":       title,
            "building":    bldg,
            "level":       _extract_level(title) if bldg else "",
            "qty":         round(qty, 4),
            "unit":        unit,
            "unit_cost":   round(uc, 4),
            "cost_type":   ct,
            "builder_cost": round(bc, 2),
            "client_price": round(cp, 2),
            "markup_pct":  round(markup, 4),
            "code_cat":    _code_category(cc),
        })

    save_bt_estimate(project, rows)
    return {"status":"ok","rows":len(rows),
            "message":f"Loaded {len(rows)} estimate rows from {file.filename}"}


@app.delete("/projects/{project}/bt-estimate")
def delete_bt_estimate(project: str):
    save_bt_estimate(project, [])
    return {"status": "ok", "message": "BT Estimate deleted"}


# ── Upload BT POs ─────────────────────────────────────────────────────────
@app.post("/projects/{project}/upload-bt-pos")
async def upload_bt_pos_file(project: str, file: UploadFile = File(...)):
    """Parse BT Purchase Orders XLSX and store rows."""
    import io
    raw = await file.read()
    try:
        import pandas as pd
        # BT PO export has a title row at row 0, headers at row 1
        df = pd.read_excel(io.BytesIO(raw), engine="openpyxl", header=1)
        df.columns = [str(c).strip() for c in df.columns]
    except Exception as e:
        raise HTTPException(400, f"Cannot read file: {e}")

    rows = []
    for _, r in df.iterrows():
        po_no = str(r.get("PO #","") or "").strip()
        if not po_no or po_no.lower() == "nan": continue
        cc    = _clean_cost_code(str(r.get("Cost Code","") or "").replace("\r",""))
        cost  = float(str(r.get("Cost",0) or 0) or 0)
        paid  = float(str(r.get("Amount Paid",0) or 0) or 0)
        rem   = float(str(r.get("Amount Remaining",0) or 0) or 0)
        pct_p = float(str(r.get("% Paid",0) or 0) or 0)
        pct_b = float(str(r.get("% Total Billed",0) or 0) or 0)
        vc_raw = r.get("Variance Code","")
        # NaN from pandas becomes float, handle it properly
        var_code = "" if (vc_raw is None or (hasattr(vc_raw,'__float__') and str(vc_raw)=='nan')) else str(vc_raw).strip()
        # VPO = PO# has "-VPO" suffix OR Variance Code column is filled
        is_vpo = ("-VPO" in po_no.upper()) or (bool(var_code) and var_code.lower() not in ('nan','none'))
        rows.append({
            "po_no":          po_no,
            "title":          str(r.get("Title","") or "").strip(),
            "sub_name":       str(r.get("Performed By","") or "").strip(),
            "work_status":    str(r.get("Work Status","") or "").strip(),
            "po_status":      str(r.get("PO Status","") or "").strip(),
            "cost_code":      cc,
            "cost":           round(cost, 2),
            "pct_billed":     round(pct_b * 100, 2),
            "pct_remaining":  round(float(str(r.get("% Remaining",0) or 0) or 0) * 100, 2),
            "pct_paid":       round(pct_p * 100, 2),
            "amount_paid":    round(paid, 2),
            "amount_remaining": round(rem, 2),
            "created_date":   str(r.get("Created Date","") or "").strip(),
            "from_estimate":  str(r.get("From Estimate","") or "").strip(),
            "is_vpo":         is_vpo,
            "variance_code":  var_code,
            "code_cat":       _code_category(cc),
        })

    save_bt_pos(project, rows)

    # NEW: auto-sync schedule progress from labor POs (90% rule)
    sync_result = None
    try:
        sync_result = _sync_schedule_from_labor_pos(project)
    except Exception as _e:
        sync_result = {"synced": 0, "error": str(_e)}

    msg = f"Loaded {len(rows)} PO rows ({sum(1 for r in rows if r['is_vpo'])} VPOs)"
    if sync_result and sync_result.get("synced"):
        msg += f" • Auto-synced {sync_result['synced']} schedule activities"
    return {"status":"ok","rows":len(rows), "message":msg, "sync": sync_result}


@app.delete("/projects/{project}/bt-pos")
def delete_bt_pos(project: str):
    save_bt_pos(project, [])
    return {"status": "ok", "message": "BT POs deleted"}


@app.put("/projects/{project}/bt-pos/{po_no}")
def edit_bt_po(project: str, po_no: str, body: dict):
    """Manually edit a PO row (paid amount, billed %, notes)."""
    rows = load_bt_pos(project)
    for r in rows:
        if r["po_no"] == po_no:
            for k in ["cost","amount_paid","amount_remaining","pct_paid",
                      "pct_billed","work_status","po_status","sub_name","notes"]:
                if k in body: r[k] = body[k]
            break
    else:
        raise HTTPException(404, f"PO {po_no} not found")
    save_bt_pos(project, rows)
    return {"status":"ok","po_no":po_no}


# ── BT Summary (4 sheets) ─────────────────────────────────────────────────

@app.put("/projects/{project}/bt-estimate-row")
def edit_bt_estimate_row(project: str, body: dict):
    """Manually edit a BT estimate row by cost_code + building + level."""
    rows = load_bt_estimate(project)
    cc   = body.get("cost_code","")
    bldg = body.get("building","")
    lvl  = body.get("level","")
    matched = False
    for r in rows:
        if r["cost_code"] == cc and r.get("building","") == bldg and r.get("level","") == lvl:
            for k in ["builder_cost","client_price","qty","unit_cost","title"]:
                if k in body: r[k] = body[k]
            matched = True
            break
    if not matched:
        raise HTTPException(404, f"Estimate row not found: {cc}/{bldg}/{lvl}")
    save_bt_estimate(project, rows)
    return {"status": "ok", "message": "Estimate row updated"}


@app.get("/projects/{project}/bt-summary")
def get_bt_summary(project: str):
    """Compute all 4 BT sheets from stored estimate + PO data + inventory."""
    est_rows = load_bt_estimate(project)
    po_rows  = load_bt_pos(project)
    items    = load_items(project)
    meta     = load_meta(project)

    # ── Build PO lookup by cost_code ───────────────────────────────────────
    from collections import defaultdict
    po_by_code = defaultdict(lambda: {"cost":0,"paid":0,"remaining":0,"vpo_cost":0,
                                       "vpo_paid":0,"vpo_remaining":0,"pos":[],"vpos":[]})
    for r in po_rows:
        cc = _clean_cost_code(r.get("cost_code",""))  # ensure clean cost code
        r["cost_code"] = cc                           # normalise in-place
        if not cc: continue
        if r["is_vpo"]:
            po_by_code[cc]["vpo_cost"]      += r["cost"]
            po_by_code[cc]["vpo_paid"]      += r["amount_paid"]
            po_by_code[cc]["vpo_remaining"] += r["amount_remaining"]
            po_by_code[cc]["vpos"].append(r)
        else:
            po_by_code[cc]["cost"]      += r["cost"]
            po_by_code[cc]["paid"]      += r["amount_paid"]
            po_by_code[cc]["remaining"] += r["amount_remaining"]
            po_by_code[cc]["pos"].append(r)

    # ── Sheet 1: Labor Summary ─────────────────────────────────────────────
    labor_codes = defaultdict(lambda: {"code":"","estimated":0,"po_total":0,"vpo_total":0,
                                        "billed":0,"paid":0,"remaining":0,"pos":[],"vpos":[]})
    for r in est_rows:
        if r.get("code_cat") != "labor": continue
        cc = r["cost_code"]
        labor_codes[cc]["code"] = cc
        labor_codes[cc]["estimated"] += r["builder_cost"]
        labor_codes[cc]["client_price"] = labor_codes[cc].get("client_price", 0) + r.get("client_price", 0)

    # Merge PO data
    for cc, ld in labor_codes.items():
        pd_ = po_by_code.get(cc, {})
        ld["po_total"]   = round(pd_.get("cost",0) + pd_.get("vpo_cost",0), 2)
        ld["vpo_total"]  = round(pd_.get("vpo_cost",0), 2)
        ld["billed"]     = round(pd_.get("cost",0) * 0 + pd_.get("paid",0) + pd_.get("vpo_paid",0), 2)
        # billed = sum of (cost × pct_billed) across POs
        billed_sum = sum(r["cost"] * r["pct_billed"] / 100 for r in pd_.get("pos",[]))  # type: ignore[arg-type]
        billed_sum += sum(r["cost"] * r["pct_billed"] / 100 for r in pd_.get("vpos",[]))  # type: ignore[arg-type]
        ld["billed"]     = round(billed_sum, 2)
        ld["paid"]       = round(pd_.get("paid",0) + pd_.get("vpo_paid",0), 2)
        ld["remaining"]  = round(pd_.get("remaining",0) + pd_.get("vpo_remaining",0), 2)
        # FIX 4: report PO and VPO separately, plus combined total, plus variance %
        ld["po_only"]     = round(pd_.get("cost",0), 2)                   # PO only (no VPO)
        ld["vpo_only"]    = round(pd_.get("vpo_cost",0), 2)               # VPO only
        ld["po_plus_vpo"] = round(pd_.get("cost",0)+pd_.get("vpo_cost",0), 2)  # PO + VPO combined
        ld["variance"]   = round(ld["po_total"] - ld["estimated"], 2)
        ld["variance_pct"] = round(ld["variance"] / ld["estimated"] * 100, 1) if ld["estimated"] else 0
        ld["pct_billed"] = round(ld["billed"] / ld["po_total"] * 100, 1) if ld["po_total"] else 0
        ld["pos"]  = pd_.get("pos",[])
        ld["vpos"] = pd_.get("vpos",[])

    # Include PO-only codes (not in estimate)
    for cc, pd_ in po_by_code.items():
        if _code_category(cc) == "labor" and cc not in labor_codes:
            total = pd_["cost"] + pd_.get("vpo_cost",0)
            billed = sum(r["cost"]*r["pct_billed"]/100 for r in pd_["pos"]+pd_["vpos"])  # type: ignore[arg-type]
            labor_codes[cc] = {
                "code": cc, "estimated": 0,
                "po_total": round(total, 2), "vpo_total": round(pd_.get("vpo_cost",0), 2),
                "po_only": round(pd_.get("cost",0), 2),
                "vpo_only": round(pd_.get("vpo_cost",0), 2),
                "po_plus_vpo": round(total, 2),
                "billed": round(billed, 2), "paid": round(pd_["paid"]+pd_.get("vpo_paid",0), 2),
                "remaining": round(pd_["remaining"]+pd_.get("vpo_remaining",0), 2),
                "variance": round(total, 2), "variance_pct": 0,
                "pct_billed": round(billed/total*100,1) if total else 0,
                "pos": pd_["pos"], "vpos": pd_["vpos"],
            }

    labor_summary = sorted(labor_codes.values(), key=lambda x: x["code"])

    # ── Sheet 2: General Conditions ────────────────────────────────────────
    gc_codes = defaultdict(lambda: {"code":"","title":"","estimated":0,"po_total":0,
                                     "vpo_total":0,"billed":0,"paid":0,"remaining":0,"variance":0})
    for r in est_rows:
        if r.get("code_cat") != "gc": continue
        cc = r["cost_code"]
        gc_codes[cc]["code"]      = cc
        gc_codes[cc]["title"]     = cc.split(" - ",1)[-1] if " - " in cc else cc
        gc_codes[cc]["estimated"] += r["builder_cost"]
        gc_codes[cc]["client_price"] = gc_codes[cc].get("client_price", 0) + r.get("client_price", 0)

    for cc, gd in gc_codes.items():
        pd_ = po_by_code.get(cc, {})
        total   = pd_.get("cost",0) + pd_.get("vpo_cost",0)
        billed  = sum(r["cost"]*r["pct_billed"]/100 for r in pd_.get("pos",[])+pd_.get("vpos",[]))  # type: ignore[arg-type]
        gd["po_total"]   = round(total, 2)
        gd["vpo_total"]  = round(pd_.get("vpo_cost",0), 2)
        gd["po_only"]    = round(pd_.get("cost",0), 2)
        gd["vpo_only"]   = round(pd_.get("vpo_cost",0), 2)
        gd["po_plus_vpo"]= round(pd_.get("cost",0)+pd_.get("vpo_cost",0), 2)
        gd["billed"]    = round(billed, 2)
        gd["paid"]      = round(pd_.get("paid",0)+pd_.get("vpo_paid",0), 2)
        gd["remaining"] = round(pd_.get("remaining",0)+pd_.get("vpo_remaining",0), 2)
        gd["variance"]  = round(total - gd["estimated"], 2)
        gd["variance_pct"] = round(gd["variance"] / gd["estimated"] * 100, 1) if gd["estimated"] else 0
        gd["pct_billed"] = round(billed/total*100,1) if total else 0
        # FIX: attach PO list so frontend "Details" popup can render them (was empty)
        gd["pos"]   = pd_.get("pos",  [])
        gd["vpos"]  = pd_.get("vpos", [])

    gc_summary = sorted(gc_codes.values(), key=lambda x: x["code"])

    # ── Sheet 3: Schedule (Building × Level × Scope) ──────────────────────
    sched = defaultdict(lambda: {"building":"","level":"","cost_code":"","scope":"",
                                  "estimated":0,"po_cost":0,"pct_billed":0,"paid":0,"remaining":0})
    for r in est_rows:
        if r.get("code_cat") != "labor" or not r.get("building"): continue
        key = f"{r['building']}|{r['level']}|{r['cost_code']}"
        s   = sched[key]
        s["building"]  = r["building"]
        s["level"]     = r["level"]
        s["cost_code"] = r["cost_code"]
        s["scope"]     = r["cost_code"].split(" - ",1)[-1] if " - " in r["cost_code"] else r["cost_code"]
        s["estimated"] += r["builder_cost"]

    # Attach PO billing % to schedule (at cost_code level, spread evenly across buildings)
    code_pct = {}
    for cc, ld in labor_codes.items():
        code_pct[cc] = ld.get("pct_billed", 0)

    for key, s in sched.items():
        cc = s["cost_code"]
        pd_ = po_by_code.get(cc, {})
        # Find matching POs by title keywords
        bldg = s["building"].lower()
        level = s["level"]
        matching_pos = [p for p in pd_.get("pos",[]) + pd_.get("vpos",[])  # type: ignore[arg-type]
                        if bldg[:8].lower() in p["title"].lower()
                        and (level in p["title"] or level == "General")]
        if matching_pos:
            s["po_cost"]    = round(sum(p["cost"] for p in matching_pos), 2)
            billed_sum      = sum(p["cost"]*p["pct_billed"]/100 for p in matching_pos)
            s["pct_billed"] = round(billed_sum/s["po_cost"]*100,1) if s["po_cost"] else 0
            s["paid"]       = round(sum(p["amount_paid"] for p in matching_pos), 2)
            s["remaining"]  = round(sum(p["amount_remaining"] for p in matching_pos), 2)
        else:
            s["pct_billed"] = code_pct.get(cc, 0)
            s["po_cost"]    = 0

    schedule_rows = sorted(sched.values(),
                            key=lambda x: (x["building"], x["level"], x["cost_code"]))

    # ── Sheet 4: Material Building-Wise ───────────────────────────────────
    # FIX: BT estimate has multiple rows per material cost code (one per building × level).
    # Previously this dict-comprehension overwrote on duplicate keys → last row won.
    # Now we SUM all rows so the total is accurate (e.g. 22 rows summing to $579,452.73).
    mat_est = defaultdict(float)
    mat_client = defaultdict(float)  # also track client price for profit calc
    for r in est_rows:
        if r.get("code_cat") == "material":
            mat_est[r["cost_code"]]    += r.get("builder_cost", 0)
            mat_client[r["cost_code"]] += r.get("client_price", 0)

    po_tax  = n(meta.get("po_tax_rate", meta.get("tax_rate", TAX_RATE)))
    del_tax = n(meta.get("delivery_tax_rate", meta.get("tax_rate", TAX_RATE)))

    inv_total = inv_del = inv_rem = 0.0
    for item in items:
        tv=n(item.get("t_num",0)); wv=n(item.get("w_num",0)); lv=n(item.get("length_num",0))
        uc=n(item.get("unit_cost",0)); typ=item.get("type",""); cf=item.get("cost_formula")
        q  = n(item.get("po_qty",0)) + n(item.get("co_qty",0))
        dt = sum(item.get("deliveries",{}).values())
        rm = max(0, q - dt)
        inv_total += round(compute_cost(typ,q, tv,wv,lv,uc,cf)*(1+po_tax), 2)
        inv_del   += round(compute_cost(typ,dt,tv,wv,lv,uc,cf)*(1+del_tax),2)
        inv_rem   += round(compute_cost(typ,rm,tv,wv,lv,uc,cf)*(1+po_tax), 2)

    # Material POs from BT
    mat_pos_by_code = {cc: po_by_code[cc] for cc in po_by_code
                       if _code_category(cc) == "material"}

    material_summary = []
    for cc in sorted(set(list(mat_est.keys()) + list(mat_pos_by_code.keys()))):
        pd_ = mat_pos_by_code.get(cc, {})
        est = mat_est.get(cc, 0)
        cp  = mat_client.get(cc, 0)
        po_total = pd_.get("cost",0) + pd_.get("vpo_cost",0)
        billed   = sum(r["cost"]*r["pct_billed"]/100 for r in pd_.get("pos",[])+pd_.get("vpos",[]))  # type: ignore[arg-type]
        variance = po_total - est
        material_summary.append({
            "cost_code": cc,
            "title":     cc.split(" - ",1)[-1] if " - " in cc else cc,
            "estimated": round(est, 2),
            "client_price": round(cp, 2),
            "po_total":  round(po_total, 2),
            "vpo_total": round(pd_.get("vpo_cost",0), 2),
            "po_only":     round(pd_.get("cost",0), 2),
            "vpo_only":    round(pd_.get("vpo_cost",0), 2),
            "po_plus_vpo": round(pd_.get("cost",0)+pd_.get("vpo_cost",0), 2),
            "billed":    round(billed, 2),
            "paid":      round(pd_.get("paid",0)+pd_.get("vpo_paid",0), 2),
            "remaining": round(pd_.get("remaining",0)+pd_.get("vpo_remaining",0), 2),
            "variance":  round(variance, 2),
            "variance_pct": round(variance / est * 100, 1) if est else 0,
            "pct_billed": round(billed/po_total*100,1) if po_total else 0,
            # FIX: attach PO list so frontend "Details" popup can render them (was empty)
            "pos":   pd_.get("pos",  []),
            "vpos":  pd_.get("vpos", []),
        })

    # Totals
    def _tot(rows, key):
        return round(sum(r.get(key,0) for r in rows), 2)

    # ── Building-wise breakdown ───────────────────────────────────────────
    # FIX: Buildings are now DYNAMIC — extracted from estimate + PO titles.
    # Works for Willow Way (known names), Cobia Cove (Bldg.1, Bldg.2…), and any other project.
    # GC has no building. Material rows with no building (Nails Framing, etc.) → "Nails"
    NAILS_LABEL = "Nails"
    bldg_wise = {}

    # Collect ALL unique buildings from estimate AND PO rows.
    # FIX: always re-extract from title (don't trust stored .building field)
    # so projects uploaded BEFORE the normalization fix get cleaned up on display.
    dynamic_buildings = set()
    for r in est_rows:
        b = _extract_building(r.get("title", ""))
        if b: dynamic_buildings.add(b)
    for r in po_rows:
        b = _extract_building(r.get("title", ""))
        if b: dynamic_buildings.add(b)

    # Sort for stable display order (known buildings first, then alphabetic)
    known_set = {b.lower() for b in _KNOWN_BUILDINGS}
    sorted_bldgs = (
        [b for b in _KNOWN_BUILDINGS if b in dynamic_buildings] +
        sorted([b for b in dynamic_buildings if b.lower() not in known_set])
    )

    # Init all discovered buildings
    for b in sorted_bldgs:
        bldg_wise[b] = {"building": b, "labor_est": 0, "gc_est": 0,
                         "mat_est": 0, "total_est": 0,
                         "po_cost": 0, "billed": 0, "pct_billed": 0,
                         "scopes": [], "cost_codes": set()}
    bldg_wise["General Conditions"] = {"building": "General Conditions",
        "labor_est": 0, "gc_est": 0, "mat_est": 0, "total_est": 0,
        "po_cost": 0, "billed": 0, "pct_billed": 0, "scopes": [], "cost_codes": set()}
    bldg_wise[NAILS_LABEL] = {"building": NAILS_LABEL,
        "labor_est": 0, "gc_est": 0, "mat_est": 0, "total_est": 0,
        "po_cost": 0, "billed": 0, "pct_billed": 0, "scopes": [], "cost_codes": set()}

    for r in est_rows:
        cat   = r.get("code_cat", "other")
        bc    = r.get("builder_cost", 0)
        # FIX: always re-extract from title (don't trust stored .building) for old uploads
        bldg  = _extract_building(r.get("title", ""))
        if cat == "gc":
            dest = "General Conditions"
        elif bldg and bldg in bldg_wise:
            dest = bldg
        else:
            dest = NAILS_LABEL  # material without building (Nails Framing)
        d = bldg_wise[dest]
        d["total_est"] += bc
        if cat == "labor":    d["labor_est"] += bc
        elif cat == "gc":     d["gc_est"]    += bc
        elif cat == "material": d["mat_est"] += bc
        d["scopes"].append({"code": r.get("cost_code",""), "title": r.get("title",""),
                             "est": bc, "level": r.get("level","")})
        d["cost_codes"].add(r.get("cost_code",""))

    # Add PO cost/paid to buildings by matching PO title against discovered buildings
    for r in po_rows:
        po_title = r.get("title", "")
        po_cost  = r.get("cost", 0)
        po_paid  = r.get("amount_paid", 0)
        matched = False
        for bldg in sorted_bldgs:
            if po_title.lower().startswith(bldg.lower()):
                bldg_wise[bldg]["po_cost"] += po_cost
                bldg_wise[bldg]["billed"]  += po_paid
                matched = True
                break
        if not matched:
            # No building match — GC or Nails
            cc_cat = _code_category(r.get("cost_code",""))
            dest = "General Conditions" if cc_cat == "gc" else NAILS_LABEL
            bldg_wise[dest]["po_cost"] += po_cost
            bldg_wise[dest]["billed"]  += po_paid

    for bldg, bd in bldg_wise.items():
        bd["pct_billed"] = round(bd["billed"]/bd["po_cost"]*100,1) if bd["po_cost"] else 0
        bd["total_est"]  = round(bd["total_est"], 2)
        bd["labor_est"]  = round(bd["labor_est"], 2)
        bd["gc_est"]     = round(bd["gc_est"], 2)
        bd["mat_est"]    = round(bd["mat_est"], 2)
        bd["po_cost"]    = round(bd["po_cost"], 2)
        bd["billed"]     = round(bd["billed"], 2)
        bd["cost_codes"] = list(bd.get("cost_codes", set()))  # convert set → list

    # Only include buildings with data
    building_wise = sorted(
        [bd for bd in bldg_wise.values() if bd["total_est"] > 0 or bd["po_cost"] > 0],
        key=lambda x: x["building"])

    # ── Vendor-wise from POs ───────────────────────────────────────────────
    vendor_wise = {}
    for r in po_rows:
        sub = r.get("sub_name","") or "Unknown"
        if sub not in vendor_wise:
            vendor_wise[sub] = {"vendor": sub, "po_cost": 0, "paid": 0,
                                 "remaining": 0, "vpo_cost": 0, "pos": []}
        if r["is_vpo"]:
            vendor_wise[sub]["vpo_cost"] += r["cost"]
        else:
            vendor_wise[sub]["po_cost"] += r["cost"]
        vendor_wise[sub]["paid"]      += r["amount_paid"]
        vendor_wise[sub]["remaining"] += r["amount_remaining"]
        vendor_wise[sub]["pos"].append({"po_no": r["po_no"], "cost": r["cost"],
                                         "is_vpo": r["is_vpo"], "pct": r["pct_billed"]})
    vendor_list = sorted(
        [{"vendor": k, **{x: round(v[x],2) for x in ["po_cost","paid","remaining","vpo_cost"]},
          "pos": v["pos"]} for k, v in vendor_wise.items()],
        key=lambda x: -x["po_cost"]
    )

    # ── Grand summary for header cards ────────────────────────────────────
    all_est        = sum(r.get("builder_cost",0) for r in est_rows)
    all_client_est = sum(r.get("client_price",0) for r in est_rows)  # NEW: client total from BT
    all_po  = sum(r["cost"] for r in po_rows if not r["is_vpo"])
    all_vpo = sum(r["cost"] for r in po_rows if r["is_vpo"])
    all_billed = sum(r["cost"]*r["pct_billed"]/100 for r in po_rows)
    all_paid   = sum(r["amount_paid"] for r in po_rows)

    lab_est = _tot(labor_summary, "estimated"); lab_po = _tot(labor_summary, "po_total")
    gc_est  = _tot(gc_summary, "estimated");    gc_po  = _tot(gc_summary, "po_total")
    mat_est_total  = _tot(material_summary, "estimated")
    mat_po_total   = _tot(material_summary, "po_total")
    mat_client_tot = _tot(material_summary, "client_price")
    lab_vpo = _tot(labor_summary, "vpo_total")
    gc_vpo  = _tot(gc_summary,    "vpo_total")
    mat_vpo = _tot(material_summary, "vpo_total")

    # Detect whether the stored estimate was machine-generated (has the download workbook)
    _est_raw = _db.load_bt_estimate(project)
    _is_generated = False
    if _est_raw:
        try:
            _d = _est_raw if isinstance(_est_raw, dict) else {}
            _is_generated = bool(_d.get("generated"))
        except Exception:
            _is_generated = False

    return {
        "project": project,
        "has_estimate": bool(est_rows),
        "has_generated": _is_generated,
        "has_pos":      bool(po_rows),
        "estimate_rows": len(est_rows),
        "po_rows":       len(po_rows),
        "vpo_count":     sum(1 for r in po_rows if r["is_vpo"]),
        "labor_summary":  labor_summary,
        "gc_summary":     gc_summary,
        "schedule":       schedule_rows,
        "material_summary": material_summary,
        "building_wise":  building_wise,
        "vendor_wise":    vendor_list,
        "inventory": {
            "total": round(inv_total,2), "delivered": round(inv_del,2),
            "remaining": round(inv_rem,2),
        },
        "totals": {
            "all_estimated":   round(all_est,2),
            "all_client":      round(all_client_est, 2),       # NEW
            "all_profit":      round(all_client_est - all_est, 2),  # NEW
            "all_margin_pct":  round((all_client_est - all_est)/all_client_est*100, 2) if all_client_est else 0,  # NEW
            "all_po":          round(all_po,2),
            "all_vpo":         round(all_vpo,2),
            "all_total":       round(all_po + all_vpo, 2),     # NEW: PO + VPO
            "all_billed":      round(all_billed,2),
            "all_paid":        round(all_paid,2),
            "labor_estimated": lab_est, "labor_po": lab_po,
            "labor_vpo":       lab_vpo,
            "labor_total":     round(lab_po + lab_vpo, 2),     # NEW: PO + VPO
            "labor_client":    _tot(labor_summary, "client_price"),  # NEW: client cost
            "labor_billed":    _tot(labor_summary, "billed"),
            "labor_paid":      _tot(labor_summary, "paid"),
            "gc_estimated":    gc_est, "gc_po": gc_po,
            "gc_vpo":          gc_vpo,
            "gc_total":        round(gc_po + gc_vpo, 2),       # NEW: PO + VPO
            "gc_client":       _tot(gc_summary, "client_price"),    # NEW: client cost
            "gc_billed":       _tot(gc_summary, "billed"),
            "gc_paid":         _tot(gc_summary, "paid"),
            "mat_estimated":   mat_est_total,
            "mat_client":      mat_client_tot,                  # NEW
            "mat_po":          mat_po_total,
            "mat_vpo":         mat_vpo,
            "mat_total":       round(mat_po_total + mat_vpo, 2),  # NEW: PO + VPO
            "mat_billed":      _tot(material_summary, "billed"),
            "mat_paid":        _tot(material_summary, "paid"),
        }
    }


# ══════════════════════════════════════════════════════════════════════════════
# P6-STYLE SCHEDULE MODULE
# Activities: Activity ID, Name, Start, Finish, Duration, Baseline dates, % Complete
# Relationships: FS/SS/FF/SF with lag days
# Baselines: locked snapshots, unlock/re-lock, comparison with current
# Auto-generate from BT Estimate
# ══════════════════════════════════════════════════════════════════════════════
from datetime import datetime, timedelta, date as _date_cls
import json as _json

def load_sched_v2(project: str):
    return _db.load_sched_v2(project) or {"activities": [], "relationships": [], "next_id": 1000}

def save_sched_v2(project: str, data: dict):
    _db.save_sched_v2(project, data)

def load_baselines(project: str):
    return _db.load_baselines(project) or {"baselines": [], "next_bl_id": 1}

def save_baselines(project: str, data: dict):
    _db.save_baselines(project, data)

# ── Per-project Working Calendar ─────────────────────────────────────────────
# Schema: {
#   "work_week": [Mon, Tue, Wed, Thu, Fri, Sat, Sun]  (booleans),
#   "exceptions": {"YYYY-MM-DD": "non_working" | "working", ...}
# }
# Mon = index 0, Sun = index 6.  Default = 6-day week (Mon-Sat working, Sun off).

def load_calendar(project: str):
    return _db.load_calendar(project) or {
        "work_week": [True, True, True, True, True, True, False],
        "exceptions": {}
    }

def save_calendar(project: str, data: dict):
    _db.save_calendar(project, data)

def _is_working_day(date_obj, calendar: dict) -> bool:
    """Check if a datetime.date is a working day per the calendar."""
    if calendar is None:
        return date_obj.weekday() != 6  # Sun off
    iso = date_obj.strftime("%Y-%m-%d") if hasattr(date_obj, 'strftime') else str(date_obj)
    ex = (calendar.get("exceptions") or {}).get(iso)
    if ex == "non_working": return False
    if ex == "working":     return True
    work_week = calendar.get("work_week") or [True]*5 + [True, False]
    wd = date_obj.weekday()  # Mon=0, Sun=6
    if 0 <= wd < len(work_week):
        return bool(work_week[wd])
    return wd != 6

def _add_workdays_cal(start_str: str, days: int, calendar: dict | None = None) -> str:
    """Calendar-aware version of _add_workdays. Falls back to Mon-Sat if calendar is None."""
    if not start_str: return ""
    d = datetime.strptime(_parse_date(start_str), "%Y-%m-%d")
    added = 0
    while added < days:
        d += timedelta(days=1)
        if _is_working_day(d, calendar): added += 1  # type: ignore[arg-type]
    return d.strftime("%Y-%m-%d")

def _shift_workdays_cal(date_str: str, days: int, calendar: dict | None = None) -> str:
    """Calendar-aware sign-aware shift. Falls back to Mon-Sat if calendar is None."""
    if not date_str: return ""
    d = datetime.strptime(_parse_date(date_str), "%Y-%m-%d")
    if days == 0: return d.strftime("%Y-%m-%d")
    step = 1 if days > 0 else -1
    remaining = abs(days)
    while remaining > 0:
        d += timedelta(days=step)
        if _is_working_day(d, calendar): remaining -= 1  # type: ignore[arg-type]
    return d.strftime("%Y-%m-%d")

def _wd_between(start_str: str, finish_str: str, calendar: dict | None = None) -> int:
    """Inclusive working days between two dates."""
    if not start_str or not finish_str: return 0
    s = datetime.strptime(_parse_date(start_str), "%Y-%m-%d")
    f = datetime.strptime(_parse_date(finish_str), "%Y-%m-%d")
    if f < s: return 0
    wd = 0
    d = s
    while d <= f:
        if _is_working_day(d, calendar): wd += 1  # type: ignore[arg-type]
        d += timedelta(days=1)
    return max(1, wd)


def _parse_date(s):
    """Parse date string to YYYY-MM-DD."""
    if not s: return ""
    s = str(s).strip().split(" ")[0]
    for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"]:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except: pass
    return s

def _add_workdays(start_str: str, days: int) -> str:
    """Add working days (Mon-Sat). Legacy helper — uses default Mon-Sat calendar."""
    if not start_str: return ""
    d = datetime.strptime(_parse_date(start_str), "%Y-%m-%d")
    added = 0
    while added < days:
        d += timedelta(days=1)
        if d.weekday() != 6:  # skip Sunday
            added += 1
    return d.strftime("%Y-%m-%d")

def _shift_workdays(date_str: str, days: int) -> str:
    """Add or subtract working days (Mon-Sat). Sign-aware.
    Used for lag computation and SS/FF/SF predecessor offsets."""
    if not date_str: return ""
    d = datetime.strptime(_parse_date(date_str), "%Y-%m-%d")
    if days == 0: return d.strftime("%Y-%m-%d")
    step = 1 if days > 0 else -1
    remaining = abs(days)
    while remaining > 0:
        d += timedelta(days=step)
        if d.weekday() != 6:
            remaining -= 1
    return d.strftime("%Y-%m-%d")

def _date_max(*dates):
    """Return latest non-empty YYYY-MM-DD; ignore empties."""
    valid = [d for d in dates if d]
    return max(valid) if valid else ""

def _cpm_forward_pass(data: dict, calendar: dict | None = None, pinned_id: str | None = None) -> int:
    """Recompute start/finish for every activity using a forward pass through relationships.
    P6-style rules:
      FS: successor.start  = max(successor.start, pred.finish + 1 work day + lag)
      SS: successor.start  = max(successor.start, pred.start            + lag)
      FF: successor.finish = max(successor.finish, pred.finish          + lag)
      SF: successor.finish = max(successor.finish, pred.start           + lag)
    Then: finish = start + duration - 1 working day (or vice versa).

    Locked activities (never recomputed):
      - status=='Complete' — actuals are frozen
      - id == pinned_id — the activity the user just edited keeps its dates;
        only its successors will cascade.  This is the P6 "user override" behavior:
        when you drag a bar earlier than its predecessor would push it, your edit wins
        for that activity, but downstream successors still re-flow.

    Returns count of activities whose dates actually changed."""
    acts = data.get("activities", [])
    rels = data.get("relationships", [])
    if not acts: return 0

    # Index activities by id
    by_id = {a["id"]: a for a in acts}
    # Index successor relationships (predecessors of each activity)
    preds_of = {}  # succ_id -> [rel, rel, ...]
    for r in rels:
        preds_of.setdefault(r["succ_id"], []).append(r)

    # Build topological order via Kahn's algorithm (so we update preds before succs)
    in_deg = {a["id"]: 0 for a in acts}
    for r in rels:
        if r["pred_id"] in in_deg and r["succ_id"] in in_deg:
            in_deg[r["succ_id"]] += 1
    queue = [aid for aid, d in in_deg.items() if d == 0]
    topo = []
    seen = set()
    # Adjacency list for forward traversal
    succs_of = {}
    for r in rels:
        succs_of.setdefault(r["pred_id"], []).append(r["succ_id"])
    while queue:
        node = queue.pop(0)
        if node in seen: continue
        seen.add(node)
        topo.append(node)
        for s in succs_of.get(node, []):
            in_deg[s] -= 1
            if in_deg[s] == 0:
                queue.append(s)
    # Append any cycle survivors at the end (won't break, just won't be optimal)
    for aid in by_id:
        if aid not in seen: topo.append(aid)

    changes = 0
    for aid in topo:
        a = by_id[aid]
        # Locked: don't touch Complete activities OR the user-edited activity
        if a.get("status") in ("Complete", "Completed"): continue
        if pinned_id and aid == pinned_id: continue  # FIX 3: user edit wins
        prs = preds_of.get(aid, [])
        if not prs: continue  # no predecessors → leave dates alone

        dur = max(1, int(a.get("duration", 1)))
        # P6-style: reschedule based on the LATEST constraint from predecessors.
        # We compute one driving start/finish from preds (not max with existing).
        driving_start = ""
        driving_finish = ""

        for r in prs:
            p = by_id.get(r["pred_id"])
            if not p: continue
            p_start  = p.get("start", "") or ""
            p_finish = p.get("finish", "") or ""
            lag = int(r.get("lag_days", 0) or 0)
            rtype = (r.get("type") or "FS").upper()

            if rtype == "FS":
                # succ start = pred finish + 1 working day + lag
                base = _shift_workdays_cal(p_finish, 1 + lag, calendar)
                driving_start = _date_max(driving_start, base)
            elif rtype == "SS":
                # succ start = pred start + lag
                base = _shift_workdays_cal(p_start, lag, calendar)
                driving_start = _date_max(driving_start, base)
            elif rtype == "FF":
                # succ finish = pred finish + lag
                base = _shift_workdays_cal(p_finish, lag, calendar)
                driving_finish = _date_max(driving_finish, base)
            elif rtype == "SF":
                # succ finish = pred start + lag
                base = _shift_workdays_cal(p_start, lag, calendar)
                driving_finish = _date_max(driving_finish, base)

        # If we have a finish driver but no start driver, finish anchors → start = finish - dur + 1
        # If we have a start driver but no finish driver, start anchors → finish = start + dur - 1
        # If both, use start driver (typical for FS/SS predecessors).
        new_start  = a.get("start","")  or ""
        new_finish = a.get("finish","") or ""
        if driving_start:
            new_start  = driving_start
            new_finish = _add_workdays_cal(new_start, dur - 1, calendar)
        elif driving_finish:
            new_finish = driving_finish
            new_start  = _shift_workdays_cal(new_finish, -(dur - 1), calendar)

        # Apply
        if new_start != a.get("start","") or new_finish != a.get("finish",""):
            a["start"]  = new_start  or a.get("start","")
            a["finish"] = new_finish or a.get("finish","")
            changes += 1
    # Sync planned_start/finish = start/finish after every CPM run
    for a in acts:
        if a.get("start"):  a["planned_start"]  = a["start"]
        if a.get("finish"): a["planned_finish"] = a["finish"]
    return changes

def _duration_from_hours(hrs: float) -> int:
    """Convert P6 hours (8h/day) to work days."""
    return max(1, round(float(hrs or 0) / 8))


# ── Import P6 XER ─────────────────────────────────────────────────────────
@app.get("/projects/{project}/schedule/xer")
def download_schedule_xer(project: str):
    """Export the V2 schedule as a Primavera P6 .xer file (P6 v24.12 format).
    The XER format is tab-separated with %T (table), %F (fields), %R (row) and a
    SINGLE %E at the very end of the file (not per-table — that was the bug).
    Generates the tables P6 actually requires for import:
       CURRTYPE, FINTMPL, OBS, PROJECT, CALENDAR, SCHEDOPTIONS, PROJWBS, TASK, TASKPRED.
    Encoded as Windows-1252 with CRLF line endings.
    Import into P6 via File → Import → Primavera XER."""
    data = load_sched_v2(project)
    acts = data.get("activities", [])
    rels = data.get("relationships", [])
    if not acts:
        raise HTTPException(400, "No schedule data. Generate a schedule first.")

    proj_short = re.sub(r"[^A-Za-z0-9_-]", "_", project)[:20] or "MATINV"
    now = datetime.now()
    now_p6 = now.strftime("%Y-%m-%d %H:%M")
    today_str = now.strftime("%Y-%m-%d")

    def _to_p6_date(d, hh="08:00"):
        """ISO YYYY-MM-DD → P6 format YYYY-MM-DD HH:MM (8am default)"""
        if not d: return ""
        return f"{str(d)[:10]} {hh}"

    def _guid(suffix=""):
        """Generate a P6-style 22-char base64-like GUID."""
        import uuid, base64
        b = uuid.uuid4().bytes
        return base64.b64encode(b).decode().rstrip('=')[:22]

    # Build unique WBS list (preserve insertion order via bldg_seq)
    wbs_list = []
    seen_wbs = set()
    for a in sorted(acts, key=lambda x: (x.get("bldg_seq", 9999), x.get("seq_no", 9999))):
        w = a.get("wbs", "") or "General"
        if w not in seen_wbs:
            seen_wbs.add(w); wbs_list.append(w)

    # Stable numeric IDs for P6
    PROJ_ID      = "1000"
    CAL_ID       = "1001"
    OBS_ID       = "540"            # match BT-exported sample
    PROJ_NODE_ID = "2000"
    # WBS numeric ids start at 2001
    wbs_id_map = {}
    for i, w in enumerate(wbs_list):
        wbs_id_map[w] = str(2001 + i)
    # Activity numeric ids
    task_id_map = {a["id"]: str(100000 + i + 1) for i, a in enumerate(acts)}
    proj_guid = _guid("PROJ")

    # Build calendar_data string from the project's working calendar
    cal = load_calendar(project)
    work_week = cal.get("work_week") or [True]*5 + [True, False]
    day_short_hrs = sum(8 for d in work_week if d)
    # P6 day-of-week: 1=Sun, 2=Mon, ..., 7=Sat
    p6_day_keys = [2, 3, 4, 5, 6, 7, 1]  # iso Mon..Sun → p6 2..7,1
    days_block = ""
    for iso_i, p6_i in enumerate(p6_day_keys):
        is_working = work_week[iso_i]
        if is_working:
            days_block += f"(0||{p6_i}()((0||0(s|08:00|f|12:00)())(0||1(s|13:00|f|17:00)())))"
        else:
            days_block += f"(0||{p6_i}()())"
    # Exception dates (non-working overrides)
    excep_block = ""
    from datetime import date as _date_cls
    for iso_date, kind in (cal.get("exceptions") or {}).items():
        if kind == "non_working":
            # P6 stores exception dates as days since 1899-12-30
            try:
                d = datetime.strptime(iso_date, "%Y-%m-%d").date()
                base = _date_cls(1899, 12, 30)
                serial = (d - base).days
                excep_block += f"(0||0(d|{serial})())"
            except: pass
    calendar_data = f"(0||CalendarData()((0||DaysOfWeek()({days_block}))(0||Exceptions()({excep_block}))))"

    lines = []
    # ── HEADER ──
    lines.append(f"ERMHDR\t24.12\t{today_str}\tProject\tADMIN\tMatInv\tdbxDatabaseNoName\tProject Management\tUSD")

    # ── CURRTYPE (required) ──
    lines.append("%T\tCURRTYPE")
    lines.append("%F\tcurr_id\tdecimal_digit_cnt\tcurr_symbol\tdecimal_symbol\tdigit_group_symbol\tpos_curr_fmt_type\tneg_curr_fmt_type\tcurr_type\tcurr_short_name\tgroup_digit_cnt\tbase_exch_rate")
    lines.append("%R\t1\t2\t$\t.\t,\t#1.1\t(#1.1)\tUS Dollar\tUSD\t3\t1")

    # ── FINTMPL (required) ──
    lines.append("%T\tFINTMPL")
    lines.append("%F\tfintmpl_id\tfintmpl_name\tdefault_flag")
    lines.append("%R\t1\tCalendar\tY")

    # ── OBS (required) ──
    lines.append("%T\tOBS")
    lines.append("%F\tobs_id\tparent_obs_id\tguid\tseq_num\tobs_name\tobs_descr")
    lines.append(f"%R\t{OBS_ID}\t\t\t0\tEnterprise\t")

    # ── PROJECT ──
    proj_start = min((a.get("start","") for a in acts if a.get("start")), default=today_str)
    proj_end   = max((a.get("finish","") for a in acts if a.get("finish")), default=today_str)
    lines.append("%T\tPROJECT")
    # 71 fields matching real P6 24.12 export — order is significant!
    lines.append("%F\tproj_id\tfy_start_month_num\trsrc_self_add_flag\tallow_complete_flag\trsrc_multi_assign_flag\tcheckout_flag\tproject_flag\tstep_complete_flag\tcost_qty_recalc_flag\tbatch_sum_flag\tname_sep_char\tdef_complete_pct_type\tproj_short_name\tacct_id\torig_proj_id\tsource_proj_id\tbase_type_id\tclndr_id\tsum_base_proj_id\ttask_code_base\ttask_code_step\tpriority_num\twbs_max_sum_level\tstrgy_priority_num\tlast_checksum\tcritical_drtn_hr_cnt\tdef_cost_per_qty\tlast_recalc_date\tplan_start_date\tplan_end_date\tscd_end_date\tadd_date\tlast_tasksum_date\tfcst_start_date\tdef_duration_type\ttask_code_prefix\tguid\tdef_qty_type\tadd_by_name\tweb_local_root_path\tproj_url\tdef_rate_type\tadd_act_remain_flag\tact_this_per_link_flag\tdef_task_type\tact_pct_link_flag\tcritical_path_type\ttask_code_prefix_flag\tdef_rollup_dates_flag\tuse_project_baseline_flag\trem_target_link_flag\treset_planned_flag\tallow_neg_act_flag\tsum_assign_level\tlast_fin_dates_id\tfintmpl_id\tlast_baseline_update_date\tcr_external_key\tapply_actuals_date\tlocation_id\tlast_schedule_date\tloaded_scope_level\texport_flag\tnew_fin_dates_id\tbaselines_to_export\tbaseline_names_to_export\tnext_data_date\tclose_period_flag\tsum_refresh_date\ttrsrcsum_loaded\tsumtask_loaded")
    # 71 fields — order matches the %F line exactly
    proj_row = [
        PROJ_ID,                           # 1  proj_id
        "1",                               # 2  fy_start_month_num
        "Y","Y","Y","N","Y","N","N","Y",   # 3-10  flags (rsrc_self_add..batch_sum)
        ".",                               # 11 name_sep_char
        "CP_Drtn",                         # 12 def_complete_pct_type
        proj_short,                        # 13 proj_short_name
        "", "", "", "",                    # 14-17  acct_id, orig_proj_id, source_proj_id, base_type_id
        CAL_ID,                            # 18 clndr_id
        "",                                # 19 sum_base_proj_id
        "1000", "10", "10", "2", "500",    # 20-24 task_code_base, step, priority, wbs_max, strgy
        "",                                # 25 last_checksum
        "0", "0.0000",                     # 26-27 critical_drtn_hr_cnt, def_cost_per_qty
        _to_p6_date(proj_start, "00:00"),  # 28 last_recalc_date
        _to_p6_date(proj_start, "00:00"),  # 29 plan_start_date
        "",                                # 30 plan_end_date
        _to_p6_date(proj_end,   "00:00"),  # 31 scd_end_date
        _to_p6_date(today_str,  "00:00"),  # 32 add_date
        "", "",                            # 33-34 last_tasksum_date, fcst_start_date
        "DT_FixedDUR2",                    # 35 def_duration_type
        "A",                               # 36 task_code_prefix
        proj_guid,                         # 37 guid
        "QT_Hour",                         # 38 def_qty_type
        "ADMIN",                           # 39 add_by_name
        "", "",                            # 40-41 web_local_root_path, proj_url
        "COST_PER_QTY",                    # 42 def_rate_type
        "N", "Y",                          # 43-44 add_act_remain_flag, act_this_per_link_flag
        "TT_Task",                         # 45 def_task_type
        "Y",                               # 46 act_pct_link_flag
        "CT_TotFloat",                     # 47 critical_path_type
        "Y","Y","Y","Y","N","N",           # 48-53 task_code_prefix_flag .. allow_neg_act_flag
        "SL_Taskrsrc",                     # 54 sum_assign_level
        "",                                # 55 last_fin_dates_id
        "1",                               # 56 fintmpl_id
        "", "", "", "",                    # 57-60 last_baseline_update_date, cr_external_key, apply_actuals_date, location_id
        _to_p6_date(today_str, "00:00"),   # 61 last_schedule_date  ← was missing
        "7",                               # 62 loaded_scope_level
        "Y",                               # 63 export_flag
        "",                                # 64 new_fin_dates_id
        "", "",                            # 65-66 baselines_to_export, baseline_names_to_export
        "1899-12-30 00:00",                # 67 next_data_date  (P6 sentinel)
        "",                                # 68 close_period_flag
        _to_p6_date(today_str, "00:00"),   # 69 sum_refresh_date
        "",                                # 70 trsrcsum_loaded
        ""                                 # 71 sumtask_loaded  ← was missing
    ]
    assert len(proj_row) == 71, f"PROJECT row must have 71 fields, got {len(proj_row)}"
    lines.append("%R\t" + "\t".join(proj_row))

    # ── CALENDAR ──
    lines.append("%T\tCALENDAR")
    lines.append("%F\tclndr_id\tdefault_flag\tclndr_name\tproj_id\tbase_clndr_id\tlast_chng_date\tclndr_type\tday_hr_cnt\tweek_hr_cnt\tmonth_hr_cnt\tyear_hr_cnt\trsrc_private\tclndr_data")
    week_hrs = day_short_hrs                  # 5d=40, 6d=48, 7d=56
    month_hrs = int(week_hrs * 52 / 12)
    year_hrs  = week_hrs * 52
    cal_name = f"MatInv {sum(work_week)}-day"
    lines.append(f"%R\t{CAL_ID}\tY\t{cal_name}\t\t\t{_to_p6_date(today_str,'00:00')}\tCA_Base\t8\t{week_hrs}\t{month_hrs}\t{year_hrs}\tN\t{calendar_data}")

    # ── SCHEDOPTIONS (required) ──
    lines.append("%T\tSCHEDOPTIONS")
    lines.append("%F\tschedoptions_id\tproj_id\tsched_outer_depend_type\tsched_open_critical_flag\tsched_lag_early_start_flag\tsched_retained_logic\tsched_setplantoforecast\tsched_float_type\tsched_calendar_on_relationship_lag\tsched_use_expect_end_flag\tsched_progress_override\tlevel_float_thrs_cnt\tlevel_outer_assign_flag\tlevel_outer_assign_priority\tlevel_over_alloc_pct\tlevel_within_float_flag\tlevel_keep_sched_date_flag\tlevel_all_rsrc_flag\tsched_use_project_end_date_for_float\tenable_multiple_longest_path_calc\tlimit_multiple_longest_path_calc\tmax_multiple_longest_path\tuse_total_float_multiple_longest_paths\tkey_activity_for_multiple_longest_paths\tLevelPriorityList")
    lines.append(f"%R\t1\t{PROJ_ID}\tSD_Both\tN\tY\tY\tN\tFT_FF\trcal_Predecessor\tY\tN\t0\tN\t5\t25\tN\tY\tY\tY\tN\tY\t10\tY\t\tpriority_type,ASC_BY_FIELD/ASC")

    # ── PROJWBS ──
    lines.append("%T\tPROJWBS")
    lines.append("%F\twbs_id\tproj_id\tobs_id\tseq_num\test_wt\tproj_node_flag\tsum_data_flag\tstatus_code\twbs_short_name\twbs_name\tphase_id\tparent_wbs_id\tev_user_pct\tev_etc_user_value\torig_cost\tindep_remain_total_cost\tann_dscnt_rate_pct\tdscnt_period_type\tindep_remain_work_qty\tanticip_start_date\tanticip_end_date\tev_compute_type\tev_etc_compute_type\tguid\ttmpl_guid\tplan_open_state")
    lines.append(f"%R\t{PROJ_NODE_ID}\t{PROJ_ID}\t{OBS_ID}\t1\t1\tY\tN\tWS_Open\t{proj_short}\t{project}\t\t\t6\t0.88\t0.0000\t0.0000\t\t\t\t\t\tEC_Cmp_pct\tEE_Rem_hr\t{_guid()}\t\t")
    for i, w in enumerate(wbs_list):
        wid = wbs_id_map[w]
        short = re.sub(r"[^A-Za-z0-9]", "", w)[:8] or str(i+1)
        lines.append(f"%R\t{wid}\t{PROJ_ID}\t{OBS_ID}\t{i+1}\t1\tN\tN\tWS_Open\t{short}\t{w}\t\t{PROJ_NODE_ID}\t6\t0.88\t0.0000\t0.0000\t\t\t\t\t\tEC_Cmp_pct\tEE_Rem_hr\t{_guid()}\t\t")

    # ── TASK ──
    lines.append("%T\tTASK")
    lines.append("%F\ttask_id\tproj_id\twbs_id\tclndr_id\tphys_complete_pct\trev_fdbk_flag\test_wt\tlock_plan_flag\tauto_compute_act_flag\tcomplete_pct_type\ttask_type\tduration_type\tstatus_code\ttask_code\ttask_name\trsrc_id\ttotal_float_hr_cnt\tfree_float_hr_cnt\tremain_drtn_hr_cnt\tact_work_qty\tremain_work_qty\ttarget_work_qty\ttarget_drtn_hr_cnt\ttarget_equip_qty\tact_equip_qty\tremain_equip_qty\tcstr_date\tact_start_date\tact_end_date\tlate_start_date\tlate_end_date\texpect_end_date\tearly_start_date\tearly_end_date\trestart_date\treend_date\ttarget_start_date\ttarget_end_date\trem_late_start_date\trem_late_end_date\tcstr_type\tpriority_type\tsuspend_date\tresume_date\tfloat_path\tfloat_path_order\tguid\ttmpl_guid\tcstr_date2\tcstr_type2\tdriving_path_flag\tact_this_per_work_qty\tact_this_per_equip_qty\texternal_early_start_date\texternal_late_end_date\tcreate_date\tupdate_date\tcreate_user\tupdate_user\tlocation_id\tcrt_path_num")
    for a in acts:
        tid     = task_id_map[a["id"]]
        wid     = wbs_id_map.get(a.get("wbs","General"), PROJ_NODE_ID)
        dur_hrs = int(a.get("duration", 1)) * 8
        pct     = float(a.get("pct_complete", 0))
        st      = a.get("status", "Not Started")
        if st in ("Complete","Completed"):
            status_code = "TK_Complete"
        elif st == "In Progress":
            status_code = "TK_Active"
        else:
            status_code = "TK_NotStart"
        ts       = _to_p6_date(a.get("start",""),  "01:00")
        tf       = _to_p6_date(a.get("finish",""), "00:00")
        actual_s = _to_p6_date(a.get("actual_start",""),  "01:00")
        actual_f = _to_p6_date(a.get("actual_finish",""), "00:00")
        remain_hrs = dur_hrs * (1 - pct/100)
        # 60 fields, matching %F exactly
        row = [
            tid, PROJ_ID, wid, CAL_ID,                # task_id, proj_id, wbs_id, clndr_id
            f"{pct:.0f}", "N", "1", "N", "N",          # phys_complete_pct, rev_fdbk, est_wt, lock_plan, auto_compute
            "CP_Drtn", "TT_Task", "DT_FixedDUR2",     # complete_pct_type, task_type, duration_type
            status_code,
            a.get("p6_code", a["id"]),                # task_code
            a.get("name",""),                          # task_name
            "",                                         # rsrc_id
            "0", "0",                                  # total_float_hr_cnt, free_float_hr_cnt
            f"{remain_hrs:.1f}",                       # remain_drtn_hr_cnt
            "0",                                        # act_work_qty
            f"{remain_hrs:.1f}",                       # remain_work_qty
            f"{dur_hrs:.1f}",                          # target_work_qty
            f"{dur_hrs:.1f}",                          # target_drtn_hr_cnt
            "0", "0", "0",                             # target_equip_qty, act_equip_qty, remain_equip_qty
            "",                                         # cstr_date
            actual_s, actual_f,                        # act_start_date, act_end_date
            ts, tf,                                    # late_start_date, late_end_date
            "",                                         # expect_end_date
            ts, tf,                                    # early_start_date, early_end_date
            ts, tf,                                    # restart_date, reend_date
            ts, tf,                                    # target_start_date, target_end_date
            ts, tf,                                    # rem_late_start_date, rem_late_end_date
            "",                                         # cstr_type
            "PT_Normal",                                # priority_type
            "", "", "", "",                             # suspend_date, resume_date, float_path, float_path_order
            _guid(),                                    # guid
            "",                                         # tmpl_guid
            "", "",                                     # cstr_date2, cstr_type2
            "Y" if status_code != "TK_Complete" else "N",  # driving_path_flag
            "0", "0",                                   # act_this_per_work_qty, act_this_per_equip_qty
            "", "",                                     # extern_early_start_date, extern_late_end_date
            _to_p6_date(today_str, "00:00"),            # create_date
            _to_p6_date(today_str, "00:00"),            # update_date
            "MatInv", "MatInv",                         # create_user, update_user
            "", ""                                      # location_id, crt_path_num
        ]
        # Ensure exactly 60 fields
        while len(row) < 60: row.append("")
        lines.append("%R\t" + "\t".join(row[:60]))

    # ── TASKPRED ──
    lines.append("%T\tTASKPRED")
    lines.append("%F\ttask_pred_id\ttask_id\tpred_task_id\tproj_id\tpred_proj_id\tpred_type\tlag_hr_cnt\tcomments\tfloat_path\taref\tarls")
    for i, r in enumerate(rels):
        pid_num = task_id_map.get(r.get("pred_id",""))
        sid_num = task_id_map.get(r.get("succ_id",""))
        if not pid_num or not sid_num: continue
        pred_type = {"FS":"PR_FS","SS":"PR_SS","FF":"PR_FF","SF":"PR_SF"}.get(r.get("type","FS"),"PR_FS")
        lag_hrs   = int(r.get("lag_days", 0)) * 8
        rel_id    = str(40000 + i + 1)
        # Look up pred/succ dates for aref/arls fields
        pred_act = next((a for a in acts if a["id"] == r.get("pred_id")), None)
        succ_act = next((a for a in acts if a["id"] == r.get("succ_id")), None)
        aref = _to_p6_date(pred_act.get("finish","") if pred_act else "", "00:00")
        arls = _to_p6_date(succ_act.get("start","")  if succ_act else "", "01:00")
        lines.append(f"%R\t{rel_id}\t{sid_num}\t{pid_num}\t{PROJ_ID}\t{PROJ_ID}\t{pred_type}\t{lag_hrs}\t\t\t{aref}\t{arls}")

    # ── SINGLE %E at end of file ──
    lines.append("%E")

    # Encode as CP1252 with CRLF (P6 format)
    xer_content = "\r\n".join(lines) + "\r\n"
    out_dir = PROJECTS / project
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{proj_short}_schedule_{now.strftime('%Y%m%d_%H%M')}.xer"
    out_path.write_bytes(xer_content.encode('cp1252', errors='replace'))
    return FileResponse(out_path, media_type='application/octet-stream',
                        filename=out_path.name)


@app.post("/projects/{project}/upload-xer")
async def upload_xer(project: str, file: UploadFile = File(...)):
    """Import P6 XER file into schedule_v2."""
    raw = await file.read()
    try:
        text = raw.decode('latin-1')
    except:
        text = raw.decode('utf-8', errors='replace')

    # Parse XER
    sections = {}; ct = None; cf = None
    for line in text.split('\r\n'):
        if line.startswith('%T\t'):
            ct = line.split('\t')[1]; sections[ct] = {'fields': [], 'rows': []}
        elif line.startswith('%F\t') and ct:
            cf = line.split('\t')[1:]; sections[ct]['fields'] = cf
        elif line.startswith('%R\t') and ct:
            vals = line.split('\t')[1:]
            sections[ct]['rows'].append(dict(zip(cf or [], vals)))

    tasks_raw = sections.get('TASK', {}).get('rows', [])
    preds_raw = sections.get('TASKPRED', {}).get('rows', [])
    wbs_raw   = sections.get('PROJWBS', {}).get('rows', [])

    # WBS lookup
    wbs_map = {w['wbs_id']: w.get('wbs_name', '') for w in wbs_raw}
    id_map  = {t['task_id']: t for t in tasks_raw}  # task_id → task

    activities = []
    for i, t in enumerate(tasks_raw):
        tid = t.get('task_id', '')
        dur_hrs = float(t.get('target_drtn_hr_cnt') or 8)
        dur_days = _duration_from_hours(dur_hrs)
        ts  = (t.get('target_start_date') or t.get('early_start_date') or '')[:10]
        tf  = (t.get('target_end_date')   or t.get('early_end_date')   or '')[:10]
        act_s = (t.get('act_start_date') or '')[:10]
        act_f = (t.get('act_end_date')   or '')[:10]
        pct   = float(t.get('phys_complete_pct') or 0)
        status = t.get('status_code', 'TK_NotStart')
        if act_f:
            disp_status = 'Complete'
        elif act_s:
            disp_status = 'In Progress'
        else:
            disp_status = 'Not Started'
        activities.append({
            "id":            f"A{str(i+1).zfill(4)}",
            "p6_task_id":    tid,
            "p6_code":       t.get('task_code', ''),
            "name":          t.get('task_name', ''),
            "wbs":           wbs_map.get(t.get('wbs_id', ''), ''),
            "start":         _parse_date(ts),
            "finish":        _parse_date(tf),
            "baseline_start": _parse_date(ts),
            "baseline_finish": _parse_date(tf),
            "actual_start":  _parse_date(act_s),
            "actual_finish": _parse_date(act_f),
            "duration":      dur_days,
            "pct_complete":  round(pct, 1),
            "status":        disp_status,
            "is_milestone":  t.get('task_type') == 'TT_Mile',
            "critical":      t.get('driving_path_flag') == 'Y',
            "notes":         "",
            "resource":      t.get('rsrc_id', ''),
            "source":        "xer",
        })

    # Build id lookup for relationships
    tid_to_aid = {t['p6_task_id']: t['id'] for t in activities}

    relationships = []
    PRED_MAP = {'PR_FS': 'FS', 'PR_SS': 'SS', 'PR_FF': 'FF', 'PR_SF': 'SF'}
    for pr in preds_raw:
        pred_aid = tid_to_aid.get(pr.get('pred_task_id', ''))
        succ_aid = tid_to_aid.get(pr.get('task_id', ''))
        if pred_aid and succ_aid:
            lag_hrs = float(pr.get('lag_hr_cnt') or 0)
            lag_days = round(lag_hrs / 8)
            relationships.append({
                "id":        len(relationships) + 1,
                "pred_id":   pred_aid,
                "succ_id":   succ_aid,
                "type":      PRED_MAP.get(pr.get('pred_type', 'PR_FS'), 'FS'),
                "lag_days":  lag_days,
            })

    data = {"activities": activities, "relationships": relationships,
            "next_id": len(activities) + 1}
    save_sched_v2(project, data)
    return {"status": "ok", "activities": len(activities),
            "relationships": len(relationships),
            "message": f"Imported {len(activities)} activities, {len(relationships)} links"}


# ── Generate Schedule from BT Estimate ───────────────────────────────────
@app.post("/projects/{project}/generate-schedule-from-estimate")
def generate_schedule_from_estimate(project: str, body: dict):
    """Auto-generate activities from BT Estimate labor rows.
    body: { start_date: "YYYY-MM-DD", days_per_scope: { "101 - Layout - Labor": 14 } }
    """
    est_rows = load_bt_estimate(project)
    if not est_rows:
        raise HTTPException(400, "Upload a BT Estimate first.")

    start_date  = body.get("start_date", str(_date_cls.today()))
    days_map    = body.get("days_per_scope", {})  # cost_code → duration days
    notes       = body.get("notes", "")
    labor_rows  = [r for r in est_rows if r.get("code_cat") == "labor"]

    # Group by building → scope → sum builder_cost
    from collections import defaultdict, OrderedDict
    by_bldg_scope = defaultdict(lambda: defaultdict(lambda: {
        "builder_cost": 0, "client_price": 0, "qty": 0, "unit": "", "cost_code": "", "description": ""}))

    for r in labor_rows:
        # FIX: always re-extract from title so old uploads get normalized too.
        bldg  = _extract_building(r.get("title", "")) or "General"
        scope = r.get("cost_code", "")
        s     = by_bldg_scope[bldg][scope]
        s["builder_cost"]  += r.get("builder_cost", 0)
        s["client_price"]  += r.get("client_price", 0)
        s["qty"]           += r.get("qty", 0)
        s["unit"]           = r.get("unit", "")
        s["cost_code"]      = scope
        s["description"]    = r.get("title", "")

    # Assign activity IDs and calculate dates
    # Default duration: if not in days_map, estimate from builder_cost
    activities = []; relationships = []; aid = 1
    current_date = start_date  # rolling date per building
    prev_aid_per_building = {}  # last activity per building for FS link

    # Sort buildings and scopes
    SCOPE_ORDER = ["101","102","103","104","106","107","109","110","113","115"]
    def scope_sort(cc):
        num = cc.split(" - ")[0].strip()
        try: return SCOPE_ORDER.index(num) if num in SCOPE_ORDER else 99
        except: return 99

    first_activity_id = None  # track for building-to-building FS link
    last_bldg_last_act = {}

    # FIX 1: assign building-level sequence (1-indexed) in alphabetical order.
    # User can re-order via /schedule/building/swap-seq later.
    sorted_bldgs = sorted(by_bldg_scope.keys())
    bldg_seq_lookup = {b: idx + 1 for idx, b in enumerate(sorted_bldgs)}

    for bldg in sorted_bldgs:
        scopes = by_bldg_scope[bldg]
        bldg_start = start_date  # each building starts at project start (parallel)
        prev_in_bldg = None
        seq_in_bldg = 0  # per-WBS sequence counter for S.NO

        for scope in sorted(scopes.keys(), key=scope_sort):
            info   = scopes[scope]
            cc     = info["cost_code"]
            dur    = int(days_map.get(cc, 0)) or max(1, round(info["builder_cost"] / 1500))
            sc_start = bldg_start  # start after prev in same building
            sc_fin   = _add_workdays(sc_start, dur)
            seq_in_bldg += 1

            a = {
                "id":              f"A{str(aid).zfill(4)}",
                "p6_task_id":      "",
                "p6_code":         f"A{str(aid).zfill(4)}",
                "name":            f"{bldg} - {cc.split(' - ',1)[-1] if ' - ' in cc else cc}",
                "wbs":             bldg,
                "bldg_seq":        bldg_seq_lookup[bldg],  # FIX 1: building-level order
                "seq_no":          seq_in_bldg,   # FIX 1: P6-style S.NO within building
                "start":           sc_start,
                "finish":          sc_fin,
                "baseline_start":  sc_start,
                "baseline_finish": sc_fin,
                "actual_start":    "",
                "actual_finish":   "",
                "duration":        dur,
                "pct_complete":    0,
                "status":          "Not Started",
                "is_milestone":    False,
                "critical":        False,
                "notes":           "",
                "cost_code":       cc,
                "builder_cost":    round(info["builder_cost"], 2),
                "client_price":    round(info["client_price"], 2),
                "source":          "bt_estimate",
            }
            activities.append(a)

            # FS link within same building
            if prev_in_bldg:
                relationships.append({"id": len(relationships)+1,
                    "pred_id": prev_in_bldg, "succ_id": a["id"],
                    "type": "FS", "lag_days": 0})
            prev_in_bldg = a["id"]
            bldg_start = sc_fin  # next scope starts when this one ends
            aid += 1

        last_bldg_last_act[bldg] = prev_in_bldg

    data = {"activities": activities, "relationships": relationships, "next_id": aid}
    save_sched_v2(project, data)
    # Auto-create baseline
    _create_baseline_internal(project, "BT Estimate Baseline", locked=True)
    return {"status": "ok", "activities": len(activities),
            "relationships": len(relationships),
            "message": f"Generated {len(activities)} activities from BT Estimate. Baseline locked."}


def _create_baseline_internal(project: str, name: str, locked: bool = True):
    """Snapshot current schedule as a baseline."""
    data = load_sched_v2(project)
    bls  = load_baselines(project)
    bl = {
        "id":          bls["next_bl_id"],
        "name":        name,
        "created":     str(_date_cls.today()),
        "locked":      locked,
        "activities":  [
            {"id": a["id"], "name": a["name"],
             "start": a["start"], "finish": a["finish"],
             "duration": a["duration"], "pct_complete": a.get("pct_complete", 0)}
            for a in data.get("activities", [])
        ]
    }
    bls["baselines"].append(bl)
    bls["next_bl_id"] += 1
    save_baselines(project, bls)
    return bl


@app.post("/projects/{project}/schedule/baseline")
def create_baseline(project: str, body: dict):
    name   = body.get("name", f"Baseline {str(_date_cls.today())}")
    locked = body.get("locked", True)
    bl = _create_baseline_internal(project, name, locked)
    return {"status": "ok", "baseline_id": bl["id"], "name": bl["name"]}

@app.get("/projects/{project}/schedule/baselines")
def get_baselines(project: str):
    return load_baselines(project)

@app.put("/projects/{project}/schedule/baseline/{bl_id}/lock")
def toggle_baseline_lock(project: str, bl_id: int, body: dict):
    bls = load_baselines(project)
    baselines_list = bls.get("baselines") if isinstance(bls, dict) else []  # type: ignore[union-attr]
    for bl in (baselines_list or []): # pyright: ignore[reportGeneralTypeIssues]
        if bl["id"] == bl_id:
            bl["locked"] = body.get("locked", not bl.get("locked", True))
            save_baselines(project, bls)
            return {"status": "ok", "locked": bl["locked"]}
    raise HTTPException(404, "Baseline not found")

@app.delete("/projects/{project}/schedule/baseline/{bl_id}")
def delete_baseline(project: str, bl_id: int):
    bls = load_baselines(project)
    baselines_list = bls.get("baselines") if isinstance(bls, dict) else []  # type: ignore[union-attr]
    bls["baselines"] = [b for b in (baselines_list or []) if b["id"] != bl_id] # pyright: ignore[reportGeneralTypeIssues]
    save_baselines(project, bls)
    return {"status": "ok"}


# ══════════════════════════════════════════════════════════════════════════════
# LABOR → SCHEDULE AUTO-SYNC
# Match labor POs (from BT PO file) to schedule activities and update %complete.
# Rule: weighted-by-cost average %billed across matching POs.
#       ≥90% billed = task completed (10% retainage rule).
# ══════════════════════════════════════════════════════════════════════════════
def _sync_schedule_from_labor_pos(project: str):
    """Sync schedule activity progress from BT labor POs.
    Returns: {"synced": int, "completed": int, "in_progress": int, "details": [...]}
    """
    po_rows = load_bt_pos(project)
    if not po_rows:
        return {"synced": 0, "completed": 0, "in_progress": 0,
                "message": "No BT POs loaded"}

    data = load_sched_v2(project)
    if not data.get("activities"):
        return {"synced": 0, "completed": 0, "in_progress": 0,
                "message": "No schedule activities — generate from BT first"}

    # Build label maps for matching
    # Labor POs only — match by (building, cost_code) -> {cost, pct_billed}
    from collections import defaultdict
    po_buckets = defaultdict(list)  # (bldg, cost_code) -> [{cost, pct_billed, po_no, title}]
    for r in po_rows:
        if r.get("code_cat") != "labor": continue
        title = r.get("title", "")
        bldg  = _extract_building(title) or "General"
        cc    = _clean_cost_code(r.get("cost_code", ""))
        if not cc: continue
        po_buckets[(bldg, cc)].append({
            "cost": r.get("cost", 0), "pct_billed": r.get("pct_billed", 0),
            "po_no": r.get("po_no", ""), "title": title, "is_vpo": r.get("is_vpo", False),
        })

    today = str(_date_cls.today())
    details = []
    completed = in_progress = synced = 0

    for a in data.get("activities", []):
        # Activity matching keys
        a_bldg = a.get("wbs", "") or "General"
        a_cc   = a.get("cost_code", "")
        if not a_cc: continue

        # Try exact match first, then fall back to building-agnostic
        bucket = po_buckets.get((a_bldg, a_cc), [])
        if not bucket:
            # Same cost code, any building (last resort fallback)
            for (bldg_k, cc_k), v in po_buckets.items():
                if cc_k == a_cc and a_bldg.lower() in bldg_k.lower():
                    bucket = v
                    break
        if not bucket: continue

        # Weighted average %billed by PO cost
        total_cost = sum(p["cost"] for p in bucket)
        if total_cost <= 0:
            avg_pct = sum(p["pct_billed"] for p in bucket) / len(bucket)
        else:
            avg_pct = sum(p["cost"] * p["pct_billed"] for p in bucket) / total_cost
        avg_pct = round(avg_pct, 2)

        # Apply 90% completion rule (10% retainage)
        new_status = "Complete" if avg_pct >= 90 else ("In Progress" if avg_pct > 0 else "Not Started")
        new_pct    = 100.0 if avg_pct >= 90 else avg_pct

        old_status = a.get("status", "Not Started")
        old_pct    = a.get("pct_complete", 0)

        if (abs(new_pct - old_pct) > 0.01) or (new_status != old_status):
            a["status"]       = new_status
            a["pct_complete"] = new_pct
            # auto fill actual_start/finish based on status
            if new_status == "In Progress" and not a.get("actual_start"):
                a["actual_start"] = a.get("start", today)
            if new_status == "Complete":
                if not a.get("actual_start"):
                    a["actual_start"] = a.get("start", today)
                if not a.get("actual_finish"):
                    a["actual_finish"] = today
            synced += 1
            if new_status == "Complete":    completed += 1
            elif new_status == "In Progress": in_progress += 1

        details.append({
            "activity_id":  a["id"],
            "activity":     a.get("name", ""),
            "cost_code":    a_cc,
            "building":     a_bldg,
            "po_count":     len(bucket),
            "weighted_pct": avg_pct,
            "status":       new_status,
            "matched_pos":  [p["po_no"] for p in bucket],
        })

    save_sched_v2(project, data)
    return {"status": "ok", "synced": synced, "completed": completed,
            "in_progress": in_progress, "details": details,
            "total_activities": len(data.get("activities", []))}


@app.post("/projects/{project}/schedule/sync-from-labor")
def sync_schedule_from_labor(project: str):
    """Manually trigger labor → schedule sync."""
    res = _sync_schedule_from_labor_pos(project)
    return res


# ── Activity CRUD ─────────────────────────────────────────────────────────
@app.get("/projects/{project}/calendar")
def get_calendar(project: str):
    return load_calendar(project)

@app.put("/projects/{project}/calendar")
def update_calendar(project: str, body: dict):
    """Save the working-day calendar. Body:
       {"work_week": [Mon, Tue, Wed, Thu, Fri, Sat, Sun] booleans,
        "exceptions": {"YYYY-MM-DD": "non_working"|"working", ...}}
    After saving, runs CPM forward-pass so all dependent dates re-flow through the new calendar."""
    cal = load_calendar(project)
    if "work_week" in body and isinstance(body["work_week"], list) and len(body["work_week"]) == 7:
        cal["work_week"] = [bool(x) for x in body["work_week"]]
    if "exceptions" in body and isinstance(body["exceptions"], dict):
        # Validate exception values
        clean_ex = {}
        for k, v in body["exceptions"].items():
            if v in ("non_working", "working"):
                clean_ex[str(k)[:10]] = v
        cal["exceptions"] = clean_ex
    save_calendar(project, cal)
    # Re-flow dates through the new calendar
    data = load_sched_v2(project)
    if data.get("activities"):
        cpm_changes = _cpm_forward_pass(data, cal)
        # Also fix activity finishes when work_week made their current span invalid
        for a in data["activities"]:
            if a.get("status") in ("Complete","Completed"): continue
            if a.get("start") and a.get("duration"):
                a["finish"] = _add_workdays_cal(a["start"], int(a["duration"]) - 1, cal)
        save_sched_v2(project, data)
        return {"status": "ok", "calendar": cal, "cpm_changes": cpm_changes}
    return {"status": "ok", "calendar": cal, "cpm_changes": 0}

@app.post("/projects/{project}/schedule/reschedule")
def reschedule_all(project: str):
    """Manually run CPM forward-pass on all activities. Useful when relationships
    or dates seem out-of-sync."""
    data = load_sched_v2(project)
    if not data.get("activities"):
        raise HTTPException(400, "No activities to reschedule")
    cpm_changes = _cpm_forward_pass(data, load_calendar(project))
    save_sched_v2(project, data)
    return {"status": "ok", "cpm_changes": cpm_changes,
            "message": f"Recalculated dates • {cpm_changes} activities moved"}


@app.get("/projects/{project}/schedule/v2")
def get_schedule_v2(project: str):
    data = load_sched_v2(project)
    bls  = load_baselines(project)
    # Attach baseline dates to each activity
    active_bl = next((b for b in reversed(bls["baselines"]) if b.get("locked")), None)
    if active_bl:
        bl_map = {a["id"]: a for a in active_bl.get("activities", [])}
        for a in data.get("activities", []):
            bla = bl_map.get(a["id"], {})
            a["baseline_start"]  = bla.get("start", a.get("baseline_start", ""))
            a["baseline_finish"] = bla.get("finish", a.get("baseline_finish", ""))

    # FIX 1: Backfill seq_no (activity within building) AND bldg_seq (building order)
    # for activities saved before these fields existed.
    from collections import defaultdict as _dd
    needs_save = False
    # ---- per-building activity seq_no ----
    wbs_counters = _dd(int)
    used = _dd(set)
    for a in data.get("activities", []):
        s = a.get("seq_no")
        if isinstance(s, int) and s > 0:
            used[a.get("wbs","")].add(s)
    for a in data.get("activities", []):
        if not isinstance(a.get("seq_no"), int) or a.get("seq_no", 0) <= 0:
            w = a.get("wbs","")
            wbs_counters[w] += 1
            while wbs_counters[w] in used[w]:
                wbs_counters[w] += 1
            a["seq_no"] = wbs_counters[w]
            used[w].add(wbs_counters[w])
            needs_save = True

    # ---- building-level bldg_seq ----
    # Each unique WBS gets a 1-indexed bldg_seq.  If activities in the same WBS
    # already have a bldg_seq, use it; otherwise compute from project-start order.
    bldg_seq_map = {}  # wbs -> bldg_seq
    # First pass: harvest any existing bldg_seq values
    for a in data.get("activities", []):
        b = a.get("wbs","")
        bs = a.get("bldg_seq")
        if isinstance(bs, int) and bs > 0:
            if b not in bldg_seq_map: bldg_seq_map[b] = bs
    # Compute next available bldg_seq per WBS using earliest start as tiebreak
    used_bldg_seq = set(bldg_seq_map.values())
    wbs_to_earliest = {}
    for a in data.get("activities", []):
        b = a.get("wbs","")
        if b in bldg_seq_map: continue
        s = a.get("start","") or "9999-99-99"
        if b not in wbs_to_earliest or s < wbs_to_earliest[b]:
            wbs_to_earliest[b] = s
    next_seq = 1
    for b, _ in sorted(wbs_to_earliest.items(), key=lambda kv: (kv[1], kv[0])):
        while next_seq in used_bldg_seq: next_seq += 1
        bldg_seq_map[b] = next_seq
        used_bldg_seq.add(next_seq)
        next_seq += 1
    # Stamp every activity
    for a in data.get("activities", []):
        b = a.get("wbs","")
        target = bldg_seq_map.get(b, 999)
        if a.get("bldg_seq") != target:
            a["bldg_seq"] = target
            needs_save = True

    if needs_save:
        save_sched_v2(project, data)

    # Sync planned_start/finish = start/finish before returning
    for a in data.get("activities", []):
        s = a.get("start","") or a.get("planned_start","")
        f = a.get("finish","") or a.get("planned_finish","")
        if s: a["start"]=s; a["planned_start"]=s
        if f: a["finish"]=f; a["planned_finish"]=f

    # Sort activities by (bldg_seq, seq_no)
    data["activities"] = sorted(data.get("activities", []),
        key=lambda a: (a.get("bldg_seq", 9999), a.get("seq_no", 9999)))

    acts = data.get("activities", [])
    from collections import defaultdict as _dd
    wbs_buckets = _dd(list)
    for a in acts:
        wbs_buckets[a.get("wbs", "") or "General"].append(a)

    subtotals = []
    for wbs in sorted(wbs_buckets.keys()):
        bucket = wbs_buckets[wbs]
        # Aggregations
        total_bc = sum(a.get("builder_cost", 0) for a in bucket)
        total_cp = sum(a.get("client_price", 0) for a in bucket)
        # Weighted % complete by builder_cost
        if total_bc > 0:
            wpct = sum(a.get("builder_cost", 0) * a.get("pct_complete", 0) for a in bucket) / total_bc
        else:
            wpct = sum(a.get("pct_complete", 0) for a in bucket) / len(bucket) if bucket else 0
        # Earliest start / Latest finish
        starts  = [a.get("start", "")  for a in bucket if a.get("start")]
        finishes = [a.get("finish", "") for a in bucket if a.get("finish")]
        # Sums of duration
        total_dur = sum(a.get("duration", 0) for a in bucket)
        # Status counts
        completed = sum(1 for a in bucket if a.get("status") in ("Complete","Completed"))
        in_prog   = sum(1 for a in bucket if a.get("status") == "In Progress")
        subtotals.append({
            "wbs":           wbs,
            "activity_count": len(bucket),
            "duration_sum":   total_dur,
            "earliest_start": min(starts) if starts else "",
            "latest_finish":  max(finishes) if finishes else "",
            "pct_complete":   round(wpct, 2),
            "builder_cost":   round(total_bc, 2),
            "client_price":   round(total_cp, 2),
            "profit":         round(total_cp - total_bc, 2),
            "margin_pct":     round((total_cp - total_bc) / total_cp * 100, 2) if total_cp else 0,
            "completed":      completed,
            "in_progress":    in_prog,
        })
    # Grand total
    grand_bc = sum(s["builder_cost"] for s in subtotals)
    grand_cp = sum(s["client_price"] for s in subtotals)
    grand_wpct = (sum(s["builder_cost"] * s["pct_complete"] for s in subtotals) / grand_bc) if grand_bc else 0
    grand = {
        "wbs":            "GRAND TOTAL",
        "activity_count": len(acts),
        "duration_sum":   sum(s["duration_sum"] for s in subtotals),
        "earliest_start": min((s["earliest_start"] for s in subtotals if s["earliest_start"]), default=""),
        "latest_finish":  max((s["latest_finish"]  for s in subtotals if s["latest_finish"]),  default=""),
        "pct_complete":   round(grand_wpct, 2),
        "builder_cost":   round(grand_bc, 2),
        "client_price":   round(grand_cp, 2),
        "profit":         round(grand_cp - grand_bc, 2),
        "margin_pct":     round((grand_cp - grand_bc) / grand_cp * 100, 2) if grand_cp else 0,
        "completed":      sum(s["completed"] for s in subtotals),
        "in_progress":    sum(s["in_progress"] for s in subtotals),
    }

    return {"schedule": data, "baselines": bls, "active_baseline": active_bl,
            "subtotals": subtotals, "grand_total": grand}

@app.put("/projects/{project}/schedule/v2")
def put_schedule_v2(project: str, body: dict):
    """Full schedule save — used by drag-to-reorder to persist seq_no changes
    and updated predecessor/successor relationships in one atomic write."""
    data = load_sched_v2(project)
    # Accept updated activities (seq_no) and relationships from the body
    if "activities" in body:
        incoming = {a["id"]: a for a in body["activities"]}
        for act in data.get("activities", []):
            if act["id"] in incoming:
                src = incoming[act["id"]]
                # Only update fields that the reorder operation changes
                if "seq_no"   in src: act["seq_no"]   = src["seq_no"]
                if "bldg_seq" in src: act["bldg_seq"] = src["bldg_seq"]
    if "relationships" in body:
        data["relationships"] = body["relationships"]
    save_sched_v2(project, data)
    return {"status": "ok"}

@app.post("/projects/{project}/schedule/activity")
def add_activity(project: str, body: dict):
    data = load_sched_v2(project)
    aid  = data.get("next_id", len(data["activities"]) + 1)
    a = {
        "id":              f"A{str(aid).zfill(4)}",
        "p6_code":         body.get("p6_code", f"A{str(aid).zfill(4)}"),
        "name":            body.get("name", "New Activity"),
        "wbs":             body.get("wbs", ""),
        "start":           _parse_date(body.get("start", str(_date_cls.today()))),
        "finish":          _parse_date(body.get("finish", "")),
        "baseline_start":  _parse_date(body.get("start", "")),
        "baseline_finish": _parse_date(body.get("finish", "")),
        "actual_start":    "",  "actual_finish":   "",
        "duration":        int(body.get("duration", 1)),
        "pct_complete":    float(body.get("pct_complete", 0)),
        "status":          "Not Started",
        "is_milestone":    bool(body.get("is_milestone", False)),
        "critical":        bool(body.get("critical", False)),
        "notes":           body.get("notes", ""),
        "cost_code":       body.get("cost_code", ""),
        "builder_cost":    float(body.get("builder_cost", 0)),
        "client_price":    float(body.get("client_price", 0)),
        "source":          "manual",
    }
    if a["start"] and not a["finish"]:
        a["finish"] = _add_workdays(a["start"], a["duration"])
    data["activities"].append(a)
    data["next_id"] = aid + 1
    save_sched_v2(project, data)
    return {"status": "ok", "activity": a}

@app.put("/projects/{project}/schedule/activity/{act_id}")
def update_activity(project: str, act_id: str, body: dict):
    """Update a single activity. Accepts BOTH V2 field names and legacy aliases:
       - start  / planned_start          → activity.start
       - finish / planned_finish         → activity.finish
       - wbs    / building               → activity.wbs
       - predecessor_ids (list)          → creates/updates relationships in V2 store
       Runs CPM forward-pass after the edit so all successors get rescheduled.
    """
    data = load_sched_v2(project)
    alias = {
        "planned_start":  "start",
        "planned_finish": "finish",
        "building":       "wbs",
    }
    for a in data["activities"]:
        if a["id"] == act_id:
            # Track which date the user explicitly edited so we can recompute the OTHER one
            edited_start  = "start"    in body or "planned_start"  in body
            edited_finish = "finish"   in body or "planned_finish" in body
            edited_dur    = "duration" in body
            for k in ["name","start","finish","duration","pct_complete","status",
                      "actual_start","actual_finish","wbs","notes","critical",
                      "cost_code","builder_cost","client_price","p6_code","level"]:
                if k in body:
                    a[k] = body[k]
            for legacy_k, v2_k in alias.items():
                if legacy_k in body and v2_k not in body:
                    a[v2_k] = body[legacy_k]

            # FIX 2: keep start/finish/duration in sync after any edit
            # User changes finish → recompute duration (start fixed)
            # User changes start  → recompute finish (duration fixed)
            # User changes duration → recompute finish (start fixed)
            try:
                _cal = load_calendar(project)
                if edited_finish and a.get("start") and a.get("finish"):
                    a["duration"] = _wd_between(a["start"], a["finish"], _cal)
                elif edited_start and a.get("start") and a.get("duration"):
                    a["finish"] = _add_workdays_cal(a["start"], int(a["duration"]) - 1, _cal)
                elif edited_dur and a.get("start"):
                    a["finish"] = _add_workdays_cal(a["start"], int(a.get("duration",1)) - 1, _cal)
            except Exception: pass

            # Sync planned dates
            if a.get("start"):  a["planned_start"]  = a["start"]
            if a.get("finish"): a["planned_finish"] = a["finish"]
            # Auto status from actuals
            if a.get("actual_finish"): a["status"] = "Complete"; a["pct_complete"] = 100
            elif a.get("actual_start") and a.get("status") not in ("Complete","Completed"):
                a["status"] = "In Progress"

            # predecessor_ids batch update (legacy path)
            if "predecessor_ids" in body and isinstance(body["predecessor_ids"], list):
                rels = data.setdefault("relationships", [])
                rels = [r for r in rels if r.get("succ_id") != act_id]
                pred_details = body.get("predecessor_details") or []
                detail_map = {d.get("pred_id"): d for d in pred_details if isinstance(d, dict)}
                for pid in body["predecessor_ids"]:
                    if not pid: continue
                    det = detail_map.get(pid, {})
                    rel_type = str(det.get("rel_type", "FS")).replace("PR_", "")
                    rels.append({
                        "id":       len(rels) + 1,
                        "pred_id":  pid,
                        "succ_id":  act_id,
                        "type":     rel_type or "FS",
                        "lag_days": int(det.get("lag_days", 0) or 0),
                    })
                data["relationships"] = rels

            # FIX 3: cascade dates to successors via CPM forward pass (calendar-aware).
            # Pin the user-edited activity so its dates aren't overridden by its preds —
            # only its successors re-flow. This is the P6 "user override" behavior.
            cpm_changes = _cpm_forward_pass(data, load_calendar(project), pinned_id=act_id)
            save_sched_v2(project, data)
            return {"status": "ok", "activity": a, "cpm_changes": cpm_changes,
                    "message": f"Activity {act_id} updated"+(f" • {cpm_changes} successor(s) rescheduled" if cpm_changes else "")}
    raise HTTPException(404, f"Activity {act_id} not found.")

@app.delete("/projects/{project}/schedule/activity/{act_id}")
def delete_activity(project: str, act_id: str):
    data = load_sched_v2(project)
    data["activities"] = [a for a in data["activities"] if a["id"] != act_id]
    data["relationships"] = [r for r in data.get("relationships", [])
                               if r.get("pred_id") != act_id and r.get("succ_id") != act_id]
    save_sched_v2(project, data)
    return {"status": "ok"}

@app.post("/projects/{project}/schedule/relationship")
def add_relationship(project: str, body: dict):
    data = load_sched_v2(project)
    # Generate next unique id (avoid collisions if some rels were deleted)
    existing_ids = [r.get("id",0) for r in data.get("relationships", [])]
    new_id = (max(existing_ids) if existing_ids else 0) + 1
    rel  = {"id": new_id,
            "pred_id": body["pred_id"], "succ_id": body["succ_id"],
            "type": body.get("type", "FS"), "lag_days": int(body.get("lag_days", 0))}
    data.setdefault("relationships", []).append(rel)
    # FIX 3: cascade dates to the new successor (calendar-aware)
    cpm_changes = _cpm_forward_pass(data, load_calendar(project))
    save_sched_v2(project, data)
    return {"status": "ok", "relationship": rel, "cpm_changes": cpm_changes}

@app.put("/projects/{project}/schedule/relationship/{rel_id}")
def update_relationship(project: str, rel_id: int, body: dict):
    """Update an existing relationship's type and/or lag (P6-style live edit).
    Runs CPM after the change so successor dates update."""
    data = load_sched_v2(project)
    for r in data.get("relationships", []):
        if r["id"] == rel_id:
            if "type" in body:     r["type"]     = body["type"]
            if "lag_days" in body: r["lag_days"] = int(body["lag_days"])
            cpm_changes = _cpm_forward_pass(data, load_calendar(project))
            save_sched_v2(project, data)
            return {"status": "ok", "relationship": r, "cpm_changes": cpm_changes}
    raise HTTPException(404, f"Relationship {rel_id} not found")


@app.post("/projects/{project}/schedule/building/swap-seq")
def swap_building_seq(project: str, body: dict):
    """Reorder buildings by changing their bldg_seq. Swaps with whoever currently
    has that bldg_seq. Body: {wbs: "Bldg 5", new_seq: 1}"""
    wbs_target = body.get("wbs", "")
    new_seq = int(body.get("new_seq", 0))
    if not wbs_target or new_seq < 1:
        raise HTTPException(400, "wbs and new_seq (>=1) required")
    data = load_sched_v2(project)
    acts = data.get("activities", [])
    if not acts:
        raise HTTPException(404, "No activities in schedule")
    # Current bldg_seq for target wbs
    cur_seq = next((a.get("bldg_seq", 0) for a in acts if a.get("wbs","") == wbs_target), 0)
    if cur_seq == 0:
        raise HTTPException(404, f"Building '{wbs_target}' not found")
    # Find any WBS that currently holds new_seq
    other_wbs = next((a.get("wbs","") for a in acts if a.get("bldg_seq") == new_seq and a.get("wbs","") != wbs_target), None)
    # Apply
    for a in acts:
        w = a.get("wbs","")
        if w == wbs_target:           a["bldg_seq"] = new_seq
        elif other_wbs and w == other_wbs: a["bldg_seq"] = cur_seq
    save_sched_v2(project, data)
    return {"status": "ok", "swapped": bool(other_wbs),
            "target": {"wbs": wbs_target, "bldg_seq": new_seq},
            "other":  {"wbs": other_wbs,  "bldg_seq": cur_seq} if other_wbs else None}


@app.post("/projects/{project}/schedule/activity/{act_id}/swap-seq")
def swap_activity_seq(project: str, act_id: str, body: dict):
    """Set a new S.NO for an activity by swapping with whoever currently has that seq_no
    within the same WBS (building). P6-style reorder by sequence number."""
    new_seq = int(body.get("new_seq", 0))
    if new_seq < 1:
        raise HTTPException(400, "new_seq must be >= 1")
    data = load_sched_v2(project)
    acts = data.get("activities", [])
    me = next((a for a in acts if a["id"] == act_id), None)
    if not me:
        raise HTTPException(404, f"Activity {act_id} not found")
    my_wbs = me.get("wbs", "")
    my_old_seq = me.get("seq_no", 0)
    # Find who currently has new_seq in same WBS
    other = next((a for a in acts if a.get("wbs","") == my_wbs and a.get("seq_no", 0) == new_seq and a["id"] != act_id), None)
    if other:
        other["seq_no"] = my_old_seq
    me["seq_no"] = new_seq
    save_sched_v2(project, data)
    return {"status": "ok", "swapped": bool(other),
            "me": {"id": me["id"], "seq_no": me["seq_no"]},
            "other": {"id": other["id"], "seq_no": other["seq_no"]} if other else None}

@app.delete("/projects/{project}/schedule/relationship/{rel_id}")
def delete_relationship(project: str, rel_id: int):
    data = load_sched_v2(project)
    data["relationships"] = [r for r in data.get("relationships", []) if r["id"] != rel_id]
    save_sched_v2(project, data)
    return {"status": "ok"}

@app.delete("/projects/{project}/schedule/v2")
def clear_schedule_v2(project: str):
    save_sched_v2(project, {"activities": [], "relationships": [], "next_id": 1000})
    return {"status": "ok", "message": "Schedule cleared"}

# ═══════════════════════════════════════════════════════════════════════════════
# MATERIAL AUTOMATION ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════
import uuid as _uuid
import tempfile as _tempfile

# ── Material Automation imports (graceful fallback) ────────────────────────
_MA_OK = False
_scan_drawings = None
_apply_qa_answers = None
_run_recipe_from_session = None
_build_backup_excel = None

try:
    from drawing_reader import scan_drawings as _scan_drawings
    from drawing_reader import apply_qa_answers as _apply_qa_answers
    _MA_OK = True
except Exception as _dr_err:
    print(f"[WARNING] drawing_reader not available: {_dr_err}")

try:
    from recipe_engine import run_recipe_from_session as _run_recipe_from_session
    from recipe_engine import build_backup_excel as _build_backup_excel
except Exception as _re_err:
    print(f"[WARNING] recipe_engine not available: {_re_err}")
    _MA_OK = False

def scan_drawings(pdf_paths, client=None):
    if _scan_drawings:
        return _scan_drawings(pdf_paths, client=client)
    raise RuntimeError("drawing_reader.py not available on server")

def apply_qa_answers(session, answers):
    if _apply_qa_answers:
        return _apply_qa_answers(session, answers)
    return session

def run_recipe_from_session(session):
    if _run_recipe_from_session:
        return _run_recipe_from_session(session)
    return [], []

def build_backup_excel(lines, path, project_name=""):
    if _build_backup_excel:
        return _build_backup_excel(lines, path, project_name=project_name)
    import openpyxl as _xl
    wb = _xl.Workbook(); wb.save(path)

# In-memory session store (keyed by session_id)
_takeoff_sessions: dict = {}

def _sessions_dir() -> Path:
    import tempfile as _tmp_sess
    d = Path(_tmp_sess.gettempdir()) / "matinv_sessions"
    d.mkdir(parents=True, exist_ok=True)
    return d

def _save_session(sid: str):
    try:
        import json
        s = {k: v for k, v in _takeoff_sessions.get(sid, {}).items()
             if k not in ("pdf_paths",)}
        (_sessions_dir() / f"{sid}.json").write_text(json.dumps(s, default=str))
    except Exception:
        pass

def _load_session(sid: str) -> dict | None:
    try:
        import json
        p = _sessions_dir() / f"{sid}.json"
        if p.exists():
            s = json.loads(p.read_text())
            _takeoff_sessions[sid] = s
            return s
    except Exception:
        pass
    return None


@app.post("/projects/{project}/takeoff/start")
async def takeoff_start(
    project: str,
    files: list[UploadFile] = File(...),
    building_type: str  = Form("multifamily"),
    building_count: int = Form(1),
):
    """Step 1 - Upload drawings, kick off background scan, return session_id immediately."""
    # Note: we proceed even if _MA_OK is False — scan_drawings will report
    # the specific error in log_lines so the user sees it clearly in the UI.
    _db.create_project(project)  # ensure project exists in DB

    # Save uploaded PDFs to temp dir
    tmp_dir = Path(_tempfile.mkdtemp(prefix="matinv_"))
    pdf_paths = []
    for f in files:
        fname = (f.filename or "upload.pdf")
        if not fname.lower().endswith(".pdf"):
            raise HTTPException(400, f"Only PDF files accepted, got: {fname}")
        dest = tmp_dir / fname
        dest.write_bytes(await f.read())
        pdf_paths.append(str(dest))

    # Create session immediately so client can start polling
    sid = str(_uuid.uuid4())[:8]
    _takeoff_sessions[sid] = {
        "status":           "scanning",
        "building_type":    building_type,
        "building_count":   building_count,
        "project":          project,
        "tmp_dir":          str(tmp_dir),
        "pdf_paths":        pdf_paths,
        "pages_scanned":    0,
        "schedules_found":  0,
        "stud_notes_found": False,
        "scales_detected":  [],
        "questions":        [],
        "log_lines":        [{"type":"info","icon":"📤","msg":f"Saved {len(pdf_paths)} drawing(s). Scanning now..."}],
        "lines":            None,
        "xlsx_path":        None,
        "error":            None,
    }

    # Run scan in background thread (non-blocking so we return session_id fast)
    import asyncio
    loop = asyncio.get_event_loop()

    def _bg_scan():
        sess = _takeoff_sessions[sid]
        try:
            client = None
            try:
                if ANTHROPIC_API_KEY:
                    import anthropic as _ant
                    client = _ant.Anthropic(api_key=ANTHROPIC_API_KEY)
            except Exception:
                pass
            result = scan_drawings(pdf_paths, client=client)
            sess.update(result)
            sess["status"] = "questions" if result.get("questions") else "ready"
            _save_session(sid)
        except Exception as exc:
            sess["status"] = "error"
            sess["error"]  = str(exc)
            sess["log_lines"].append({"type":"err","icon":"✗","msg":f"Scan failed: {exc}"})
            _save_session(sid)

    loop.run_in_executor(None, _bg_scan)

    return {"session_id": sid, "status": "scanning",
            "message": f"Scanning {len(pdf_paths)} drawing(s) in background"}


@app.get("/projects/{project}/takeoff/status/{session_id}")
def takeoff_status(project: str, session_id: str):
    """Poll endpoint — loads from disk if server was restarted."""
    sess = _takeoff_sessions.get(session_id) or _load_session(session_id)
    if not sess:
        raise HTTPException(404, "Session not found — please re-upload drawings")
    resp = {
        "session_id":       session_id,
        "status":           sess.get("status", "scanning"),
        "pages_scanned":    sess.get("pages_scanned", 0),
        "schedules_found":  sess.get("schedules_found", 0),
        "stud_notes_found": sess.get("stud_notes_found", False),
        "scales_detected":  sess.get("scales_detected", []),
        "questions":        sess.get("questions", []),
        "log_lines":        sess.get("log_lines", []),
        "error":            sess.get("error"),
    }
    if sess.get("status") == "complete":
        resp["lines"]  = sess.get("lines", [])
        resp["result"] = sess.get("result", {})
    return resp


@app.post("/projects/{project}/takeoff/compute")
async def takeoff_compute(project: str, body: dict):
    """Apply Q&A answers, compute in background, return immediately.
    Client polls /takeoff/status — lines returned when status=complete."""
    sid     = body.get("session_id")
    answers = body.get("answers", {})

    sess = _takeoff_sessions.get(sid) or _load_session(sid)
    if not sess:
        raise HTTPException(404, "Session not found — please re-upload drawings")

    sess = apply_qa_answers(sess, answers)
    sess["status"] = "computing"
    sess["log_lines"].append({"type":"info","icon":"⚙","msg":"Running framing recipe..."})
    _takeoff_sessions[sid] = sess
    _save_session(sid)

    import asyncio
    loop = asyncio.get_event_loop()

    def _bg_compute():
        s = _takeoff_sessions.get(sid, {})
        try:
            lines, log_lines = run_recipe_from_session(s)
            xlsx_path = UPLOAD_DIR / f"material_automation_{sid}.xlsx"
            build_backup_excel(lines, str(xlsx_path), project_name=s.get("project", project))
            s["xlsx_path"] = str(xlsx_path)
            s["lines"]     = [_line_to_dict(l) for l in lines]
            s["log_lines"].extend(log_lines)
            from collections import Counter
            total_lf = sum(l.total_lf for l in lines if l.uom in ("LF", "Pcs"))
            s["result"] = {
                "total_lines":  len(lines),
                "total_lf":     round(total_lf, 0),
                "total_panels": sum(l.total_pcs for l in lines if l.item_type == "Panels"),
                "by_level":     dict(Counter(l.level for l in lines)),
            }
            s["log_lines"].append({"type":"ok","icon":"✓","msg":f"Done — {len(lines)} lines generated"})
            s["status"] = "complete"
            _save_session(sid)
        except Exception as exc:
            s["status"] = "error"
            s["error"]  = str(exc)
            s["log_lines"].append({"type":"err","icon":"✗","msg":f"Compute failed: {exc}"})
            _save_session(sid)

    loop.run_in_executor(None, _bg_compute)
    return {"status": "computing", "session_id": sid}


@app.get("/projects/{project}/takeoff/download-excel")
def takeoff_download_excel(project: str, session_id: str):
    """Download the generated material list Excel."""
    sess = _takeoff_sessions.get(session_id) or _load_session(session_id)
    if not sess or not sess.get("xlsx_path"):
        raise HTTPException(404, "No Excel generated — run compute first.")
    path = Path(sess["xlsx_path"])
    if not path.exists():
        raise HTTPException(404, "Excel file not found on disk.")
    return FileResponse(str(path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{project}_material_list.xlsx")


def _line_to_dict(line) -> dict:
    """Serialize a RecipeLine dataclass to a JSON-safe dict."""
    return {
        "level":       line.level,
        "location":    line.location,
        "item_type":   line.item_type,
        "size_d1":     line.size_d1,
        "size_d2":     line.size_d2,
        "stock_len":   line.stock_len,
        "uom":         line.uom,
        "spec":        line.spec,
        "remarks":     line.remarks,
        "actuals_lf":  round(line.actuals_lf, 1),
        "actuals_pcs": line.actuals_pcs,
        "wastage_lf":  round(line.wastage_lf, 1),
        "wastage_pcs": line.wastage_pcs,
        "total_lf":    round(line.total_lf, 1),
        "total_pcs":   line.total_pcs,
    }
