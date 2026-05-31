"""
recipe_engine.py — MATINV Phase 2: Framing Recipe Engine
=========================================================
Takes wall-run LF measurements (from takeoff_parser.py) and produces the
complete lumber material list in Backup Levelwise format — exactly matching
the Willow Way P-0348 Overall Backup structure.

RULES (locked from Silver City structural notes + user confirmation):
───────────────────────────────────────────────────────────────────────
Stud heights (L1–L4 Exterior 12' floor):
  Stud = 144" − 20" (truss) − 3" (DTP) − 1.5" (BP) = 119.5" ≈ 9'11½"
  → order as 10' studs (120" stock, trim to 119.5")

Stud heights (L1–L4 Interior 9' ceiling):
  Stud = 108" − 20" (truss) − 3" (DTP) − 1.5" (BP) = 83.5" ≈ 6'11½"
  → order as 8' studs (96" stock, trim to 83.5") — or 7' precut

Stud heights (L5 Exterior + Interior 9' ceiling):
  Same as interior above → 83.5"

Wall recipe per LF of wall run:
  1 × Sill Plate    (PT, 1 LF per LF)
  2 × Top Plate     (SYP#2, 2 LF per LF = double top plate)
  2 × Blocking      (SYP#2, 2 LF per LF = mid-height blocking)
  Studs             (SYP#2, LF/run ÷ OC_in_inches × 12 × stud_height_in_feet)

Sheathing:
  Wall sheathing area = run_LF × wall_height_ft
  Panels = area ÷ 32 SF/sheet

Headers:
  From S0.06 Header/Beam Schedule — looked up by mark

FRT rule:
  All EXTERIOR walls → PT lumber for sill plate, FRT for studs/plates
  (per Silver City structural notes: all 2-hr exterior LBW = FRT)

Wastage: 5% on all items (matching Willow Way V4 format)

Building count multiplier:
  Silver City = 1 building (5-story student housing, no mirror repeat)
  Willow Way  = 2 (North + South identical)
"""

import math
import re
from dataclasses import dataclass, field
from typing import Optional
from collections import defaultdict


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

WASTAGE_PCT = 0.05   # 5% wastage
SF_PER_PANEL = 32    # Standard 4×8 sheet

# Stud height lookup (inches) keyed by (level, wall_type)
# wall_type: 'exterior' | 'interior'
STUD_HEIGHT_IN = {
    # L1–L4 exterior walls have 12' floor-to-floor
    ("L1", "exterior"): 119.5,
    ("L2", "exterior"): 119.5,
    ("L3", "exterior"): 119.5,
    ("L4", "exterior"): 119.5,
    # L1–L4 interior/demising/corridor = 9' ceiling
    ("L1", "interior"): 83.5,
    ("L2", "interior"): 83.5,
    ("L3", "interior"): 83.5,
    ("L4", "interior"): 83.5,
    # L5 all walls = 9' ceiling
    ("L5", "exterior"): 83.5,
    ("L5", "interior"): 83.5,
}
DEFAULT_STUD_HEIGHT_IN = 83.5  # fallback

# Stock length to order (nearest standard length ≥ stud height)
def stud_stock_length_ft(stud_height_in: float) -> int:
    """Return the stock length in feet to order for a given stud height."""
    stud_ft = stud_height_in / 12
    for stock in [7, 8, 9, 10, 12, 14, 16]:
        if stock >= stud_ft:
            return stock
    return 16

# OC spacing defaults by wall category + level load
# Based on Silver City structural general notes (Image 2):
# Exterior load-bearing: 2x6 @ 12" OC (supporting ≥2 floors)
# Interior load-bearing: (2)2x4 @ 8" OC or 2x6 @ 16" OC
# This maps item name patterns to (stud_size, oc_inches)
OC_DEFAULTS = {
    "exterior_wall":  ("2x6", 12),
    "corridor_wall":  ("2x6", 12),
    "demising_wall":  ("2x4", 16),
    "interior_wall":  ("2x4", 16),
    "shear_wall":     ("2x6", 16),
    "stair_wall":     ("2x4", 16),
    "elevator_wall":  ("2x4", 16),
}

# Header/Beam Schedule from S0.06 (Silver City)
# Format: mark → {lumber_size, material, ply_count, flitch_plate, stock_length_ft}
HEADER_SCHEDULE = {
    "H2-08": {"lumber_size": "2x8",           "material": "SYP#2", "ply_count": 2,  "flitch_plate": "1/2\" PLY", "stock_ft": 8},
    "H2-10": {"lumber_size": "2x10",          "material": "SYP#2", "ply_count": 2,  "flitch_plate": "1/2\" PLY", "stock_ft": 10},
    "H2-12": {"lumber_size": "2x12",          "material": "SYP#2", "ply_count": 2,  "flitch_plate": "1/2\" PLY", "stock_ft": 12},
    "H3-08": {"lumber_size": "2x8",           "material": "SYP#2", "ply_count": 3,  "flitch_plate": "1/2\" PLY", "stock_ft": 8},
    "H3-10": {"lumber_size": "2x10",          "material": "SYP#2", "ply_count": 3,  "flitch_plate": "1/2\" PLY", "stock_ft": 10},
    "H3-12": {"lumber_size": "2x12",          "material": "SYP#2", "ply_count": 3,  "flitch_plate": "1/2\" PLY", "stock_ft": 12},
    "H4-08": {"lumber_size": "3-1/2\"x9-1/4\"",  "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 8},
    "H4-10": {"lumber_size": "5-1/4\"x9-1/4\"",  "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 10},
    "H4-12": {"lumber_size": "3-1/2\"x11-7/8\"", "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 12},
    "H4-14": {"lumber_size": "3-1/2\"x14\"",  "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 14},
    "H4-16": {"lumber_size": "3-1/2\"x14\"",  "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 16},
    "H4-18": {"lumber_size": "3-1/2\"x16\"",  "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 18},
    "H4-20": {"lumber_size": "3/4\"x20\"",    "material": "LVL",   "ply_count": 2,  "flitch_plate": None, "stock_ft": 20},
    "H5-10": {"lumber_size": "5-1/4\"x9-1/4\"",  "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 10},
    "H5-12": {"lumber_size": "5-1/4\"x11-7/8\"", "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 12},
    "H6-10": {"lumber_size": "5-1/4\"x9-1/4\"",  "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 10},
    "H6-12": {"lumber_size": "5-1/4\"x11-7/8\"", "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 12},
    "H6-14": {"lumber_size": "5-1/4\"x14\"",  "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 14},
    "H6-16": {"lumber_size": "5-1/4\"x14\"",  "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 16},
    "H6-18": {"lumber_size": "5-1/4\"x14\"",  "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 18},
    "H6-20": {"lumber_size": "5-1/4\"x16\"",  "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 20},
    "H8-18": {"lumber_size": "7\"x14\"",      "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 18},
    "H8-20": {"lumber_size": "7\"x16\"",      "material": "PSL",   "ply_count": 1,  "flitch_plate": None, "stock_ft": 20},
    "H6-20L":{"lumber_size": "1-3/4\"x20\"",  "material": "LVL",   "ply_count": 4,  "flitch_plate": None, "stock_ft": 20},
}


# ─────────────────────────────────────────────────────────────────────────────
# Data classes
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class RecipeLine:
    """One line in the Backup Levelwise output — mirrors Overall Backup - V4 columns."""
    building_type: str        # "Silver City" / "North & South" etc.
    level:         str        # L1, L2, L3, L4, L5, Roof
    mark:          str        # header mark, shear wall mark, or ""
    location:      str        # Exterior Wall, Interior Wall, Corridor Wall, etc.
    item_type:     str        # Lumber / EWP / Panels / Each
    size_d1:       str        # first dimension (e.g. "2")
    size_d2:       str        # second dimension (e.g. "6")
    stock_len:     int        # standard stock length in ft (or SF for panels)
    uom:           str        # LF / Pcs / SF
    spec:          str        # SYP#2 / PT / FRT / PSL / LVL / 7/16\" OSB Zip
    compile_code:  str        # material code key (e.g. "2X6-9-SYP#2")
    takeoff_lf:    float      # measured LF input
    oc_in:         int        # OC spacing in inches (0 if not applicable)
    actuals_lf:    float      # computed LF/SF
    actuals_pcs:   int        # computed pieces
    remarks:       str        # Sill Plate / Double Top Plate / Wall Studs / etc.
    building_count:int = 1    # multiplier (Silver City = 1)

    @property
    def wastage_lf(self) -> float:
        return round(self.actuals_lf * WASTAGE_PCT, 1)

    @property
    def wastage_pcs(self) -> int:
        return math.ceil(self.actuals_pcs * WASTAGE_PCT)

    @property
    def total_lf(self) -> float:
        return round((self.actuals_lf + self.wastage_lf) * self.building_count, 1)

    @property
    def total_pcs(self) -> int:
        return math.ceil((self.actuals_pcs + self.wastage_pcs) * self.building_count)


# ─────────────────────────────────────────────────────────────────────────────
# Wall type classifier
# ─────────────────────────────────────────────────────────────────────────────

def _is_exterior(category: str, raw_name: str) -> bool:
    """Determine if a wall is exterior (→ FRT) or interior."""
    ext_cats = {"exterior_wall", "shear_wall", "stair_wall", "elevator_wall"}
    if category in ext_cats:
        return True
    name_up = raw_name.upper()
    if any(k in name_up for k in ["EXTERIOR", "EXT", "SW", "SHEAR"]):
        return True
    return False


def _wall_height_ft(level: str, is_exterior: bool) -> float:
    """Return wall height in feet for sheathing area calculation."""
    wtype = "exterior" if is_exterior else "interior"
    stud_in = STUD_HEIGHT_IN.get((level, wtype), DEFAULT_STUD_HEIGHT_IN)
    # Full wall height = stud + bottom plate + double top plate
    return (stud_in + 1.5 + 3.0) / 12.0


def _stud_height_ft(level: str, is_exterior: bool) -> float:
    wtype = "exterior" if is_exterior else "interior"
    return STUD_HEIGHT_IN.get((level, wtype), DEFAULT_STUD_HEIGHT_IN) / 12.0


# ─────────────────────────────────────────────────────────────────────────────
# Core recipe: wall run → lumber lines
# ─────────────────────────────────────────────────────────────────────────────

def _studs_per_lf(oc_in: int) -> float:
    """Number of studs per LF of wall run at given OC spacing."""
    return 12.0 / oc_in  # e.g. 16" OC → 0.75 studs/LF


def wall_recipe(
    building_type: str,
    level: str,
    location: str,
    category: str,
    raw_name: str,
    run_lf: float,
    stud_size: Optional[str],
    oc_in: Optional[int],
    building_count: int = 1,
) -> list[RecipeLine]:
    """
    Expand one wall run into all lumber recipe lines.

    Returns a list of RecipeLine objects:
      - Sill Plate (PT)
      - Double Top Plate (SYP#2 or FRT)
      - Wall Blocking (SYP#2 or FRT)
      - Wall Studs (SYP#2 or FRT, in stock lengths)
      - Wall Sheathing (panels) if exterior
    """
    if run_lf <= 0:
        return []

    lines: list[RecipeLine] = []
    is_ext = _is_exterior(category, raw_name)
    stud_spec  = "FRT" if is_ext else "SYP#2"
    sill_spec  = "PT"

    # Resolve stud size and OC from item name or defaults
    if not stud_size or not oc_in:
        default_size, default_oc = OC_DEFAULTS.get(category, ("2x4", 16))
        stud_size = stud_size or default_size
        oc_in     = oc_in     or default_oc

    # Parse stud size into d1, d2
    m = re.match(r"(\d+)[xX](\d+)", stud_size)
    d1 = m.group(1) if m else "2"
    d2 = m.group(2) if m else "4"

    stud_ht_ft    = _stud_height_ft(level, is_ext)
    stud_stock_ft = stud_stock_length_ft(stud_ht_ft * 12)
    wall_ht_ft    = _wall_height_ft(level, is_ext)

    run = round(run_lf, 1)

    def make(location_tag, item_type, d1_, d2_, stock, uom, spec, code, lf, oc, pcs, remarks):
        return RecipeLine(
            building_type=building_type, level=level, mark="",
            location=location_tag, item_type=item_type,
            size_d1=str(d1_), size_d2=str(d2_), stock_len=stock, uom=uom, spec=spec,
            compile_code=code, takeoff_lf=run_lf, oc_in=oc,
            actuals_lf=round(lf, 1), actuals_pcs=int(pcs), remarks=remarks,
            building_count=building_count,
        )

    # ── 1. Sill Plate (1× run LF, PT) ────────────────────────────────────────
    lines.append(make(
        location, "Lumber", d1, d2, 1, "LF", sill_spec,
        f"{d1}X{d2}-1-PT", run, 0, run, "Sill Plate"
    ))

    # ── 2. Double Top Plate (2× run LF) ──────────────────────────────────────
    dtp_lf = run * 2
    lines.append(make(
        location, "Lumber", d1, d2, 1, "LF", stud_spec,
        f"{d1}X{d2}-1-{stud_spec}", dtp_lf, 0, dtp_lf, "Double Top Plate"
    ))

    # ── 3. Wall Blocking (2× run LF) ─────────────────────────────────────────
    blk_lf = run * 2
    lines.append(make(
        location, "Lumber", d1, d2, 1, "LF", stud_spec,
        f"{d1}X{d2}-1-{stud_spec}", blk_lf, 0, blk_lf, "Wall Blocking"
    ))

    # ── 4. Wall Studs ─────────────────────────────────────────────────────────
    n_studs    = math.ceil(_studs_per_lf(oc_in) * run)
    stud_lf    = n_studs * stud_ht_ft
    lines.append(make(
        location, "Lumber", d1, d2, stud_stock_ft, "Pcs", stud_spec,
        f"{d1}X{d2}-{stud_stock_ft}-{stud_spec}",
        round(stud_lf, 1), oc_in, n_studs, "Wall Studs"
    ))

    # ── 5. Exterior Wall Sheathing (panels, only for exterior walls) ──────────
    if is_ext:
        sheath_sf   = round(run * wall_ht_ft, 1)
        sheath_pcs  = math.ceil(sheath_sf / SF_PER_PANEL)
        sheath_spec = "7/16\" OSB Zip Sheath"
        lines.append(RecipeLine(
            building_type=building_type, level=level, mark="",
            location=f"{location} Sheathing",
            item_type="Panels", size_d1="4", size_d2="8",
            stock_len=32, uom="SF", spec=sheath_spec,
            compile_code="4X8-12-7/16\" OSB - Zip Sheath",
            takeoff_lf=run_lf, oc_in=0,
            actuals_lf=sheath_sf, actuals_pcs=sheath_pcs,
            remarks="Exterior Wall Sheathing",
            building_count=building_count,
        ))

    return lines


# ─────────────────────────────────────────────────────────────────────────────
# Header recipe
# ─────────────────────────────────────────────────────────────────────────────

def header_recipe(
    building_type: str,
    level: str,
    mark: str,
    count: int,
    building_count: int = 1,
) -> list[RecipeLine]:
    """Expand a header mark + count into lumber lines (EWP or Lumber + flitch)."""
    spec = HEADER_SCHEDULE.get(mark)
    if not spec or count <= 0:
        return []

    lines = []
    ls = spec["lumber_size"]
    mat = spec["material"]
    ply = spec["ply_count"] or 1
    stock = spec["stock_ft"]
    flitch = spec["flitch_plate"]

    # Parse size
    m = re.match(r"(\d+)[xX](\d+)", ls)
    if m:
        d1, d2, item_type = m.group(1), m.group(2), "Lumber"
    else:
        # EWP like "3-1/2\"x11-7/8\""
        parts = ls.split("x") if "x" in ls else ls.split("X")
        d1 = parts[0].strip().replace('"', '')
        d2 = parts[1].strip().replace('"', '') if len(parts) > 1 else "?"
        item_type = "EWP"

    total_lf    = count * stock * ply
    total_pcs   = count * ply
    code        = f"{d1}X{d2}-{stock}-{mat}"

    lines.append(RecipeLine(
        building_type=building_type, level=level, mark=mark,
        location="Headers", item_type=item_type,
        size_d1=d1, size_d2=d2, stock_len=stock, uom="Pcs", spec=mat,
        compile_code=code, takeoff_lf=0, oc_in=0,
        actuals_lf=round(total_lf, 1), actuals_pcs=total_pcs,
        remarks=f"Header {mark}", building_count=building_count,
    ))

    # Flitch plate
    if flitch:
        fp_sf  = count * stock * 1.5  # 1.5 SF per LF of header (approx height)
        fp_pcs = math.ceil(fp_sf / SF_PER_PANEL)
        lines.append(RecipeLine(
            building_type=building_type, level=level, mark=f"{mark}-FP",
            location="Headers", item_type="Panels",
            size_d1="4", size_d2="8", stock_len=stock, uom="SF", spec="1/2\" Ply",
            compile_code=f"4X8-{stock}-1/2\" Ply",
            takeoff_lf=0, oc_in=0,
            actuals_lf=fp_sf, actuals_pcs=fp_pcs,
            remarks=f"Flitch Plate {mark}", building_count=building_count,
        ))

    return lines


# ─────────────────────────────────────────────────────────────────────────────
# Main recipe runner
# ─────────────────────────────────────────────────────────────────────────────

def run_recipe(
    takeoff_items: list,   # list of TakeoffItem from takeoff_parser
    building_type: str = "Silver City",
    building_count: int = 1,
    scale_ft_per_pt: float = None,
) -> list[RecipeLine]:
    """
    Process all takeoff items and return a complete list of RecipeLines.

    takeoff_items: output of takeoff_parser.parse_takeoff_zip()
    scale_ft_per_pt: if not None, apply to raw_total_pt to get LF
                     (only needed if takeoff_parser didn't apply scale)
    """
    all_lines: list[RecipeLine] = []
    header_counts: dict[tuple, dict[str, int]] = defaultdict(lambda: defaultdict(int))

    for item in takeoff_items:
        # Skip non-wood / concrete / items with no measurement
        if item.is_skip:
            continue
        if item.item_class not in ("Linear", "Area"):
            continue

        # Resolve LF
        lf = item.total_lf
        if lf is None and scale_ft_per_pt is not None:
            lf = item.raw_total_pt * scale_ft_per_pt
        if not lf or lf <= 0:
            continue

        level = item.level
        cat   = item.category

        # Skip non-wall categories (blocking/headers/sheathing handled separately)
        wall_cats = {
            "exterior_wall", "interior_wall", "demising_wall",
            "corridor_wall", "shear_wall", "stair_wall", "elevator_wall"
        }
        if cat not in wall_cats:
            continue

        # Friendly location name
        loc_map = {
            "exterior_wall":  "Exterior Wall",
            "interior_wall":  "Interior Wall",
            "demising_wall":  "Demising Wall",
            "corridor_wall":  "Corridor Wall",
            "shear_wall":     "Shear Wall",
            "stair_wall":     "Stair Wall",
            "elevator_wall":  "Elevator Wall",
        }
        location = loc_map.get(cat, cat.replace("_", " ").title())

        # Collect header marks from item name for later processing
        if item.sw_mark:
            header_counts[(level, location)][item.sw_mark] += 1

        lines = wall_recipe(
            building_type=building_type,
            level=level,
            location=location,
            category=cat,
            raw_name=item.raw_name,
            run_lf=lf,
            stud_size=item.stud_size,
            oc_in=item.oc_spacing,
            building_count=building_count,
        )
        all_lines.extend(lines)

    return all_lines


# ─────────────────────────────────────────────────────────────────────────────
# Aggregator — consolidate lines to Final List format
# ─────────────────────────────────────────────────────────────────────────────

def aggregate_to_final_list(lines: list[RecipeLine]) -> list[dict]:
    """
    Roll up RecipeLines into a Final List (one row per compile_code + stock_len + spec).
    Matches the 'Final List Updated-V4' sheet format.
    """
    totals: dict[str, dict] = {}

    for line in lines:
        key = line.compile_code
        if key not in totals:
            totals[key] = {
                "item_type":   line.item_type,
                "compile_code": key,
                "size_d1":     line.size_d1,
                "size_d2":     line.size_d2,
                "stock_len":   line.stock_len,
                "uom":         line.uom,
                "spec":        line.spec,
                "actuals_lf":  0.0,
                "actuals_pcs": 0,
                "wastage_lf":  0.0,
                "wastage_pcs": 0,
                "total_lf":    0.0,
                "total_pcs":   0,
                "remarks":     line.remarks,
            }
        t = totals[key]
        t["actuals_lf"]  += line.total_lf / line.building_count  # already scaled
        t["actuals_pcs"] += line.actuals_pcs * line.building_count
        t["wastage_lf"]  += line.wastage_lf * line.building_count
        t["wastage_pcs"] += line.wastage_pcs * line.building_count
        t["total_lf"]    += line.total_lf
        t["total_pcs"]   += line.total_pcs

    # Round
    for t in totals.values():
        t["actuals_lf"]  = round(t["actuals_lf"], 1)
        t["wastage_lf"]  = round(t["wastage_lf"], 1)
        t["total_lf"]    = round(t["total_lf"], 1)

    return sorted(totals.values(), key=lambda r: (r["item_type"], r["spec"], r["compile_code"]))


# ─────────────────────────────────────────────────────────────────────────────
# Excel output builder
# ─────────────────────────────────────────────────────────────────────────────

def build_backup_excel(
    lines: list[RecipeLine],
    output_path: str,
    project_name: str = "Silver City",
    gsf: str = "",
) -> str:
    """
    Write the Backup Levelwise Excel:
      Sheet 1: 'Overall Backup'  — per-level detail (mirrors Overall Backup - V4)
      Sheet 2: 'Final List'      — aggregated material list (mirrors Final List Updated-V4)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    THIN   = Side(border_style="thin", color="CCCCCC")
    BDR    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_F  = PatternFill("solid", start_color="1B3A5C")
    HDR_FT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    SUB_F  = PatternFill("solid", start_color="D9E1F2")
    SUB_FT = Font(bold=True, name="Arial", size=10)
    NRM_FT = Font(name="Arial", size=9)
    TITLE_FT = Font(bold=True, name="Arial", size=11)

    def hdr(ws, r, c, v, fill=HDR_F, font=HDR_FT):
        cell = ws.cell(row=r, column=c, value=v)
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BDR

    def val(ws, r, c, v, fmt=None, bold=False, fill=None, halign="center"):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=bold, name="Arial", size=9)
        cell.alignment = Alignment(horizontal=halign, vertical="center")
        cell.border = BDR
        if fmt: cell.number_format = fmt
        if fill: cell.fill = fill
        return cell

    # ── Sheet 1: Overall Backup ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Overall Backup"
    ws1.row_dimensions[1].height = 14
    ws1.row_dimensions[2].height = 30

    # Project title row
    ws1.merge_cells("A1:X1")
    t = ws1["A1"]
    t.value = f"{project_name} — Framing Backup Levelwise{' | GSF = ' + gsf if gsf else ''}"
    t.font = TITLE_FT
    t.alignment = Alignment(horizontal="left", vertical="center")

    # Column headers (matching Overall Backup - V4)
    cols = [
        "Sl No", "Building Type", "Level", "Mark", "Location",
        "Count/Unit", "×", "Size Count", "×", "Size", "×", "Size",
        "L/H", "UOM", "Specification", "Compile Code", "Takeoff LF", "OC",
        "Actuals LF/SF", "Actuals Pcs",
        "Wastage LF/SF", "Wastage Pcs",
        "Total LF/SF", "Total Pcs", "Remarks", "Bldg Count"
    ]
    for ci, col in enumerate(cols, 1):
        hdr(ws1, 2, ci, col)

    # Set column widths
    widths = [6,14,6,8,18,8,4,8,4,6,4,6,6,6,18,20,10,6,12,10,12,10,12,10,20,8]
    for ci, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    row = 3
    prev_level = None
    for sl, line in enumerate(lines, 1):
        # Level separator
        if line.level != prev_level:
            prev_level = line.level
            ws1.merge_cells(f"A{row}:Z{row}")
            sep = ws1.cell(row=row, column=1, value=f"── {line.level} ──")
            sep.fill = PatternFill("solid", start_color="E2EFD9")
            sep.font = Font(bold=True, name="Arial", size=9)
            sep.alignment = Alignment(horizontal="left")
            row += 1

        fill = PatternFill("solid", start_color="EBF3FB") if line.item_type == "Lumber" else \
               PatternFill("solid", start_color="FFF3CD") if line.item_type in ("EWP","Panels") else None

        vals = [
            sl, line.building_type, line.level, line.mark, line.location,
            1, "×", 1, "×", line.size_d1, "×", line.size_d2,
            line.stock_len, line.uom, line.spec, line.compile_code,
            round(line.takeoff_lf, 1), line.oc_in or "",
            round(line.actuals_lf, 1), line.actuals_pcs,
            round(line.wastage_lf, 1), line.wastage_pcs,
            round(line.total_lf, 1), line.total_pcs,
            line.remarks, line.building_count,
        ]
        for ci, v in enumerate(vals, 1):
            val(ws1, row, ci, v, fill=fill, halign="right" if ci > 12 else "left")
        row += 1

    # Freeze header rows
    ws1.freeze_panes = "A3"

    # ── Sheet 2: Final List ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Final List")
    ws2.merge_cells("A1:O1")
    t2 = ws2["A1"]
    t2.value = f"{project_name} — Lumber & EWP Material List | Revision 1 | Date: {__import__('datetime').date.today()}"
    t2.font = TITLE_FT

    fl_cols = ["Sl No","Material","Code","Size d1","×","Size d2","L/H","Specification",
               "Actuals LF/SF","Actuals Pcs","Wastage LF/SF","Wastage Pcs","Total LF/SF","Total Pcs"]
    for ci, col in enumerate(fl_cols, 1):
        hdr(ws2, 2, ci, col)

    fl_widths = [6,10,24,8,4,8,6,20,12,10,12,10,12,10]
    for ci, w in enumerate(fl_widths, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    final = aggregate_to_final_list(lines)
    grand_lf = 0
    for sl, item in enumerate(final, 1):
        fill = PatternFill("solid", start_color="DDEEFF") if item["item_type"] == "Lumber" else \
               PatternFill("solid", start_color="D4EDDA") if item["item_type"] == "EWP" else \
               PatternFill("solid", start_color="FFF3CD") if item["item_type"] == "Panels" else None
        r = sl + 2
        row_vals = [
            sl, item["item_type"], item["compile_code"],
            item["size_d1"], "×", item["size_d2"],
            item["stock_len"], item["spec"],
            item["total_lf"], item["total_pcs"],
            item["wastage_lf"], item["wastage_pcs"],
            item["total_lf"], item["total_pcs"],
        ]
        for ci, v in enumerate(row_vals, 1):
            val(ws2, r, ci, v, fill=fill)
        grand_lf += item["total_lf"]

    # Grand total row
    gr = len(final) + 3
    ws2.cell(row=gr, column=1, value="TOTAL").font = Font(bold=True, name="Arial")
    ws2.cell(row=gr, column=9, value=round(grand_lf, 0)).font = Font(bold=True, name="Arial")
    ws2.freeze_panes = "A3"

    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# CLI test runner
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(__import__("pathlib").Path(__file__).parent))
    from takeoff_parser import parse_takeoff_zip, SCALE_PRESETS

    zip_path = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/Takeoff.zip"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "/tmp/silver_city_backup.xlsx"

    print(f"Parsing {zip_path}...")
    # Use 1/4"=1' at 72 dpi (per scale callout on drawings)
    scale = SCALE_PRESETS["1/4\"=1'-0\""]
    items = parse_takeoff_zip(zip_path, scale_ft_per_pt=scale, wood_only=True)
    print(f"  Loaded {len(items)} wood items")

    print("Running recipe engine...")
    lines = run_recipe(
        takeoff_items=items,
        building_type="Silver City",
        building_count=1,
    )
    print(f"  Generated {len(lines)} recipe lines")

    # Summary by level
    from collections import Counter
    lvl_count = Counter(l.level for l in lines)
    for lvl in sorted(lvl_count):
        lf_total = sum(l.total_lf for l in lines if l.level == lvl and l.uom in ("LF","Pcs"))
        print(f"  {lvl}: {lvl_count[lvl]} lines, ~{lf_total:,.0f} total LF")

    print(f"\nBuilding Excel → {out_path}")
    build_backup_excel(lines, out_path, project_name="Silver City")
    print("Done ✓")


# ─────────────────────────────────────────────────────────────────────────────
# Bridge: session dict → recipe lines
# ─────────────────────────────────────────────────────────────────────────────

def run_recipe_from_session(session: dict) -> tuple[list, list]:
    """
    Convert a drawing_reader session into RecipeLines.

    This is the core bridge used by the FastAPI endpoint.
    It reads the session's stud_specs, truss_info, floor_heights,
    header_schedule, and sheet data to produce the full material list.

    Returns: (lines, log_lines)
    """
    logs: list[dict] = []
    def log(t, icon, msg):
        logs.append({"type": t, "icon": icon, "msg": msg})

    stud_specs     = session.get("stud_specs", {})
    truss_info     = session.get("truss_info", {})
    floor_heights  = session.get("floor_heights", {})
    header_sched   = session.get("header_schedule", {}) or HEADER_SCHEDULE
    building_type  = session.get("building_type", "multifamily")
    building_count = int(session.get("building_count", 1))
    project        = session.get("project", "Project")
    sheets         = session.get("sheets", [])

    truss_depth_in = truss_info.get("depth_in", 20)  # default 20" if not found
    log("info", "🏗", f"Truss depth: {truss_depth_in}\"")

    # ── Compute stud heights per level ────────────────────────────────────────
    # Get floor-to-floor height per level (from session or default)
    # Default: L1-L4 exterior = 12', all interior = 9', L5 all = 9'
    def _stud_height(level: str, is_ext: bool) -> float:
        ftf_key = level
        if ftf_key in floor_heights:
            ftf_in = float(floor_heights[ftf_key])
        else:
            # Silver City style defaults
            if is_ext and level in ("L1", "L2", "L3", "L4"):
                ftf_in = 144.0  # 12 ft
            else:
                ftf_in = 108.0  # 9 ft
        return ftf_in - truss_depth_in - 3.0 - 1.5  # - truss - DTP - BP

    # ── Build wall runs from sheet data ───────────────────────────────────────
    # Since drawings were read but no spatial measurement was done (Tier 1 = no vision measurement),
    # we need to read wall runs from the takeoff data if provided,
    # OR use the sheet scan to identify what was found and prompt for runs.
    #
    # For now: use any TakeoffItem data attached to session (from takeoff_parser if provided)
    # OR synthesize from schedule data + floor areas.
    #
    # The session may contain "takeoff_items" if the user also uploaded a Takeoff.zip
    # (future), or "wall_runs" from a spatial vision pass (Tier 2).
    # For Tier 1 (drawings only), we use the schedule data to build a skeleton.

    takeoff_items = session.get("takeoff_items", [])
    all_lines: list[RecipeLine] = []

    if takeoff_items:
        # Use measured wall runs from takeoff parser
        log("info", "📐", f"Using {len(takeoff_items)} measured wall runs")
        for item in takeoff_items:
            if item.is_skip or item.item_class != "Linear":
                continue
            lf = item.total_lf or 0
            if lf <= 0:
                continue
            wall_cats = {"exterior_wall","interior_wall","demising_wall",
                         "corridor_wall","shear_wall","stair_wall","elevator_wall"}
            if item.category not in wall_cats:
                continue
            loc_map = {
                "exterior_wall":"Exterior Wall", "interior_wall":"Interior Wall",
                "demising_wall":"Demising Wall", "corridor_wall":"Corridor Wall",
                "shear_wall":"Shear Wall", "stair_wall":"Stair Wall",
                "elevator_wall":"Elevator Wall",
            }
            location = loc_map.get(item.category, item.category.replace("_"," ").title())
            is_ext   = _is_exterior(item.category, item.raw_name)

            # Resolve stud spec from session stud_specs
            stud_size, oc_in = item.stud_size, item.oc_spacing
            if not stud_size or not oc_in:
                wtype = "exterior" if is_ext else item.category.replace("_wall","")
                wspec = stud_specs.get(wtype, stud_specs.get("exterior" if is_ext else "interior", {}))
                # Pick spec for the level's load (estimate: higher levels support fewer floors)
                lvl_num = int(item.level[1]) if item.level and item.level[1:].isdigit() else 1
                max_floors = max(wspec.keys()) if wspec else 0
                load = max(0, max_floors - (lvl_num - 1))
                spec_tuple = wspec.get(load, wspec.get(0, None))
                if spec_tuple:
                    stud_size = stud_size or spec_tuple[0]
                    oc_in     = oc_in     or spec_tuple[1]

            # Override stud height from session
            stud_ht = _stud_height(item.level, is_ext)
            # Temporarily patch STUD_HEIGHT_IN for this call
            _orig = STUD_HEIGHT_IN.copy()
            for k in list(STUD_HEIGHT_IN.keys()):
                del STUD_HEIGHT_IN[k]
            wtype_key = "exterior" if is_ext else "interior"
            STUD_HEIGHT_IN[(item.level, wtype_key)] = stud_ht

            lines = wall_recipe(
                building_type=project,
                level=item.level,
                location=location,
                category=item.category,
                raw_name=item.raw_name,
                run_lf=lf,
                stud_size=stud_size,
                oc_in=oc_in,
                building_count=building_count,
            )
            all_lines.extend(lines)
            STUD_HEIGHT_IN.update(_orig)

    else:
        # No takeoff items — generate a skeleton message and ask for wall run data
        log("warn", "⚠",
            "No measured wall runs found. "
            "Upload a PlanSwift Takeoff.zip alongside the drawings for full quantities. "
            "Showing schedule data only.")

        # At minimum, return the header schedule data as lines
        if header_sched:
            log("info", "📊", f"Header schedule loaded: {len(header_sched)} marks")
            for mark, spec in header_sched.items():
                hlines = header_recipe(
                    building_type=project,
                    level="All Levels",
                    mark=mark,
                    count=1,
                    building_count=building_count,
                )
                all_lines.extend(hlines)

    log("ok", "✓", f"Recipe complete: {len(all_lines)} lines generated")
    return all_lines, logs
